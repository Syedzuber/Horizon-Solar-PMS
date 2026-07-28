import json
import logging
import re
import uuid as _uuid
from datetime import date, timedelta, datetime
from decimal import Decimal, InvalidOperation
from urllib.parse import urlparse as _urlparse

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.db import transaction, IntegrityError
from django.db.models import Count, DecimalField, Exists, F, Max, Min, OuterRef, Prefetch, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.core.exceptions import PermissionDenied
from django.http import Http404, HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt

from .models import (
    UserProfile, Project, ProjectPhase, Task, DueDateChangeLog, ProjectFieldEditLog,
    Vendor, VendorCategory, VendorBrand,
    BOQ, BOQItem, BOQItemMaster, BOQRevision, Notification, get_standard_boq_items,
    PaymentMilestone, ProjectDocument, TaskAttachment,
    Issue, ActivityLog, Comment, log_activity,
    DeliveryChallan, DCLineItem, recalculate_dc_status, get_material_status,
    PaymentRequest, NotificationLog, SystemSettings, DesignSubmission,
    TaskDurationTemplate,
    Checklist, ChecklistItem, ChecklistTaskLink, ChecklistItemCompletion,
    Program, program_rollup_annotations, get_program_rollup,
)
from .notifications import send_notification, send_raw_email
from .forms import UserCreateForm, UserEditForm, AdminUserEditForm, ProjectCreateForm, ProjectEditForm, PostActivationFieldEditForm, TaskAddForm, VendorForm, ProgramForm, OpexSiteForm, BOQItemMasterForm, normalize_program_code
from .decorators import (
    login_required, role_required, get_user_dashboard,
    get_post_login_url, LANDING_ROLES,
)
from .permissions import (
    user_can_manage_project, project_managers, user_can_manage_program,
    user_can_view_project_boq, user_can_edit_project_boq,
    project_boq_is_group_locked,
)
from .utils import (
    attach_residential_template, calculate_due_dates, recalculate_from_task,
    get_residential_template_task_names, compute_gantt_schedule, build_gantt_view,
)
from .gantt_constants import GANTT_PHASE_DISPLAY_NAME_MAP, GANTT_TASK_DISPLAY_NAME_MAP
# Dashboard integration for the OPEX design module (Part 4.5). Context helpers only —
# they read design state and decide which single action to offer; every endpoint they
# point at re-checks authority for itself. design_views does not import views, so this
# direction of the dependency is safe.
from .design_views import (
    design_head_dashboard_counts, designer_dashboard_context, pm_change_request_targets,
    scm_opex_tender_rows,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# File upload constants
# ---------------------------------------------------------------------------

ALLOWED_DOCUMENT_EXTENSIONS = ['pdf', 'doc', 'docx', 'xls', 'xlsx']
ALLOWED_PHOTO_EXTENSIONS    = ['jpg', 'jpeg', 'png']
ALLOWED_EXTENSIONS          = ALLOWED_DOCUMENT_EXTENSIONS + ALLOWED_PHOTO_EXTENSIONS
MAX_FILE_SIZE_BYTES         = 20 * 1024 * 1024  # 20 MB

MIME_TYPE_MAP = {
    'pdf':  'application/pdf',
    'doc':  'application/msword',
    'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    'xls':  'application/vnd.ms-excel',
    'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'jpg':  'image/jpeg',
    'jpeg': 'image/jpeg',
    'png':  'image/png',
}


# ---------------------------------------------------------------------------
# Dashboard context (EPC Residential / Tenders)
# ---------------------------------------------------------------------------
# A pure DISPLAY filter for the three department-level dashboards (CEO, Finance,
# SCM) whose querysets today mix Residential projects with OPEX/CAPEX tender
# sites. It changes what is *shown*, never what is *permitted* — no permission
# function is consulted, called differently, or modified anywhere in this feature.
#
# Context lives in the URL query string only (?context=...). Nothing is written
# to the session or to UserProfile, so a deep link to a project is completely
# unaffected: no context, no filter, no redirect.

CONTEXT_RESIDENTIAL = 'residential'
CONTEXT_TENDERS     = 'tenders'
VALID_CONTEXTS      = (CONTEXT_RESIDENTIAL, CONTEXT_TENDERS)

# Values are the exact Project.PROJECT_TYPE_CHOICES strings (models.py:8-12).
# CAPEX has zero rows today; it is included so those sites appear automatically
# the moment the first one is created — no code change needed then.
CONTEXT_PROJECT_TYPES = {
    CONTEXT_RESIDENTIAL: ['Residential'],
    CONTEXT_TENDERS:     ['OPEX', 'CAPEX'],
}

CONTEXT_LABELS = {
    CONTEXT_RESIDENTIAL: 'EPC Residential',
    CONTEXT_TENDERS:     'Tenders',
}


def _read_context(request):
    """
    Pull the context out of the query string. Returns None for absent, blank or
    unrecognised values — and None means 'apply no filter at all', i.e. exactly
    the pre-existing portfolio-wide behaviour. There is no way for a bad value
    to narrow or widen what a user sees.
    """
    raw = (request.GET.get('context') or '').strip().lower()
    return raw if raw in VALID_CONTEXTS else None


def _context_filter(context, prefix=''):
    """
    Filter kwargs restricting a queryset to the given context's project types.

    Returns {} when context is None, so `**_context_filter(None)` added to any
    .filter() call is a no-op and leaves the queryset byte-identical.

    prefix is the ORM path from the model being queried to Project, e.g.
    'project__' for PaymentRequest, 'phase__project__' for Task.
    """
    types = CONTEXT_PROJECT_TYPES.get(context)
    if not types:
        return {}
    return {f'{prefix}project_type__in': types}


def _context_nav(request, context):
    """
    Template context for the header switcher rendered by base.html.

    Only the three LANDING_ROLES dashboards supply this, so no other page grows
    a switcher. Returns None for anyone else, which renders nothing.
    """
    try:
        role = request.user.profile.role
    except Exception:
        return None
    if role not in LANDING_ROLES:
        return None

    path = request.path
    return {
        'current':      context,
        'current_label': CONTEXT_LABELS.get(context, 'All Projects'),
        'residential_url': f'{path}?context={CONTEXT_RESIDENTIAL}',
        'tenders_url':     f'{path}?context={CONTEXT_TENDERS}',
    }


# ---------------------------------------------------------------------------
# Dashboard sectioning (Tenders / EPC Residential)
# ---------------------------------------------------------------------------
# Presentation only, for the PM / Project Coordinator, Design and Sales & BD
# dashboards. Each of those views has already built its row list; this partitions
# that list in place. It issues NO query — every project it inspects is an object
# the view already fetched — and it never adds or drops a row.

SECTION_TENDERS     = 'Tenders'
SECTION_RESIDENTIAL = 'EPC Residential'

# Same source of truth the context filter uses, so the two features can't drift.
_TENDER_TYPES = tuple(CONTEXT_PROJECT_TYPES[CONTEXT_TENDERS])   # ('OPEX', 'CAPEX')


def _apply_project_sections(rows):
    """
    Group an already-built dashboard row list into Tenders then EPC Residential.

    rows: list of dicts, each carrying the fetched Project under key 'project' —
    the shape dashboard_pm, dashboard_design and dashboard_bd all already use.

    Behaviour:
      * Both types present -> returns tenders + residential (Tenders first), with
        'section_label'/'section_count' set on the FIRST row of each group. The
        template renders a header only where those keys appear.
      * Only one type present (or the list is empty) -> returns `rows` completely
        untouched, with no section keys at all, so the dashboard renders exactly
        as it does today.

    Relative order WITHIN each group is preserved: the partition is stable, so
    whatever the view already sorted by still holds inside a section.

    No DB access. `project_type` is read off the in-memory Project instance; none
    of the three views defers or restricts loaded fields, so it is always present.
    A type outside PROJECT_TYPE_CHOICES cannot occur today (verified: 0 rows) and
    would fall into the Residential group rather than vanish from the page.
    """
    tenders     = [r for r in rows if r['project'].project_type in _TENDER_TYPES]
    residential = [r for r in rows if r['project'].project_type not in _TENDER_TYPES]

    # One type only (or none) — no headers, original list and order returned as-is.
    if not tenders or not residential:
        return rows

    tenders[0]['section_label']     = SECTION_TENDERS
    tenders[0]['section_count']     = len(tenders)
    residential[0]['section_label'] = SECTION_RESIDENTIAL
    residential[0]['section_count'] = len(residential)
    return tenders + residential


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

def login_view(request):
    """
    Render login form and authenticate. Redirects already-authenticated users
    onward immediately. CEO/Finance/SCM land on the context chooser; every other
    role goes straight to its role dashboard exactly as before. Access: public.
    """
    if request.user.is_authenticated:
        return redirect(get_post_login_url(request.user))

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect(get_post_login_url(user))
        messages.error(request, 'Invalid username or password')
        return render(request, 'registration/login.html')

    return render(request, 'registration/login.html')


def logout_view(request):
    """Log out the current user and redirect to login page. Access: any authenticated user."""
    logout(request)
    return redirect('/login/')


@login_required
@role_required(list(LANDING_ROLES))
def landing(request):
    """
    Two-card context chooser shown after login to CEO / Finance / SCM.

    Each card links to the user's own role dashboard with ?context= set. The
    counts use the identical portfolio scope the three dashboards use
    (is_deleted=False, status in Active/In Progress) so a card's number always
    matches the dashboard it opens. A zero-count context still renders its card.
    """
    base_qs = Project.objects.filter(is_deleted=False, status__in=['Active', 'In Progress'])
    dashboard_url = get_user_dashboard(request.user)

    cards = []
    for key in VALID_CONTEXTS:
        cards.append({
            'key':   key,
            'label': CONTEXT_LABELS[key],
            'count': base_qs.filter(**_context_filter(key)).count(),
            'url':   f'{dashboard_url}?context={key}',
        })

    return render(request, 'landing.html', {'cards': cards})


# ---------------------------------------------------------------------------
# Dashboards
# ---------------------------------------------------------------------------

@login_required
def dashboard_admin(request):
    """Admin landing page. Access: any authenticated user (no role restriction here — Admin nav is in the template)."""
    return render(request, 'dashboard/admin.html')


def _project_material_badge(project_id, delivery_lookup):
    """
    Compute the PM dashboard Materials badge value from a pre-fetched delivery_lookup dict.
    Called inside the projects_with_progress loop — lookup was built from a single query
    before the loop to avoid N+1 (one query per project would be too expensive).
    Returns 'Pending', 'Partial', or 'Received'.
    """
    project_data = delivery_lookup.get(project_id, {})
    if not project_data:
        return 'Pending'

    categories = ['Solar Modules', 'Structure', 'Inverter', 'BOS']
    statuses = []
    for cat in categories:
        cat_data = project_data.get(cat)
        if not cat_data:
            statuses.append('Pending')
            continue
        received = cat_data['total_received'] or 0
        ordered  = cat_data['total_ordered'] or 0
        has_damage = cat_data['has_damage']
        if received == 0:
            statuses.append('Pending')
        elif received >= ordered and not has_damage:
            statuses.append('Received')
        else:
            statuses.append('Partial')

    if all(s == 'Pending' for s in statuses):
        return 'Pending'
    elif all(s == 'Received' for s in statuses):
        return 'Received'
    else:
        return 'Partial'


def _build_delivery_lookup(project_ids):
    """
    Build a per-project per-category aggregation dict from DCLineItem data.
    Single query across all given project IDs — avoids N+1 in multi-project dashboard loops.
    Returns: {project_id: {category: {'total_ordered', 'total_received', 'has_damage'}}}
    Called by both dashboard_pm and dashboard_scm — shared to prevent independent
    implementations drifting apart (the root cause of the Q4 representation-5 drift).
    """
    delivery_rows = DCLineItem.objects.filter(
        challan__project__project_id__in=project_ids
    ).values(
        'challan__project__project_id', 'boq_category'
    ).annotate(
        total_ordered=Sum('ordered_quantity'),
        total_received=Sum('received_quantity'),
        # Use damaged_quantity (precise numeric field) — not the coarse condition string
        damage_count=Count('pk', filter=Q(damaged_quantity__gt=0)),
    )
    delivery_lookup = {}
    for row in delivery_rows:
        pid = row['challan__project__project_id']
        cat = row['boq_category']
        if pid not in delivery_lookup:
            delivery_lookup[pid] = {}
        delivery_lookup[pid][cat] = {
            'total_ordered':  row['total_ordered'] or 0,
            'total_received': row['total_received'] or 0,
            'has_damage':     row['damage_count'] > 0,
        }
    return delivery_lookup


_ROLE_DASHBOARD = {
    'PM':           'dashboard_pm',
    'Design':       'dashboard_design',
    'SCM':          'dashboard_scm',
    'Site Engineer': 'dashboard_site_engineer',
    'Finance':       'dashboard_finance',
    'BD / Sales':    'dashboard_bd',
    'CEO':           'dashboard_ceo',
}

_FILTER_TITLES = {
    'due-today': 'Tasks Due Today',
    'due-soon':  'Tasks Due in Next 7 Days',
    'overdue':   'Overdue Tasks',
}


@login_required
def tasks_drill_down(request, filter_type):
    """Read-only list of tasks grouped by project, filtered by due-date window.
    Scoping mirrors the logged-in user's dashboard (PM/Design/SCM/SE/etc.)."""
    if filter_type not in _FILTER_TITLES:
        raise Http404

    today   = date.today()
    profile = request.user.profile
    role    = profile.role

    base_qs = Task.objects.filter(
        phase__project__is_deleted=False,
        phase__project__status__in=['Active', 'In Progress'],
        due_date__isnull=False,
    ).select_related('phase__project')

    if role in ('PM', 'Project Coordinator'):
        # Coordinators are scoped exactly like a PM, but to the projects they
        # coordinate. For a pure PM the coordinators clause matches nothing, so
        # this is identical to the old assigned_pm-only filter (additive-only).
        base_qs = base_qs.filter(
            Q(phase__project__assigned_pm=profile) |
            Q(phase__project__coordinators=profile)
        ).distinct()
    elif role == 'Design':
        # Union: FK-owned projects (assigned_design) plus projects where this
        # Design user has been given a task by the Design lead — task-driven
        # visibility, same pattern as the SE dashboard fix.
        base_qs = base_qs.filter(
            Q(phase__project__assigned_design=profile) |
            Q(phase__project__phases__tasks__assigned_to=profile)
        ).distinct()
    elif role == 'Site Engineer':
        base_qs = base_qs.filter(assigned_to=profile)
    # SCM and others: all active non-deleted projects

    active_statuses = [Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED]

    if filter_type == 'due-today':
        tasks_qs = base_qs.filter(due_date=today, status__in=active_statuses)
    elif filter_type == 'due-soon':
        tasks_qs = base_qs.filter(
            due_date__gt=today,
            due_date__lte=today + timedelta(days=7),
            status__in=active_statuses,
        )
    else:  # overdue
        tasks_qs = base_qs.filter(due_date__lt=today).exclude(status=Task.DONE)

    project_map = {}
    for task in tasks_qs.order_by('phase__project__project_id', 'due_date'):
        proj = task.phase.project
        if proj.project_id not in project_map:
            project_map[proj.project_id] = {'project': proj, 'tasks': []}
        project_map[proj.project_id]['tasks'].append(task)

    groups       = list(project_map.values())
    total_count  = sum(len(g['tasks']) for g in groups)
    back_url     = reverse(_ROLE_DASHBOARD.get(role, 'dashboard_pm'))

    return render(request, 'tasks/task_drill_down.html', {
        'page_title':    _FILTER_TITLES[filter_type],
        'filter_type':   filter_type,
        'groups':        groups,
        'total_count':   total_count,
        'project_count': len(groups),
        'today':         today,
        'back_url':      back_url,
    })


@login_required
@role_required(['PM', 'Project Coordinator'])
def dashboard_pm(request):
    """
    PM dashboard: summary cards + per-project progress + task lists.
    Scoped to the projects the user manages — projects they are assigned PM of,
    plus (for a Project Coordinator) projects they coordinate.
    Access: PM and Project Coordinator.
    TODO: the per-project loop runs multiple queries per project — consider
    annotating with Subquery or moving to prefetch_related before scaling.
    """
    pm_profile = request.user.profile

    # Projects this user manages: own PM projects OR projects they coordinate.
    # For a pure PM (no coordinator rows) this equals their assigned projects
    # exactly, so PM behaviour is unchanged — this is strictly additive.
    managed_project_ids = list(
        Project.objects.filter(
            Q(assigned_pm=pm_profile) | Q(coordinators=pm_profile)
        ).values_list('id', flat=True).distinct()
    )

    # Role-appropriate label for the dashboard heading (PM vs Project Coordinator).
    user_role  = pm_profile.role
    role_label = 'Project Coordinator' if user_role == 'Project Coordinator' else 'PM'

    # Summary card counts — each is a single COUNT query
    active_projects = Project.objects.filter(
        id__in=managed_project_ids,
        status__in=['Active', 'In Progress'],
    ).count()

    due_today = Task.objects.filter(
        phase__project_id__in=managed_project_ids,
        due_date=date.today(),
        due_date__isnull=False,
        task_type=Task.INTERNAL,
        status__in=['Not Started', 'In Progress'],
    ).count()

    blocked_tasks = Task.objects.filter(
        phase__project_id__in=managed_project_ids,
        status='Blocked',
    ).count()

    pending_approvals = Task.objects.filter(
        phase__project_id__in=managed_project_ids,
        assigned_role=Task.PM,
        status='Not Started',
    ).count()

    external_pending = Task.objects.filter(
        phase__project_id__in=managed_project_ids,
        task_type=Task.EXTERNAL,
        status__in=['Not Started', 'In Progress'],
    ).count()

    # Draft projects assigned to this PM (Zoho-created or manually created and not yet activated)
    draft_projects = list(
        Project.objects.filter(
            id__in=managed_project_ids,
            status='Draft',
            is_deleted=False,
        ).order_by('-created_at')
    )

    projects_with_progress = []
    for project in Project.objects.filter(id__in=managed_project_ids, status__in=['Active', 'In Progress'], is_deleted=False):
        try:
            project.boq_status = project.boq.status
            project.boq_url    = f'/projects/{project.project_id}/boq/'
        except Exception:
            project.boq_status = None
            project.boq_url    = None

        total_tasks    = Task.objects.filter(phase__project=project).count()
        done_tasks     = Task.objects.filter(phase__project=project, status='Done').count()
        internal_total = Task.objects.filter(phase__project=project, task_type=Task.INTERNAL).count()
        internal_done  = Task.objects.filter(phase__project=project, task_type=Task.INTERNAL, status='Done').count()
        internal_percent = int(internal_done / internal_total * 100) if internal_total else 0
        ext_pending    = Task.objects.filter(
            phase__project=project, task_type=Task.EXTERNAL,
            status__in=['Not Started', 'In Progress'],
        ).count()
        overdue_count  = Task.objects.filter(
            phase__project=project, task_type=Task.INTERNAL,
            due_date__lt=date.today(), due_date__isnull=False,
            status__in=['Not Started', 'In Progress'],
        ).count()
        blocked_count = Task.objects.filter(
            phase__project=project, status='Blocked',
        ).count()
        blocked_tasks_for_project = list(
            Task.objects.filter(phase__project=project, status='Blocked')
            .select_related('phase')[:5]
        )
        overdue_tasks_for_project = list(
            Task.objects.filter(
                phase__project=project, task_type=Task.INTERNAL,
                due_date__lt=date.today(), due_date__isnull=False,
                status__in=['Not Started', 'In Progress'],
            ).select_related('phase')[:5]
        )
        is_delayed = bool(
            project.target_commissioning_date
            and project.target_commissioning_date < date.today()
        )
        delay_days = (
            (date.today() - project.target_commissioning_date).days
            if is_delayed else None
        )
        current_phase  = (
            project.phases
            .filter(tasks__status__in=['Not Started', 'In Progress', 'Blocked'])
            .order_by('phase_order').first()
            or project.phases.order_by('-phase_order').first()
        )
        due_today_for_project = Task.objects.filter(
            phase__project=project,
            due_date=date.today(), due_date__isnull=False,
            status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED],
        ).count()

        milestones_list = []
        for _m in project.milestones.all():
            milestones_list.append({
                'milestone_name':        _m.milestone_name,
                'milestone_description': _m.milestone_description,
                'amount':                _m.amount,
                'status':                _m.status,
            })

        projects_with_progress.append({
            'project':                   project,
            'total_tasks':               total_tasks,
            'done_tasks':                done_tasks,
            'internal_total':            internal_total,
            'internal_done':             internal_done,
            'internal_percent':          internal_percent,
            'external_pending':          ext_pending,
            'overdue_count':             overdue_count,
            'current_phase':             current_phase,
            'blocked_count':             blocked_count,
            'blocked_tasks_for_project': blocked_tasks_for_project,
            'overdue_tasks_for_project': overdue_tasks_for_project,
            'is_delayed':                is_delayed,
            'delay_days':                delay_days,
            'urgency_count':             blocked_count + overdue_count,
            'due_today_count':           due_today_for_project,
            'milestones':                milestones_list,
        })

    projects_with_progress.sort(
        key=lambda row: (row['overdue_count'] + row['blocked_count'], row['is_delayed']),
        reverse=True,
    )

    # Annotate material summary badges — one call to shared helper, not N+1 per project
    if projects_with_progress:
        pm_project_ids = [row['project'].project_id for row in projects_with_progress]
        delivery_lookup = _build_delivery_lookup(pm_project_ids)
        for row in projects_with_progress:
            row['material_summary'] = _project_material_badge(
                row['project'].project_id, delivery_lookup
            )

    # Delivery issues — one batch query grouped by project, attached to each row
    if projects_with_progress:
        _all_dc_issues = list(
            Issue.objects.filter(
                delivery_challan__project_id__in=managed_project_ids,
                status__in=[Issue.OPEN, Issue.IN_PROGRESS],
            )
            .select_related('project', 'delivery_challan', 'raised_by__user')
            .order_by('-raised_at')
        )
        _dc_issues_by_project = {}
        for _issue in _all_dc_issues:
            _pid = _issue.project.project_id
            _dc_issues_by_project.setdefault(_pid, []).append(_issue)
        for row in projects_with_progress:
            row['delivery_issues'] = _dc_issues_by_project.get(row['project'].project_id, [])

    # OPEX design change requests (Part 4.5). True only for OPEX sites whose change
    # window is currently open — QC has started on the live attempt and the site is not
    # released. Residential rows are never in the map, so their cards are unchanged.
    # design_change_request() re-decides authority on POST; this only offers the link.
    #
    # BOTH LISTS ARE ANNOTATED, and that is not belt-and-braces. `projects_with_progress`
    # is filtered to Active/In Progress, but `create_opex_site()` hardcodes status='Draft'
    # and nothing promotes it — so today every OPEX site reaches this PM through
    # `draft_projects` instead, and annotating only the progress rows would put the button
    # somewhere no OPEX site appears. Widening the progress queryset would move those
    # sites from the draft strip into full progress cards, which is a reorganisation of
    # the PM dashboard rather than an integration, so the flag is carried to both places
    # and the button renders wherever the site actually is.
    _cr_targets = pm_change_request_targets(
        request.user,
        [row['project'] for row in projects_with_progress] + draft_projects)
    for row in projects_with_progress:
        row['can_request_design_change'] = row['project'].pk in _cr_targets
    for dp in draft_projects:
        dp.can_request_design_change = dp.pk in _cr_targets

    # Group into Tenders / EPC Residential for display. Runs last, after every
    # per-row annotation and the sort above, so within-section order is exactly
    # the order this dashboard already produced. No query, no row added or lost.
    projects_with_progress = _apply_project_sections(projects_with_progress)

    due_today_tasks = Task.objects.filter(
        phase__project_id__in=managed_project_ids,
        due_date=date.today(), due_date__isnull=False,
        task_type=Task.INTERNAL,
        status__in=['Not Started', 'In Progress'],
    ).select_related('phase__project', 'assigned_to').order_by('phase__project__project_id')

    blocked_tasks_list = Task.objects.filter(
        phase__project_id__in=managed_project_ids,
        status='Blocked',
    ).select_related('phase__project')

    pending_approvals_list = Task.objects.filter(
        phase__project_id__in=managed_project_ids,
        assigned_role=Task.PM,
        status='Not Started',
    ).select_related('phase__project')

    external_pending_list = Task.objects.filter(
        phase__project_id__in=managed_project_ids,
        task_type=Task.EXTERNAL,
        status__in=['Not Started', 'In Progress'],
    ).select_related('phase__project')

    team_due_today = Task.objects.filter(
        phase__project_id__in=managed_project_ids,
        due_date=date.today(), due_date__isnull=False,
        status__in=['Not Started', 'In Progress'],
    ).exclude(assigned_role=Task.PM).select_related('phase__project')

    seven_days_ago = date.today() - timedelta(days=7)
    due_date_changes = DueDateChangeLog.objects.filter(
        task__phase__project_id__in=managed_project_ids,
        changed_at__date__gte=seven_days_ago,
    ).select_related('task__phase__project', 'changed_by__user').order_by('-changed_at')[:30]

    _pm_task_base = Task.objects.filter(
        phase__project_id__in=managed_project_ids,
        phase__project__is_deleted=False,
        phase__project__status__in=['Active', 'In Progress'],
        due_date__isnull=False,
    )
    _active = [Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED]
    _soon   = date.today() + timedelta(days=7)
    tasks_due_today_count = _pm_task_base.filter(due_date=date.today(), status__in=_active).count()
    tasks_due_soon_count  = _pm_task_base.filter(due_date__gt=date.today(), due_date__lte=_soon, status__in=_active).count()
    tasks_overdue_count   = _pm_task_base.filter(due_date__lt=date.today()).exclude(status=Task.DONE).count()

    return render(request, 'dashboard/pm.html', {
        'summary': {
            'active_projects':   active_projects,
            'due_today':         due_today,
            'blocked_tasks':     blocked_tasks,
            'pending_approvals': pending_approvals,
            'external_pending':  external_pending,
            'tasks_due_today':   tasks_due_today_count,
            'tasks_due_soon':    tasks_due_soon_count,
            'tasks_overdue':     tasks_overdue_count,
        },
        'draft_projects':          draft_projects,
        'projects_with_progress': projects_with_progress,
        'due_today_tasks':        due_today_tasks,
        'blocked_tasks_list':     blocked_tasks_list,
        'pending_approvals_list': pending_approvals_list,
        'external_pending_list':  external_pending_list,
        'team_due_today':         team_due_today,
        'due_date_changes':       due_date_changes,
        'today':                  date.today(),
        'all_profiles':           UserProfile.objects.select_related('user').filter(is_active=True).order_by('user__first_name'),
        'design_candidates':      UserProfile.objects.filter(role='Design', is_active=True).select_related('user'),
        'user_role':              user_role,
        'role_label':             role_label,
    })


@login_required
@role_required(['Site Engineer'])
def dashboard_site_engineer(request):
    """SE dashboard. Renders one card per assigned project, sorted by urgency. Site Engineer role only."""
    #
    # Stop-and-report findings (verified against models.py 2026-06-19):
    # 1. DeliveryChallan.EXPECTED='Expected' = GRN not yet confirmed (no line items have received_quantity)
    # 3. is_delayed: view-computed (same logic as PM dashboard) via target_commissioning_date
    # 4. Blocked field: Task.status == 'Blocked' (string); no separate is_blocked boolean
    # 5. task_type values: Task.INTERNAL='Internal', Task.EXTERNAL='External'
    #
    # SE role only — other roles are redirected at the decorator level
    today      = date.today()
    se_profile = request.user.profile

    # Annotate each project with per-SE urgency counts in a single DB round-trip.
    # overdue_count: internal tasks assigned to this SE, past due, not done.
    # Excludes external/authority tasks (task_type='External') — DISCOM delays
    # must never corrupt internal execution metrics. See Day 2 architectural decision.
    # blocked_count: tasks assigned to SE with status='Blocked' (no is_blocked field — verified).
    # pending_grn_count: DCs for this project with status='Expected' (no GRN confirmed yet).
    # issue_count: open issues on this project (Open or In Progress).
    projects = Project.objects.filter(
        is_deleted=False,
        status__in=['Active', 'In Progress'],
        phases__tasks__assigned_to=se_profile,
    ).distinct().annotate(
        overdue_count=Count(
            'phases__tasks',
            filter=Q(
                phases__tasks__task_type=Task.INTERNAL,
                phases__tasks__assigned_to=se_profile,
                phases__tasks__due_date__lt=today,
                phases__tasks__due_date__isnull=False,
                phases__tasks__status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED],
            ),
            distinct=True,
        ),
        blocked_count=Count(
            'phases__tasks',
            filter=Q(
                phases__tasks__assigned_to=se_profile,
                phases__tasks__status=Task.BLOCKED,
            ),
            distinct=True,
        ),
        pending_grn_count=Count(
            'delivery_challans',
            filter=Q(delivery_challans__status=DeliveryChallan.EXPECTED),
            distinct=True,
        ),
        issue_count=Count(
            'issues',
            filter=Q(issues__status__in=[Issue.OPEN, Issue.IN_PROGRESS]),
            distinct=True,
        ),
    )

    projects_data = []
    for project in projects:
        # Compute SE task progress: done / total for tasks explicitly assigned to this SE
        se_tasks_qs = Task.objects.filter(phase__project=project, assigned_to=se_profile)
        se_total    = se_tasks_qs.count()
        se_done     = se_tasks_qs.filter(status=Task.DONE).count()
        progress    = int(se_done / se_total * 100) if se_total else 0

        # is_delayed: same logic as PM dashboard — uses target_commissioning_date
        is_delayed = bool(
            project.target_commissioning_date
            and project.target_commissioning_date < today
        )
        delay_days = (today - project.target_commissioning_date).days if is_delayed else 0

        # Current phase: first phase with an incomplete task (same query pattern as PM dashboard)
        current_phase_obj = (
            project.phases
            .filter(tasks__status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED])
            .order_by('phase_order').first()
            or project.phases.order_by('-phase_order').first()
        )
        phase_name = current_phase_obj.phase_name if current_phase_obj else None

        # Next task: earliest not-done SE-assigned task ordered by due_date.
        # PostgreSQL sorts NULLs last in ASC order, so tasks with dates appear first.
        next_task_obj = (
            Task.objects.filter(
                phase__project=project,
                assigned_to=se_profile,
                status__in=[Task.NOT_STARTED, Task.IN_PROGRESS],
            )
            .order_by('due_date')
            .first()
        )

        next_task    = None
        next_due     = None
        next_overdue = False
        if next_task_obj:
            next_task = next_task_obj.task_name
            if next_task_obj.due_date:
                # Use .day (no leading zero) + strftime('%b') — avoids %-d platform differences
                next_due     = '{} {}'.format(next_task_obj.due_date.day, next_task_obj.due_date.strftime('%b'))
                next_overdue = next_task_obj.due_date < today
            else:
                next_due = 'No due date set'

        urgency_count = (
            project.overdue_count
            + project.blocked_count
            + project.pending_grn_count
            + project.issue_count
        )

        projects_data.append({
            'project_id':        project.project_id,
            'pk':                project.pk,
            'customer_name':     project.customer_name,
            'phase':             phase_name,
            'progress':          progress,
            'is_delayed':        is_delayed,
            'delay_days':        delay_days,
            'overdue_count':     project.overdue_count,
            'blocked_count':     project.blocked_count,
            'pending_grn_count': project.pending_grn_count,
            'issue_count':       project.issue_count,
            'urgency_count':     urgency_count,
            'next_task':         next_task,
            'next_due':          next_due,
            'next_overdue':      next_overdue,
        })

    # Sort in Python after annotation — Django ORM cannot order by a computed
    # urgency_count (sum of multiple annotations) without a subquery.
    # With 3-5 projects per SE, a Python sort is negligible overhead.
    projects_data.sort(key=lambda p: p['urgency_count'], reverse=True)

    total_overdue     = sum(p['overdue_count'] for p in projects_data)
    total_pending_grn = sum(p['pending_grn_count'] for p in projects_data)
    total_issues      = sum(p['issue_count'] for p in projects_data)
    total_urgent      = sum(p['urgency_count'] for p in projects_data)

    _se_task_base = Task.objects.filter(
        assigned_to=se_profile,
        phase__project__is_deleted=False,
        phase__project__status__in=['Active', 'In Progress'],
        due_date__isnull=False,
    )
    _se_active = [Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED]
    _se_soon   = today + timedelta(days=7)
    se_tasks_due_today = _se_task_base.filter(due_date=today, status__in=_se_active).count()
    se_tasks_due_soon  = _se_task_base.filter(due_date__gt=today, due_date__lte=_se_soon, status__in=_se_active).count()
    se_tasks_overdue   = _se_task_base.filter(due_date__lt=today).exclude(status=Task.DONE).count()

    return render(request, 'dashboard/site-engineer.html', {
        'projects':          projects_data,
        'se_first_name':     request.user.first_name or request.user.username,
        'total_overdue':     total_overdue,
        'total_pending_grn': total_pending_grn,
        'total_issues':      total_issues,
        'total_urgent':      total_urgent,
        'tasks_due_today':   se_tasks_due_today,
        'tasks_due_soon':    se_tasks_due_soon,
        'tasks_overdue':     se_tasks_overdue,
        'today':             today,
        'all_profiles':      UserProfile.objects.select_related('user').filter(is_active=True).order_by('user__first_name'),
    })


@login_required
@role_required(['Design'])
def dashboard_design(request):
    """Design dashboard. One card per assigned project, combined urgency circle. Design role only."""
    today          = date.today()
    design_profile = request.user.profile

    # Trigger 1: assigned_design FK confirmed on Project model.
    # Union: FK-owned projects plus projects where this Design user has been
    # given a task by the Design lead — task-driven visibility, same pattern
    # as the SE dashboard fix. assigned_design stays as the ownership field
    # (Phase 2 Design Head role will manage it); this only widens visibility.
    # Prefetch BOQ for each project — avoids N+1 when reading boq.status in loop.
    #
    # OPEX SITES ARE EXEMPT FROM THE STATUS FILTER (Part 4.5).
    # `create_opex_site()` hardcodes status='Draft' and nothing in the OPEX flow ever
    # promotes it — `project_activate` is a manual per-project action that also attaches
    # Residential milestone defaults. Measured before this change: 11 of 12 OPEX sites
    # were 'Draft', so the Active/In Progress filter hid EVERY site under design work
    # from the designer doing it (5 of 5 allocated sites invisible). Residential
    # behaviour is unchanged — Draft Residential projects are still excluded, exactly
    # as before; the exemption is scoped to project_type='OPEX' and nothing else.
    projects_qs = (
        Project.objects.filter(
            Q(assigned_design=design_profile) |
            Q(phases__tasks__assigned_to=design_profile),
            Q(status__in=['Active', 'In Progress']) | Q(project_type='OPEX'),
            is_deleted=False,
        )
        .select_related('program', 'design_assignment', 'design_assignment__assigned_to__user')
        .prefetch_related('boq')
        .distinct()
        .order_by('project_id')
    )

    # Summary totals — scoped to this Design user's visible portfolio (same
    # union as projects_qs, so the stat matches which cards are shown).
    # The OPEX exemption is repeated here on purpose: this stat's whole claim is that it
    # counts the same project set the cards show, and A1 would otherwise have made that
    # false — a Draft OPEX site could render a card whose BOQ the number above ignored.
    total_revisions = BOQ.objects.filter(
        Q(project__assigned_design=design_profile) |
        Q(project__phases__tasks__assigned_to=design_profile),
        Q(project__status__in=['Active', 'In Progress']) | Q(project__project_type='OPEX'),
        status='Revision Requested',
    ).distinct().count()

    # Trigger 5: no SLA/due-date mechanism exists for Design or BOQ submission.
    # total_design_overdue and total_boq_overdue hardcoded to 0 until Zuber
    # defines explicit thresholds in Claude Chat (same pattern as SCM aging thresholds).
    total_design_overdue = 0
    total_boq_overdue    = 0

    project_rows = []
    for project in projects_qs:
        # Trigger 4: is_delayed — same view-computed logic as PM/SE/Finance dashboards
        is_delayed = bool(
            project.target_commissioning_date
            and project.target_commissioning_date < today
        )
        delay_days = (
            (today - project.target_commissioning_date).days
            if is_delayed else None
        )

        # Trigger 3: BOQ status has 4 values — Draft / Submitted / Acknowledged /
        # Revision Requested. No BOQ at all is treated as Draft (not yet started).
        try:
            boq_status = project.boq.status  # Draft/Submitted/Acknowledged/Revision Requested
        except BOQ.DoesNotExist:
            boq_status = 'Draft'

        # Trigger 2: design_status derived from BOQ — 'submitted' when Design has
        # delivered a BOQ (Submitted or Acknowledged); 'pending' otherwise.
        design_status = 'submitted' if boq_status in ('Submitted', 'Acknowledged') else 'pending'

        # Trigger 4: revision_requested is a BOQ status value, not a separate model/flag.
        revision_requested = (boq_status == 'Revision Requested')

        # Trigger 5: overdue flags hardcoded False — no SLA/due-date mechanism exists
        # for Design or BOQ submission. Wire to real SLA logic once thresholds are
        # defined in Claude Chat. Status chips still display correctly without this.
        design_overdue = False  # TODO: wire to real SLA once thresholds defined
        boq_overdue    = False  # TODO: wire to real SLA once thresholds defined

        # Urgency formula: overdue items + explicit revision flag.
        # With both overdue flags False today, only revision_requested counts.
        urgency_count = (
            (1 if (design_status == 'pending' and design_overdue) else 0) +
            (1 if (boq_status not in ('Submitted', 'Acknowledged') and boq_overdue) else 0) +
            (1 if revision_requested else 0)
        )

        project_rows.append({
            'project':             project,
            'customer_name':       project.customer_name,
            'project_id':          project.project_id,
            'is_delayed':          is_delayed,
            'delay_days':          delay_days,
            'design_status':       design_status,
            'design_overdue':      design_overdue,
            'boq_status':          boq_status,
            'boq_overdue':         boq_overdue,
            'revision_requested':  revision_requested,
            'urgency_count':       urgency_count,
        })

    # Most-urgent-first; all-clear projects sink to bottom
    project_rows.sort(key=lambda r: r['urgency_count'], reverse=True)

    # Group into Tenders / EPC Residential for display — presentation only, applied
    # after the sort so within-section order is unchanged. No query, no row change.
    project_rows = _apply_project_sections(project_rows)

    # phase__project__in=projects_qs guarantees this stat block counts tasks
    # from exactly the same project set as the cards above — no drift between
    # the two if the visibility union ever changes.
    _design_task_base = Task.objects.filter(
        phase__project__in=projects_qs,
        due_date__isnull=False,
    ).distinct()
    _d_active = [Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED]
    _d_soon   = today + timedelta(days=7)
    design_tasks_due_today = _design_task_base.filter(due_date=today, status__in=_d_active).count()
    design_tasks_due_soon  = _design_task_base.filter(due_date__gt=today, due_date__lte=_d_soon, status__in=_d_active).count()
    design_tasks_overdue   = _design_task_base.filter(due_date__lt=today).exclude(status=Task.DONE).count()

    # OPEX design state for the tender cards (Part 4.5). Computed in design_views so the
    # workflow's own rules stay in the design module; this view only carries the result
    # to the template. Keyed by project pk — Residential rows are absent from the map, so
    # their cards render exactly as before.
    design_ctx = designer_dashboard_context(
        design_profile, [r['project'] for r in project_rows])
    for row in project_rows:
        row['design'] = design_ctx.get(row['project'].pk)

    return render(request, 'dashboard/design.html', {
        'design_first_name':    request.user.first_name,
        'total_revisions':      total_revisions,
        'total_design_overdue': total_design_overdue,
        'total_boq_overdue':    total_boq_overdue,
        'project_rows':         project_rows,
        'today':                today,
        'tasks_due_today':      design_tasks_due_today,
        'tasks_due_soon':       design_tasks_due_soon,
        'tasks_overdue':        design_tasks_overdue,
        # None unless this user holds Design Head authority (flag or named deputy).
        'head_counts':          design_head_dashboard_counts(request.user),
    })


@login_required
@role_required(['Finance'])
def dashboard_finance(request):
    """Finance dashboard. One card per project, milestone + payment-request status. Finance role only."""
    today = date.today()

    # Display context (EPC Residential / Tenders). None => no filter => the exact
    # portfolio-wide behaviour this dashboard had before the context feature.
    # Applied to the project list AND to every summary aggregate below, so the
    # header tiles never report a wider scope than the cards underneath them.
    ctx = _read_context(request)

    # Trigger 3: Finance is department-level — no assigned_finance field on Project.
    # Show all active/in-progress projects across the portfolio.
    # Trigger 1: milestones live on PaymentMilestone model (related_name='milestones').
    # Trigger 2: PaymentRequest exists — prefetch pending requests only, with vendor name.
    projects_qs = (
        Project.objects.filter(is_deleted=False, status__in=['Active', 'In Progress'],
                               **_context_filter(ctx))
        .prefetch_related(
            'milestones',
            Prefetch(
                'payment_requests',
                queryset=PaymentRequest.objects.filter(
                    status=PaymentRequest.PENDING,
                ).select_related('vendor'),
                to_attr='pending_payment_requests',
            ),
        )
        .order_by('project_id')
    )

    # Top-level summary counts — single query each, scoped to active portfolio
    total_milestones_awaiting = PaymentMilestone.objects.filter(
        project__status__in=['Active', 'In Progress'],
        status=PaymentMilestone.PENDING,
        **_context_filter(ctx, 'project__'),
    ).count()

    # Trigger 2: PaymentRequest confirmed present — query real pending count and value
    total_payment_requests = PaymentRequest.objects.filter(
        project__status__in=['Active', 'In Progress'],
        status=PaymentRequest.PENDING,
        **_context_filter(ctx, 'project__'),
    ).count()

    total_payment_request_value = (
        PaymentRequest.objects.filter(
            project__status__in=['Active', 'In Progress'],
            status=PaymentRequest.PENDING,
            **_context_filter(ctx, 'project__'),
        ).aggregate(s=Sum('amount'))['s'] or 0
    )

    total_client_contract_value = (
        Project.objects.filter(
            is_deleted=False,
            status__in=['Active', 'In Progress'],
            **_context_filter(ctx),
        ).aggregate(s=Sum('contract_value'))['s'] or 0
    )

    project_rows = []
    for project in projects_qs:
        # Trigger 4: is_delayed — view-computed, same logic as PM/SE dashboards
        is_delayed = bool(
            project.target_commissioning_date
            and project.target_commissioning_date < today
        )
        delay_days = (
            (today - project.target_commissioning_date).days
            if is_delayed else None
        )

        # Trigger 1: build milestone list from PaymentMilestone rows
        milestones = []
        milestones_awaiting = 0
        for m in project.milestones.all():
            # Show the most recent relevant date: received > invoiced > due
            display_date = m.received_date or m.invoice_date or m.due_date
            milestones.append({
                'label':        m.milestone_name,
                'description':  m.milestone_description,
                'amount':       m.amount,
                'status':       m.status,
                'date':         display_date,
            })
            if m.status == PaymentMilestone.PENDING:
                milestones_awaiting += 1

        # Trigger 2: pending payment requests from prefetch (no extra query)
        payment_requests = []
        for pr in project.pending_payment_requests:
            payment_requests.append({
                'pk':             pr.pk,
                'vendor':         pr.vendor.name if pr.vendor else '—',
                'amount':         pr.amount,
                'invoice_number': pr.invoice_number,
                'requested_date': pr.requested_date,
            })

        # urgency drives card sort order: most attention needed rises to top
        urgency = milestones_awaiting + len(payment_requests)

        project_rows.append({
            'project':                project,
            'customer_name':          project.customer_name,
            'project_id':             project.project_id,
            'is_delayed':             is_delayed,
            'delay_days':             delay_days,
            'milestones':             milestones,
            'milestones_total':       len(milestones),
            'milestones_awaiting':    milestones_awaiting,
            'payment_requests':       payment_requests,
            'payment_requests_count': len(payment_requests),
            'urgency':                urgency,
        })

    # Most-urgent-first; all-clear projects sink to bottom
    project_rows.sort(key=lambda r: r['urgency'], reverse=True)

    return render(request, 'dashboard/finance.html', {
        'finance_first_name':           request.user.first_name,
        'total_milestones_awaiting':    total_milestones_awaiting,
        'total_payment_requests':       total_payment_requests,
        'total_payment_request_value':  total_payment_request_value,
        'total_client_contract_value':  total_client_contract_value,
        'project_rows':                 project_rows,
        'today':                        today,
        'context_nav':                  _context_nav(request, ctx),
    })


@login_required
@role_required(['SCM'])
def dashboard_scm(request):
    """
    SCM dashboard: BOQs awaiting acknowledgment, per-project DC status,
    pending deliveries, material badges, and delivery issues.
    All delivery/procurement signals read from DeliveryChallan/DCLineItem —
    not Task proxies — so they reflect actual receipt state. Access: SCM only.
    """
    today = date.today()

    # Display context (EPC Residential / Tenders). None => no filter => the exact
    # portfolio-wide behaviour this dashboard had before the context feature.
    ctx = _read_context(request)

    # BOQ awaiting SCM acknowledgment (Q3 — confirmed working, left unchanged)
    boq_awaiting = BOQ.objects.filter(
        status='Submitted', **_context_filter(ctx, 'project__'),
    ).count()

    # Deliveries today: DCs with expected_delivery_date=today not yet fully received
    deliveries_today = DeliveryChallan.objects.filter(
        project__status__in=['Active', 'In Progress'],
        expected_delivery_date=today,
        status__in=[DeliveryChallan.EXPECTED, DeliveryChallan.PARTIALLY_RECEIVED],
        **_context_filter(ctx, 'project__'),
    ).count()

    # Overdue: DCs whose expected date has passed and are still unresolved (including Rejected = severe)
    overdue = DeliveryChallan.objects.filter(
        project__status__in=['Active', 'In Progress'],
        expected_delivery_date__lt=today,
        expected_delivery_date__isnull=False,
        status__in=[DeliveryChallan.EXPECTED, DeliveryChallan.PARTIALLY_RECEIVED, DeliveryChallan.REJECTED],
        **_context_filter(ctx, 'project__'),
    ).count()

    # Active projects SCM tracks: all Active/In Progress non-deleted projects
    active_projects = list(
        Project.objects.filter(is_deleted=False, status__in=['Active', 'In Progress'],
                               **_context_filter(ctx))
        .select_related('assigned_pm__user')
        .order_by('project_id')
    )
    active_project_ids = [p.project_id for p in active_projects]

    # dc_by_project: group all existing DCs by project for the Procurement section.
    # SCM uses this to see which projects have DCs and their current receipt state.
    # Projects with no DCs yet appear with an empty challan list.
    raw_challans = (
        DeliveryChallan.objects.filter(
            project__status__in=['Active', 'In Progress'],
            **_context_filter(ctx, 'project__'),
        )
        .select_related('project', 'vendor')
        .order_by('project__project_id', '-dc_date')
    )
    dc_map = {p.project_id: {'project': p, 'challans': []} for p in active_projects}
    for challan in raw_challans:
        pid = challan.project.project_id
        if pid in dc_map:
            dc_map[pid]['challans'].append(challan)
    # Return as list of (project, challans) tuples — mirrors old pos_by_project structure
    # so the template loop shape stays the same (project header + per-challan rows)
    dc_by_project = [(row['project'], row['challans']) for row in dc_map.values()]

    # delivery_challans: unresolved DCs for the Deliveries section — includes Rejected (severe) at top
    # Rejected DCs need urgent SCM attention alongside Expected/Partial
    delivery_challans = (
        DeliveryChallan.objects.filter(
            project__status__in=['Active', 'In Progress'],
            status__in=[
                DeliveryChallan.EXPECTED,
                DeliveryChallan.PARTIALLY_RECEIVED,
                DeliveryChallan.REJECTED,
            ],
            **_context_filter(ctx, 'project__'),
        )
        .select_related('project', 'vendor')
        .order_by('expected_delivery_date', 'project__project_id')
    )

    # Material badges — reuse shared helper; same aggregation as dashboard_pm, no drift risk
    delivery_lookup = _build_delivery_lookup(active_project_ids) if active_project_ids else {}

    # BOQ map: one query for all active projects (avoids N+1 in the loop below).
    # Key by b.project.project_id (string e.g. 'HRP-RES-2026-001'), NOT b.project_id
    # which is the integer auto-PK — mismatches the string pid used in the loop.
    boq_map = {
        b.project.project_id: b
        for b in BOQ.objects.filter(project__project_id__in=active_project_ids)
                             .select_related('project')
    }

    # DC grouping per project — reuse raw_challans (already fetched, ordered -dc_date per project)
    dc_per_project = {pid: row['challans'] for pid, row in dc_map.items()}

    # Open issue counts + oldest raised_at per project — one aggregation query
    issue_agg = (
        Issue.objects.filter(
            project__project_id__in=active_project_ids,
            status__in=[Issue.OPEN, Issue.IN_PROGRESS],
        )
        .values('project__project_id')
        .annotate(count=Count('pk'), oldest_at=Min('raised_at'))
    )
    issue_map = {}
    for _row in issue_agg:
        _pid = _row['project__project_id']
        _oldest = (_row['oldest_at'].date() if _row['oldest_at'] else None)
        issue_map[_pid] = {
            'count':      _row['count'],
            'oldest_age': (today - _oldest).days if _oldest else 0,
        }

    # Per-project pipeline rows with 4-stage status + stall computation
    project_rows = []
    for project in active_projects:
        pid = project.project_id
        boq          = boq_map.get(pid)
        challans     = dc_per_project.get(pid, [])
        material_bdg = _project_material_badge(pid, delivery_lookup)
        issue_data   = issue_map.get(pid, {'count': 0, 'oldest_age': None})

        # — BOQ stage —
        boq_status   = boq.status if boq else None
        boq_age_days = None
        if boq and boq.status == 'Submitted' and boq.submitted_at:
            boq_age_days = (today - boq.submitted_at.date()).days

        # — Order/Procurement stage —
        pending_challans = [c for c in challans if c.status in (
            DeliveryChallan.EXPECTED, DeliveryChallan.PARTIALLY_RECEIVED, DeliveryChallan.REJECTED
        )]
        if not challans:
            order_status   = 'None'
            order_age_days = None
        elif not pending_challans:
            order_status   = 'Received'
            order_age_days = None
        else:
            _severity_rank = {DeliveryChallan.REJECTED: 3,
                              DeliveryChallan.PARTIALLY_RECEIVED: 2,
                              DeliveryChallan.EXPECTED: 1}
            worst_pending  = max(pending_challans, key=lambda c: _severity_rank.get(c.status, 0))
            order_status   = ('Rejected' if worst_pending.status == DeliveryChallan.REJECTED else
                              'Partial'  if worst_pending.status == DeliveryChallan.PARTIALLY_RECEIVED else
                              'Expected')
            oldest_dc_date = min(c.dc_date for c in pending_challans)
            order_age_days = (today - oldest_dc_date).days

        # — Delivery stage —
        if material_bdg == 'Received':
            delivery_status   = 'Received'
            delivery_age_days = None
        else:
            overdue_challans = [
                c for c in challans
                if c.expected_delivery_date and c.expected_delivery_date < today
                and c.status != DeliveryChallan.RECEIVED
            ]
            unscheduled_challans = [
                c for c in challans
                if not c.expected_delivery_date and c.status != DeliveryChallan.RECEIVED
            ]
            if overdue_challans:
                delivery_status   = 'Overdue'
                delivery_age_days = max((today - c.expected_delivery_date).days
                                        for c in overdue_challans)
            elif unscheduled_challans:
                delivery_status   = 'Not Scheduled'
                oldest_dc         = min(c.dc_date for c in unscheduled_challans)
                delivery_age_days = (today - oldest_dc).days
            elif challans:
                delivery_status   = 'Scheduled'
                delivery_age_days = None
            else:
                delivery_status   = 'None'
                delivery_age_days = None

        # — Issues stage —
        open_issue_count = issue_data['count']
        issue_age_days   = issue_data['oldest_age']

        # — Stall computation — worst wins (red > amber; within same level, most days) —
        stall_candidates = []

        if boq_status == 'Submitted' and boq_age_days is not None:
            if boq_age_days > 5:
                stall_candidates.append(('boq', boq_age_days, 'red'))
            elif boq_age_days > 2:
                stall_candidates.append(('boq', boq_age_days, 'amber'))

        if order_status in ('Expected', 'Partial', 'Rejected') and order_age_days is not None:
            if order_age_days > 7:
                stall_candidates.append(('order', order_age_days, 'red'))
            elif order_age_days > 3:
                stall_candidates.append(('order', order_age_days, 'amber'))

        if delivery_status in ('Overdue', 'Not Scheduled') and delivery_age_days is not None:
            if delivery_age_days > 10:
                stall_candidates.append(('delivery', delivery_age_days, 'red'))
            elif delivery_age_days > 5:
                stall_candidates.append(('delivery', delivery_age_days, 'amber'))

        if open_issue_count > 0:
            _iage = issue_age_days or 0
            if _iage > 2:
                stall_candidates.append(('issues', _iage, 'red'))
            else:
                stall_candidates.append(('issues', _iage, 'amber'))

        if stall_candidates:
            _red = [(s, d, l) for (s, d, l) in stall_candidates if l == 'red']
            if _red:
                _worst     = max(_red, key=lambda x: x[1])
                stall_level = 'red'
            else:
                _worst     = max(stall_candidates, key=lambda x: x[1])
                stall_level = 'amber'
            is_stalled    = True
            stalled_stage = _worst[0]
            days_in_stage = _worst[1]
        else:
            is_stalled    = False
            stalled_stage = None
            stall_level   = None
            days_in_stage = None

        # — Action URLs —
        boq_url               = f'/projects/{pid}/boq/'
        schedule_delivery_url = f'/projects/{pid}/delivery-challans/create/'
        finance_url           = f'/projects/{pid}/overview/'
        # payment_request_url: POST endpoint for SCM to raise a payment request against a vendor invoice
        payment_request_url   = f'/projects/{pid}/payment-requests/raise/'
        # raise_issue_url: scope to the most recent pending DC if one exists
        latest_pending_dc = (pending_challans[0] if pending_challans
                             else (challans[0] if challans else None))
        if latest_pending_dc:
            raise_issue_url = (f'/projects/{pid}/delivery-challans/'
                               f'{latest_pending_dc.pk}/issues/create/')
        else:
            raise_issue_url = f'/projects/{pid}/issues/create/'

        project_rows.append({
            'project':             project,
            'material_badge':      material_bdg,
            'boq_status':          boq_status,
            'boq_age_days':        boq_age_days,
            'order_status':        order_status,
            'order_age_days':      order_age_days,
            'delivery_status':     delivery_status,
            'delivery_age_days':   delivery_age_days,
            'open_issue_count':    open_issue_count,
            'issue_age_days':      issue_age_days,
            'is_stalled':          is_stalled,
            'stalled_stage':       stalled_stage,
            'stall_level':         stall_level,
            'days_in_stage':       days_in_stage,
            'boq_url':             boq_url,
            'schedule_delivery_url': schedule_delivery_url,
            'finance_url':         finance_url,
            'payment_request_url': payment_request_url,
            'raise_issue_url':     raise_issue_url,
        })

    # Delivery issues: open/in-progress issues linked to DCs on active SCM-tracked projects
    # SCM scope: all active projects (SCM is not PM-scoped; it sees all active projects)
    delivery_issues = (
        Issue.objects.filter(
            delivery_challan__project__status__in=['Active', 'In Progress'],
            status__in=[Issue.OPEN, Issue.IN_PROGRESS],
            **_context_filter(ctx, 'delivery_challan__project__'),
        )
        .select_related('project', 'delivery_challan', 'raised_by__user')
        .order_by('-raised_at')[:20]
    )

    # BOQ items per project — grouped by category for the raise-payment-request dropdown.
    # Loaded here (not per-request in the modal) so the template can render them as JSON
    # once and the JS layer filters by project without extra round-trips.
    # BOQItem FK chain: BOQItem.boq → BOQ (OneToOne) → Project.
    all_boq_items = (
        BOQItem.objects.filter(boq__project__project_id__in=active_project_ids)
        .select_related('boq__project')
        .order_by('boq__project__project_id', 'serial_no')
    )
    boq_items_by_project = {}
    for item in all_boq_items:
        pid = item.boq.project.project_id
        cat = item.category
        boq_items_by_project.setdefault(pid, {}).setdefault(cat, []).append({
            'id':          item.pk,
            'serial_no':   item.serial_no,
            'description': item.description,
        })

    # All active vendors for the raise-payment-request vendor dropdown
    scm_vendors = list(
        Vendor.objects.filter(is_active=True).order_by('name')
        .values('id', 'name')
    )

    _scm_task_base = Task.objects.filter(
        phase__project__is_deleted=False,
        phase__project__status__in=['Active', 'In Progress'],
        due_date__isnull=False,
        **_context_filter(ctx, 'phase__project__'),
    )
    _s_active = [Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED]
    _s_soon   = today + timedelta(days=7)
    scm_tasks_due_today = _scm_task_base.filter(due_date=today, status__in=_s_active).count()
    scm_tasks_due_soon  = _scm_task_base.filter(due_date__gt=today, due_date__lte=_s_soon, status__in=_s_active).count()
    scm_tasks_overdue   = _scm_task_base.filter(due_date__lt=today).exclude(status=Task.DONE).count()

    # ── OPEX tenders (Part 6) ──────────────────────────────────────────────────
    # A SEPARATE QUERY SET, NOT A WIDENING OF THE ONE ABOVE. `active_projects` keeps its
    # status__in=['Active','In Progress'] filter untouched; OPEX sites are `Draft` and
    # would never appear there (deferred finding H1). This section keys off SiteGroup and
    # DesignAssignment instead, exactly as Part 6's settled decision 9 requires, so nothing
    # on the existing Residential half of this dashboard changes.
    #
    # Not context-filtered: the two-context switch (_context_filter) separates Residential
    # from Tenders, and this section IS the tender half — filtering it by the Residential
    # context would empty it.
    opex_tender_rows = scm_opex_tender_rows()

    return render(request, 'dashboard/scm.html', {
        'summary': {
            'boq_awaiting':     boq_awaiting,
            'deliveries_today': deliveries_today,
            'overdue':          overdue,
            'tasks_due_today':  scm_tasks_due_today,
            'tasks_due_soon':   scm_tasks_due_soon,
            'tasks_overdue':    scm_tasks_overdue,
        },
        'project_rows':         project_rows,
        'opex_tender_rows':     opex_tender_rows,
        'delivery_issues':      delivery_issues,
        'today':                today,
        'all_profiles':         UserProfile.objects.select_related('user').filter(is_active=True).order_by('user__first_name'),
        'boq_items_by_project': json.dumps(boq_items_by_project),
        'scm_vendors':          json.dumps(scm_vendors),
        'context_nav':          _context_nav(request, ctx),
    })


def _get_ceo_dashboard_context(context=None):
    """
    Aggregates portfolio-wide metrics for the CEO dashboard in exactly 3 DB queries.
    Query 1: Active project list with Exists annotations for blocked/at_risk classification.
    Query 2: Full task aggregate — status counts, dept rollup (18 cells), KPI time windows,
             blocked KPI, external-dependency KPI — all via a single .aggregate() call.
    Query 3: Full issue aggregate — status counts and resolution time windows.
    Returns a dict ready to be unpacked into the template context.
    """
    today  = date.today()
    now_dt = timezone.now()

    # -- Date-window boundaries (computed once; reused across all conditional filters) --
    week_start       = today - timedelta(days=today.weekday())           # Monday of current ISO week
    week_end         = week_start + timedelta(days=7)                    # Monday of next week (exclusive)
    last_week_start  = week_start - timedelta(days=7)
    last_week_end    = week_start                                        # == this week's Monday
    this_month_start = today.replace(day=1)
    if today.month == 12:
        next_month_start = date(today.year + 1, 1, 1)
    else:
        next_month_start = date(today.year, today.month + 1, 1)
    last_month_start = (this_month_start - timedelta(days=1)).replace(day=1)

    # completed_at / resolved_at are DateTimeFields — boundaries must be datetimes
    from django.utils.timezone import make_aware
    def _to_dt(d):
        return make_aware(datetime(d.year, d.month, d.day, 0, 0, 0))

    week_start_dt       = _to_dt(week_start)
    week_end_dt         = _to_dt(week_end)
    last_week_start_dt  = _to_dt(last_week_start)
    last_week_end_dt    = week_start_dt
    this_month_start_dt = _to_dt(this_month_start)
    next_month_start_dt = _to_dt(next_month_start)
    last_month_start_dt = _to_dt(last_month_start)
    aged_block_cutoff   = now_dt - timedelta(days=7)

    active_statuses = ['Active', 'In Progress']

    # -- QUERY 1: Active project list + Exists subquery annotations --
    # Subquery: does this project have any task with status='Blocked'?
    blocked_subq = Task.objects.filter(
        phase__project=OuterRef('pk'),
        status=Task.BLOCKED,
    )
    # Subquery: does this project have any overdue internal task (still future target date handled in Python)?
    at_risk_subq = Task.objects.filter(
        phase__project=OuterRef('pk'),
        task_type=Task.INTERNAL,
        due_date__lt=today,
        due_date__isnull=False,
        status__in=[Task.NOT_STARTED, Task.IN_PROGRESS],
    )
    projects_qs = (
        Project.objects
        .filter(is_deleted=False, status__in=active_statuses, **_context_filter(context))
        .annotate(
            has_blocked_task=Exists(blocked_subq),
            has_at_risk_task=Exists(at_risk_subq),
        )
        .order_by('target_commissioning_date', 'project_id')
    )

    # Classify each project in Python (no extra DB hits)
    # Precedence: Blocked > Delayed > At Risk > On Time (worst condition wins)
    proj_total = proj_on_time = proj_at_risk = proj_delayed = proj_blocked = 0
    project_cards = []
    for p in projects_qs:
        proj_total += 1
        is_delayed = bool(p.target_commissioning_date and p.target_commissioning_date < today)

        if p.has_blocked_task:
            badge = 'blocked'
            proj_blocked += 1
        elif is_delayed:
            badge = 'delayed'
            proj_delayed += 1
        elif p.has_at_risk_task and not is_delayed:
            # At Risk = overdue internal task but target date still in the future
            badge = 'at_risk'
            proj_at_risk += 1
        else:
            badge = 'on_time'
            proj_on_time += 1

        project_cards.append({'project': p, 'badge': badge, 'is_delayed': is_delayed})

    # -- QUERY 2: Task aggregate — single .aggregate() call, ~40 conditional Counts --
    task_agg = Task.objects.filter(
        phase__project__is_deleted=False,
        phase__project__status__in=active_statuses,
        **_context_filter(context, 'phase__project__'),
    ).aggregate(
        task_total     =Count('pk'),
        # Status summary (portfolio-wide)
        task_unassigned=Count('pk', filter=Q(assigned_to__isnull=True)),
        task_inprogress=Count('pk', filter=Q(status=Task.IN_PROGRESS)),
        task_completed =Count('pk', filter=Q(status=Task.DONE)),
        # Overdue = internal tasks only; external delays are not team overdue (SE has DISCOM tasks)
        task_overdue   =Count('pk', filter=Q(
            task_type=Task.INTERNAL, due_date__lt=today, due_date__isnull=False,
            status__in=[Task.NOT_STARTED, Task.IN_PROGRESS],
        )),
        # Blocked KPI (two distinct numbers — see Layer 5 §3: they will not reconcile, that is expected)
        blocked_open   =Count('pk', filter=Q(status=Task.BLOCKED)),
        blocked_aged_7d=Count('pk', filter=Q(
            status=Task.BLOCKED,
            blocked_since__lte=aged_block_cutoff,
            blocked_since__isnull=False,
        )),
        # External dependency KPI
        ext_closed =Count('pk', filter=Q(task_type=Task.EXTERNAL, status=Task.DONE)),
        ext_overdue=Count('pk', filter=Q(
            task_type=Task.EXTERNAL, due_date__lt=today, due_date__isnull=False,
            status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED],
        )),
        # Due-date windows — internal-only (matches Overdue convention); Blocked counts as still-open.
        # "This week" reuses the Mon–Sun boundaries above; today's tasks intentionally count in both.
        due_today_count    =Count('pk', filter=Q(
            task_type=Task.INTERNAL, due_date=today,
            status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED],
        )),
        due_this_week_count=Count('pk', filter=Q(
            task_type=Task.INTERNAL, due_date__gte=week_start, due_date__lt=week_end,
            status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED],
        )),
        # KPI time windows — use completed_at (DateTimeField, set by task_status_update on Done)
        task_done_this_week  =Count('pk', filter=Q(status=Task.DONE, completed_at__gte=week_start_dt,       completed_at__lt=week_end_dt)),
        task_done_last_week  =Count('pk', filter=Q(status=Task.DONE, completed_at__gte=last_week_start_dt,  completed_at__lt=last_week_end_dt)),
        task_done_this_month =Count('pk', filter=Q(status=Task.DONE, completed_at__gte=this_month_start_dt, completed_at__lt=next_month_start_dt)),
        task_done_last_month =Count('pk', filter=Q(status=Task.DONE, completed_at__gte=last_month_start_dt, completed_at__lt=this_month_start_dt)),
        # -- Department rollup: 6 roles × 3 columns = 18 conditional Counts --
        # PM
        dept_pm_assigned=Count('pk', filter=Q(assigned_role=Task.PM, assigned_to__isnull=False)),
        dept_pm_pending =Count('pk', filter=Q(assigned_role=Task.PM, status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED])),
        dept_pm_overdue =Count('pk', filter=Q(assigned_role=Task.PM, task_type=Task.INTERNAL, due_date__lt=today, due_date__isnull=False, status__in=[Task.NOT_STARTED, Task.IN_PROGRESS])),
        # Site Engineer — overdue MUST filter Internal; SE tasks include DISCOM/authority External tasks
        dept_se_assigned=Count('pk', filter=Q(assigned_role=Task.SITE_ENGINEER, assigned_to__isnull=False)),
        dept_se_pending =Count('pk', filter=Q(assigned_role=Task.SITE_ENGINEER, status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED])),
        dept_se_overdue =Count('pk', filter=Q(assigned_role=Task.SITE_ENGINEER, task_type=Task.INTERNAL, due_date__lt=today, due_date__isnull=False, status__in=[Task.NOT_STARTED, Task.IN_PROGRESS])),
        # SCM
        dept_scm_assigned=Count('pk', filter=Q(assigned_role=Task.SCM, assigned_to__isnull=False)),
        dept_scm_pending =Count('pk', filter=Q(assigned_role=Task.SCM, status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED])),
        dept_scm_overdue =Count('pk', filter=Q(assigned_role=Task.SCM, task_type=Task.INTERNAL, due_date__lt=today, due_date__isnull=False, status__in=[Task.NOT_STARTED, Task.IN_PROGRESS])),
        # Design
        dept_design_assigned=Count('pk', filter=Q(assigned_role=Task.DESIGN, assigned_to__isnull=False)),
        dept_design_pending =Count('pk', filter=Q(assigned_role=Task.DESIGN, status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED])),
        dept_design_overdue =Count('pk', filter=Q(assigned_role=Task.DESIGN, task_type=Task.INTERNAL, due_date__lt=today, due_date__isnull=False, status__in=[Task.NOT_STARTED, Task.IN_PROGRESS])),
        # BD / Sales — filter on assigned_role='BD / Sales' (Task.BD constant); 'BD' alone returns zero
        dept_bd_assigned=Count('pk', filter=Q(assigned_role=Task.BD, assigned_to__isnull=False)),
        dept_bd_pending =Count('pk', filter=Q(assigned_role=Task.BD, status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED])),
        dept_bd_overdue =Count('pk', filter=Q(assigned_role=Task.BD, task_type=Task.INTERNAL, due_date__lt=today, due_date__isnull=False, status__in=[Task.NOT_STARTED, Task.IN_PROGRESS])),
        # Finance — overdue will read near-zero if finance tasks lack due dates; that is expected, not a bug
        dept_finance_assigned=Count('pk', filter=Q(assigned_role=Task.FINANCE, assigned_to__isnull=False)),
        dept_finance_pending =Count('pk', filter=Q(assigned_role=Task.FINANCE, status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED])),
        dept_finance_overdue =Count('pk', filter=Q(assigned_role=Task.FINANCE, task_type=Task.INTERNAL, due_date__lt=today, due_date__isnull=False, status__in=[Task.NOT_STARTED, Task.IN_PROGRESS])),
        # -- Per-department due-date windows: 6 roles × 2 columns (Due Today / Due This Week) --
        # Same convention as the portfolio-wide due counts: internal-only, open incl. Blocked.
        # "This week" reuses the Mon–Sun boundaries; today's tasks intentionally count in both.
        dept_pm_due_today =Count('pk', filter=Q(assigned_role=Task.PM,            task_type=Task.INTERNAL, due_date=today,                                        status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED])),
        dept_pm_due_week  =Count('pk', filter=Q(assigned_role=Task.PM,            task_type=Task.INTERNAL, due_date__gte=week_start, due_date__lt=week_end, status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED])),
        dept_scm_due_today=Count('pk', filter=Q(assigned_role=Task.SCM,           task_type=Task.INTERNAL, due_date=today,                                        status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED])),
        dept_scm_due_week =Count('pk', filter=Q(assigned_role=Task.SCM,           task_type=Task.INTERNAL, due_date__gte=week_start, due_date__lt=week_end, status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED])),
        dept_design_due_today=Count('pk', filter=Q(assigned_role=Task.DESIGN,     task_type=Task.INTERNAL, due_date=today,                                        status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED])),
        dept_design_due_week =Count('pk', filter=Q(assigned_role=Task.DESIGN,     task_type=Task.INTERNAL, due_date__gte=week_start, due_date__lt=week_end, status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED])),
        dept_bd_due_today =Count('pk', filter=Q(assigned_role=Task.BD,            task_type=Task.INTERNAL, due_date=today,                                        status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED])),
        dept_bd_due_week  =Count('pk', filter=Q(assigned_role=Task.BD,            task_type=Task.INTERNAL, due_date__gte=week_start, due_date__lt=week_end, status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED])),
        dept_se_due_today =Count('pk', filter=Q(assigned_role=Task.SITE_ENGINEER, task_type=Task.INTERNAL, due_date=today,                                        status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED])),
        dept_se_due_week  =Count('pk', filter=Q(assigned_role=Task.SITE_ENGINEER, task_type=Task.INTERNAL, due_date__gte=week_start, due_date__lt=week_end, status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED])),
        dept_finance_due_today=Count('pk', filter=Q(assigned_role=Task.FINANCE,   task_type=Task.INTERNAL, due_date=today,                                        status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED])),
        dept_finance_due_week =Count('pk', filter=Q(assigned_role=Task.FINANCE,   task_type=Task.INTERNAL, due_date__gte=week_start, due_date__lt=week_end, status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED])),
    )

    # -- QUERY 3: Issue aggregate — status counts + resolution time windows --
    issue_agg = Issue.objects.filter(
        project__is_deleted=False,
        project__status__in=active_statuses,
        **_context_filter(context, 'project__'),
    ).aggregate(
        issue_total     =Count('pk'),
        issue_unassigned=Count('pk', filter=Q(assigned_to__isnull=True)),
        issue_inprogress=Count('pk', filter=Q(status=Issue.IN_PROGRESS)),
        # Include Closed: terminal state after Resolved; PM closes issues from the Resolved state
        issue_resolved  =Count('pk', filter=Q(status__in=[Issue.RESOLVED, Issue.CLOSED])),
        # Overdue uses due_date (nullable); will read low if few issues have due dates — accepted limitation
        issue_overdue   =Count('pk', filter=Q(
            due_date__lt=today, due_date__isnull=False,
            status__in=[Issue.OPEN, Issue.IN_PROGRESS],
        )),
        # Time windows use resolved_at; Closed issues that went through Resolved will also have it set
        issue_done_this_week  =Count('pk', filter=Q(resolved_at__isnull=False, resolved_at__gte=week_start_dt,       resolved_at__lt=week_end_dt)),
        issue_done_last_week  =Count('pk', filter=Q(resolved_at__isnull=False, resolved_at__gte=last_week_start_dt,  resolved_at__lt=last_week_end_dt)),
        issue_done_this_month =Count('pk', filter=Q(resolved_at__isnull=False, resolved_at__gte=this_month_start_dt, resolved_at__lt=next_month_start_dt)),
        issue_done_last_month =Count('pk', filter=Q(resolved_at__isnull=False, resolved_at__gte=last_month_start_dt, resolved_at__lt=this_month_start_dt)),
    )

    # Build dept_rows list for clean template iteration
    dept_rows = [
        {'label': 'PM',        'assigned': task_agg['dept_pm_assigned'],     'pending': task_agg['dept_pm_pending'],     'overdue': task_agg['dept_pm_overdue']},
        {'label': 'SCM',       'assigned': task_agg['dept_scm_assigned'],    'pending': task_agg['dept_scm_pending'],    'overdue': task_agg['dept_scm_overdue']},
        {'label': 'Design',    'assigned': task_agg['dept_design_assigned'], 'pending': task_agg['dept_design_pending'], 'overdue': task_agg['dept_design_overdue']},
        {'label': 'BD',        'assigned': task_agg['dept_bd_assigned'],     'pending': task_agg['dept_bd_pending'],     'overdue': task_agg['dept_bd_overdue']},
        {'label': 'Execution', 'assigned': task_agg['dept_se_assigned'],     'pending': task_agg['dept_se_pending'],     'overdue': task_agg['dept_se_overdue']},
        {'label': 'Finance',   'assigned': task_agg['dept_finance_assigned'],'pending': task_agg['dept_finance_pending'],'overdue': task_agg['dept_finance_overdue']},
    ]

    # -- QUERY 4: Finance summary (payment requests + contract value) --
    active_filter = {
        'project__is_deleted': False,
        'project__status__in': ['Active', 'In Progress'],
        **_context_filter(context, 'project__'),
    }
    fin_payment_requests_pending = PaymentRequest.objects.filter(
        status=PaymentRequest.PENDING, **active_filter
    ).count()
    fin_vendor_payments_outstanding = (
        PaymentRequest.objects.filter(
            status=PaymentRequest.PENDING, **active_filter
        ).aggregate(s=Sum('amount'))['s'] or 0
    )
    fin_client_contract_value = (
        Project.objects.filter(
            is_deleted=False, status__in=['Active', 'In Progress'],
            **_context_filter(context),
        ).aggregate(s=Sum('contract_value'))['s'] or 0
    )
    # Milestones Finance has invoiced (triggered by task Done) but not yet fully collected.
    # status='Invoiced' → all of amount is owed; status='Received' with partial → amount-amount_received.
    # Overpaid/exact rows excluded at DB level (amount_received__gte=amount).
    _DECIMAL_ZERO = Value(Decimal('0'), output_field=DecimalField())
    fin_client_payment_pending = (
        PaymentMilestone.objects.filter(
            project__is_deleted=False,
            project__status__in=['Active', 'In Progress'],
            status__in=['Invoiced', 'Received'],
            amount__isnull=False,
            **_context_filter(context, 'project__'),
        ).filter(
            Q(amount_received__isnull=True) | Q(amount_received__lt=F('amount'))
        ).aggregate(
            s=Sum(F('amount') - Coalesce(F('amount_received'), _DECIMAL_ZERO))
        )['s'] or 0
    )

    # -- QUERY 5: Top-5 assignee leaderboard (open tasks per user) --
    # Open = Not Started / In Progress / Blocked (Blocked counts as still on the plate).
    # select_related pulls the profile→user names in the same query — no per-row round-trip.
    top_assignees = list(
        Task.objects.filter(
            phase__project__is_deleted=False,
            phase__project__status__in=active_statuses,
            assigned_to__isnull=False,
            status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED],
            **_context_filter(context, 'phase__project__'),
        )
        .values(
            'assigned_to',
            'assigned_to__user__first_name',
            'assigned_to__user__last_name',
            'assigned_to__user__username',
        )
        .annotate(count=Count('pk'))
        .order_by('-count')[:5]
    )
    # Resolve a display name (get_full_name equivalent) without a second query.
    for row in top_assignees:
        full_name = f"{row['assigned_to__user__first_name']} {row['assigned_to__user__last_name']}".strip()
        row['display_name'] = full_name or row['assigned_to__user__username']

    ctx = {
        'proj_total':    proj_total,
        'proj_on_time':  proj_on_time,
        'proj_at_risk':  proj_at_risk,
        'proj_delayed':  proj_delayed,
        'proj_blocked':  proj_blocked,
        'project_cards': project_cards,
        'dept_rows':     dept_rows,
        'top_assignees': top_assignees,
        'fin_payment_requests_pending':    fin_payment_requests_pending,
        'fin_vendor_payments_outstanding': fin_vendor_payments_outstanding,
        'fin_client_contract_value':       fin_client_contract_value,
        'fin_client_payment_pending':      fin_client_payment_pending,
    }
    ctx.update(task_agg)
    ctx.update(issue_agg)
    return ctx


@login_required
def dashboard_ceo(request):
    """CEO portfolio overview. Access: CEO role only. Renders in 3 DB queries via _get_ceo_dashboard_context."""
    # Display context (EPC Residential / Tenders). None => no filter => the exact
    # portfolio-wide behaviour this dashboard had before the context feature.
    selected_context = _read_context(request)
    ctx = _get_ceo_dashboard_context(selected_context)
    ctx['now'] = timezone.now()
    ctx['context_nav'] = _context_nav(request, selected_context)
    return render(request, 'dashboard/ceo.html', ctx)


# ---------------------------------------------------------------------------
# User management (Admin only)
# ---------------------------------------------------------------------------

@login_required
@role_required(['Admin'])
def user_list(request):
    """List all user profiles with their roles. Access: Admin only."""
    profiles = UserProfile.objects.select_related('user', 'created_by').order_by('user__username')
    return render(request, 'users/user_list.html', {'profiles': profiles})


@login_required
@role_required(['Admin'])
def user_create(request):
    """
    Create a new Django User + UserProfile in one step.
    Sets is_staff=True when role is Admin so Django admin access is granted automatically.
    Access: Admin only.
    """
    if request.method == 'POST':
        form = UserCreateForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            user = User.objects.create_user(
                username=cd['username'],
                password=cd['password'],
                first_name=cd['first_name'],
                last_name=cd['last_name'],
                email=cd['email'],
                is_active=cd['is_active'],
                is_staff=(cd['role'] == 'Admin'),
            )
            # UserProfile is auto-created by a post_save signal on User creation
            profile = user.profile
            profile.role = cd['role']
            profile.phone_number = cd['phone_number']
            profile.is_active = cd['is_active']
            profile.created_by = request.user
            profile.save()

            messages.success(
                request,
                f"User {cd['username']} created successfully as {cd['role']}"
            )
            return redirect('user_list')
    else:
        form = UserCreateForm()

    return render(request, 'users/user_form.html', {'form': form, 'action': 'Create'})


@login_required
@role_required(['Admin'])
def user_edit(request, user_id):
    """
    Edit an existing user's name, email, role, and active status.
    Uses filter().update() on UserProfile to avoid a second .save() round-trip.
    Access: Admin only.
    """
    target_user = get_object_or_404(User, pk=user_id)
    try:
        profile = target_user.profile
    except UserProfile.DoesNotExist:
        # Safety net: create a blank profile if one somehow doesn't exist
        profile = UserProfile.objects.create(user=target_user)

    if request.method == 'POST':
        form = UserEditForm(request.POST, instance_user=target_user)
        if form.is_valid():
            cd = form.cleaned_data
            target_user.first_name = cd['first_name']
            target_user.last_name = cd['last_name']
            target_user.email = cd['email']
            target_user.is_active = cd['is_active']
            target_user.is_staff = (cd['role'] == 'Admin')
            target_user.save()

            UserProfile.objects.filter(user=target_user).update(
                role=cd['role'],
                phone_number=cd['phone_number'],
                is_active=cd['is_active'],
            )

            messages.success(request, f"User {target_user.username} updated successfully.")
            return redirect('user_list')
    else:
        form = UserEditForm(
            initial={
                'first_name':   target_user.first_name,
                'last_name':    target_user.last_name,
                'email':        target_user.email,
                'role':         profile.role,
                'phone_number': profile.phone_number,
                'is_active':    profile.is_active,
            },
            instance_user=target_user,
        )

    return render(request, 'users/user_form.html', {
        'form': form,
        'action': 'Edit',
        'target_user': target_user,
    })


# ---------------------------------------------------------------------------
# Projects
# ---------------------------------------------------------------------------

def _get_user_role(request):
    """Return the role string for the request user, or None if no profile exists."""
    try:
        return request.user.profile.role
    except Exception:
        return None


def _pm_owns_project(request, project):
    """Return True if the request user is the assigned PM on this project.

    Thin adapter over the canonical user_can_manage_project() — kept so its
    existing callers stay unchanged. No ownership comparison lives here anymore.
    """
    return user_can_manage_project(request.user, project)


def _user_can_complete_checklist_item(user, task, project):
    """Return True if `user` may tick / photograph checklist items on `task`.

    Locked decision (Prompt 1): the SAME model as task_status_update's permission
    check — the user's role matches task.assigned_role OR they have PM/coordinator
    authority on the project. This is deliberately broader than the assigned-user-only
    rule task_detail_status_update uses; both surfaces live on the same page.
    Defined once and reused by both the completion view and the template context so
    the two never drift.
    """
    profile = getattr(user, 'profile', None)
    if profile is None:
        return False
    if user_can_manage_project(user, project):
        return True
    # Task.BD = 'BD / Sales' but UserProfile stores 'BD' — normalise before comparison
    _PROFILE_TO_TASK_ROLE = {'BD': 'BD / Sales'}
    normalised_user_role = _PROFILE_TO_TASK_ROLE.get(profile.role, profile.role)
    return normalised_user_role == task.assigned_role


@login_required
@role_required(['PM', 'Admin', 'CEO'])
def project_list(request):
    """
    Redirect to the role-appropriate projects view.
    Admin  → Admin Panel project list (full table with assign-PM + delete)
    PM     → PM dashboard (draft + active project cards)
    CEO    → CEO dashboard
    Kept as a named URL so '← Projects' links in project_overview still resolve.
    """
    role = _get_user_role(request)
    if role == 'Admin':
        return redirect('admin_project_list')
    if role == 'CEO':
        return redirect('dashboard_ceo')
    return redirect('dashboard_pm')


@login_required
@role_required(['PM'])
def project_create(request):
    """
    Create a new Draft project. The PM is auto-set to the current user.
    project_id is generated inside Project.save() via generate_project_id().
    Access: PM only.
    """
    if request.method == 'POST':
        form = ProjectCreateForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                project = form.save(commit=False)
                project.assigned_pm = request.user.profile
                project.status = 'Draft'
                project.created_by = request.user
                project.save()

            messages.success(request, f"Project {project.project_id} created successfully.")
            return redirect('project_overview', project_id=project.project_id)
    else:
        form = ProjectCreateForm(initial={'project_type': 'Residential'})

    return render(request, 'projects/project_form.html', {
        'form': form,
        'action': 'Create',
        'assigned_pm_display': request.user.get_full_name() or request.user.username,
    })


@login_required
def project_detail(request, project_id):
    """Merged into project_overview — redirect all traffic there."""
    return redirect('project_overview', project_id=project_id)


@login_required
@role_required(['Admin'])
def project_delete(request, project_id):
    """Soft-delete a project. Admin only, POST only."""
    if request.method != 'POST':
        return redirect('project_overview', project_id=project_id)
    project = get_object_or_404(Project, project_id=project_id, is_deleted=False)
    project.is_deleted = True
    project.deleted_at = timezone.now()
    project.save(update_fields=['is_deleted', 'deleted_at'])
    messages.success(request, f"Project {project.project_id} ({project.customer_name}) has been deleted.")
    return redirect('project_list')


@login_required
@role_required(['PM', 'Project Coordinator'])
def project_edit(request, project_id):
    """
    Edit a Draft project's fields. Active+ projects are locked — edit is blocked
    with a warning redirect. Access: assigned PM only, Draft status only.
    """
    project = get_object_or_404(Project, project_id=project_id)

    if not _pm_owns_project(request, project):
        raise Http404

    if project.status != 'Draft':
        messages.warning(request, 'Active projects cannot be edited.')
        return redirect('project_overview', project_id=project.project_id)

    if request.method == 'POST':
        form = ProjectEditForm(request.POST, instance=project)
        if form.is_valid():
            form.save()
            messages.success(request, f"Project {project.project_id} updated successfully.")
            return redirect('project_overview', project_id=project.project_id)
    else:
        form = ProjectEditForm(instance=project)

    return render(request, 'projects/project_form.html', {
        'form':               form,
        'action':             'Edit',
        'project':            project,
        'assigned_pm_display': project.assigned_pm.user.get_full_name() or project.assigned_pm.user.username,
    })


def _fmt_field_value(value):
    """Stringify a Project field value for ProjectFieldEditLog storage; None -> ''.
    Decimals render like '5.00', dates like '2026-07-21', so the audit log diffs
    every field type uniformly as text."""
    return '' if value is None else str(value)


def _render_field_edit_modal_hx(request, project, form):
    """Render the post-activation field-edit modal body — for the GET load and for
    re-rendering with inline errors on a failed POST. Swapped into
    #fieldEditModalContent (mirrors the task-form modal convention)."""
    return render(request, 'projects/partials/_project_field_edit_form.html', {
        'project': project,
        'form':    form,
    })


@login_required
def project_field_edit(request, project_id):
    """Post-activation edit of a project's capacity / contract value / target
    commissioning date. Separate from project_edit (Draft-only, full form): this path
    is for NON-Draft projects, applies immediately with no approval gate, and writes
    one ProjectFieldEditLog row per CHANGED field (no-op fields are skipped).

    Editing target_commissioning_date does NOT trigger due-date cascade recalculation
    (recalculate_from_task / calculate_due_dates are untouched) — the contractual
    target stays decoupled from the operational task schedule by design.

    Authorization is user_can_manage_project() ALONE (assigned PM or Coordinator) —
    no role-string / assigned_pm comparison here. GET returns the modal body; POST
    validates, saves, logs, and (HTMX) OOB-swaps the header pills.
    """
    project = get_object_or_404(Project, project_id=project_id)

    # Server-side authority check — the button is only rendered for managers, but the
    # POST handler must not rely on that (see spec §5).
    if not user_can_manage_project(request.user, project):
        raise Http404

    # Draft projects must use the full edit flow (project_edit); reject here so the two
    # paths stay disjoint and a Draft can never reach this no-cascade branch (spec §5).
    if project.status == 'Draft':
        messages.warning(request, 'Draft projects are edited from the full edit form.')
        if _is_hx(request):
            return render(request, 'projects/partials/_hx_messages.html')
        return redirect('project_overview', project_id=project.project_id)

    if request.method == 'POST':
        # Capture originals BEFORE binding: ModelForm.is_valid() mutates `project` in
        # place (construct_instance in _post_clean), so getattr after save() would
        # already return the NEW values.
        old_values = {name: getattr(project, name) for name, _ in ProjectFieldEditLog.FIELD_CHOICES}
        form = PostActivationFieldEditForm(request.POST, instance=project)
        if form.is_valid():
            profile = getattr(request.user, 'profile', None)
            reason  = form.cleaned_data.get('reason', '')
            changed = []
            with transaction.atomic():
                form.save()
                for name, _ in ProjectFieldEditLog.FIELD_CHOICES:
                    new_val = getattr(project, name)
                    if old_values[name] != new_val:  # skip no-op edits — no audit noise
                        ProjectFieldEditLog.objects.create(
                            project=project,
                            field_name=name,
                            old_value=_fmt_field_value(old_values[name]),
                            new_value=_fmt_field_value(new_val),
                            edited_by=profile,
                            reason=reason,
                        )
                        changed.append(name)
                if changed:
                    log_activity(
                        project, profile,
                        f"Edited project fields: {', '.join(changed)}",
                        entity_type='Project', entity_id=project.pk,
                    )
            if changed:
                messages.success(request, f"Updated {len(changed)} field{'' if len(changed) == 1 else 's'}.")
            else:
                messages.info(request, 'No changes to save.')
            if _is_hx(request):
                resp = render(request, 'projects/partials/_project_field_edit_success.html', {
                    'project': project,
                    'oob':     True,
                })
                resp['HX-Trigger'] = 'fieldEditDone'  # closes the modal client-side
                return resp
            return redirect('project_overview', project_id=project.project_id)

        # Invalid submit → re-render the modal with inline errors (HTMX), else redirect.
        if _is_hx(request):
            return _render_field_edit_modal_hx(request, project, form)
        messages.error(request, 'Please correct the errors and try again.')
        return redirect('project_overview', project_id=project.project_id)

    # GET → modal body prefilled from the current instance values.
    form = PostActivationFieldEditForm(instance=project)
    return _render_field_edit_modal_hx(request, project, form)


@login_required
@role_required(['PM', 'Project Coordinator'])
def project_activate(request, project_id):
    """
    Activate a Draft project: sets status=Active, stamps activated_at, assigns
    the designer, attaches the Residential template (phases + tasks), and
    creates M1/M2/M3 milestones.
    A designer must be selected before activation — this is what makes the
    project visible on the Design dashboard once Design tasks are seeded.
    All DB writes are wrapped in transaction.atomic() — a failure rolls back everything.
    Access: assigned PM only. POST only.
    """
    if request.method != 'POST':
        return redirect('project_overview', project_id=project_id)

    project = get_object_or_404(Project, project_id=project_id)

    if not _pm_owns_project(request, project):
        raise Http404

    # Status check before any DB write — avoids partial state if already active
    if project.status != 'Draft':
        messages.warning(request, 'Project is already active.')
        return redirect('project_overview', project_id=project.project_id)

    assigned_design_id = request.POST.get('assigned_design_id', '').strip()
    if not assigned_design_id:
        messages.error(request, 'Please select a Designer before activating.')
        return redirect('project_overview', project_id=project.project_id)

    designer = get_object_or_404(UserProfile, pk=assigned_design_id, role='Design', is_active=True)

    with transaction.atomic():
        project.assigned_design = designer
        project.status = 'Active'
        project.activated_at = timezone.now()
        project.save()

        if project.project_type == 'Residential':
            attach_residential_template(project)

        milestone_defaults = [
            ('M1', 'On Survey Completion'),
            ('M2', 'On Material Supply'),
            ('M3', 'On Commissioning'),
        ]
        for name, desc in milestone_defaults:
            PaymentMilestone.objects.create(
                project=project,
                milestone_name=name,
                milestone_description=desc,
                created_by=request.user.profile,
            )

    # Log project activation after the transaction commits
    log_activity(project, request.user.profile, f"Activated project: {project.project_id}", entity_type='Project', entity_id=project.pk)

    if project.project_type == 'Residential':
        messages.success(request, 'Project activated. 53 tasks created. Set the first task due date to calculate all dates.')
    else:
        messages.success(request, 'Project activated. Add tasks manually using Add Task.')

    return redirect('project_overview', project_id=project.project_id)


@login_required
@role_required(['PM', 'Project Coordinator'])
def project_recalculate_dates(request, project_id):
    """
    Recalculate all task due dates from project.activated_at using the duration_days chain.
    Intended as a bulk reset; PM normally sets dates task-by-task via task_set_due_date.
    Access: assigned PM only. POST only.
    """
    if request.method != 'POST':
        return redirect('project_overview', project_id=project_id)

    project = get_object_or_404(Project, project_id=project_id)

    if not _pm_owns_project(request, project):
        raise Http404

    if project.status == 'Draft':
        messages.warning(request, 'Project must be activated before calculating due dates.')
        return redirect('project_overview', project_id=project.project_id)

    if not project.activated_at:
        messages.warning(request, 'Project has no activation date — cannot calculate due dates.')
        return redirect('project_overview', project_id=project.project_id)

    calculate_due_dates(project, user=request.user)
    messages.success(request, 'Due dates recalculated from activation date.')
    return redirect('project_overview', project_id=project.project_id)


@login_required
def enable_cascade_scheduling(request, project_id):
    """
    Irreversibly enable cascading scheduling for a project.
    POST only, PM only, feature gate must be ON.
    Once set to True, cascade_scheduling cannot be reverted.
    """
    if request.method != 'POST':
        return redirect('project_overview', project_id=project_id)

    project = get_object_or_404(Project, project_id=project_id)

    if not _pm_owns_project(request, project):
        raise PermissionDenied

    settings_obj = SystemSettings.get()
    if not settings_obj.cascade_scheduling_enabled:
        messages.error(request, 'Cascading scheduling is not enabled by Admin.')
        return redirect('project_overview', project_id=project_id)

    # Idempotent — already on, nothing to do
    if project.cascade_scheduling:
        return redirect('project_overview', project_id=project_id)

    project.cascade_scheduling = True
    project.save(update_fields=['cascade_scheduling'])

    # Trigger full recalculation if project has an activation date
    if project.activated_at:
        calculate_due_dates(project, user=request.user)

    log_activity(
        project=project,
        actor=request.user.profile,
        action=f"Cascading scheduling enabled for project '{project.project_id}'",
        entity_type='Project',
        entity_id=project.id,
    )

    messages.success(request, 'Cascading scheduling is now active for this project.')
    return redirect('project_overview', project_id=project.project_id)


@login_required
@role_required(['PM', 'Project Coordinator'])
def task_add(request, project_id):
    """
    Add a single manual task to an active project. Only allowed when status=Active.
    task_order is set to last+1 within the chosen phase.
    Access: assigned PM only.
    """
    project = get_object_or_404(Project, project_id=project_id)

    if not _pm_owns_project(request, project):
        raise Http404

    hx = _is_hx(request)

    if project.status != 'Active':
        messages.warning(request, 'Tasks can only be added to active projects.')
        if hx:
            # Never redirect an hx request into the modal — surface the warning inside it.
            return render(request, 'projects/task_add_modal.html', {'project': project, 'blocked': True})
        return redirect('project_overview', project_id=project.project_id)

    if request.method == 'POST':
        form = TaskAddForm(request.POST, project=project)
        if form.is_valid():
            cd = form.cleaned_data
            phase = cd['phase']
            last_order = phase.tasks.aggregate(Max('task_order'))['task_order__max'] or 0
            Task.objects.create(
                phase=phase,
                task_name=cd['task_name'],
                task_order=last_order + 1,
                assigned_role=cd['assigned_role'],
                due_date=cd['due_date'],
            )
            messages.success(request, f"Task '{cd['task_name']}' added successfully.")
            if hx:
                # New row(s) + updated count swapped OOB into the chosen phase; modal closes.
                return _render_task_add_success_hx(request, project, phase)
            return redirect('project_overview', project_id=project.project_id)
        # Invalid POST — re-render the modal body with inline field errors (modal stays open).
        if hx:
            return render(request, 'projects/task_add_modal.html', {'form': form, 'project': project})
    else:
        form = TaskAddForm(project=project)
        if hx:
            # hx-get → load the form into the modal instead of the standalone page.
            return render(request, 'projects/task_add_modal.html', {'form': form, 'project': project})

    return render(request, 'projects/task_add_form.html', {
        'form':    form,
        'project': project,
    })


# ---------------------------------------------------------------------------
# Programs (OPEX tender / multi-site CAPEX contract parent)
#
# Foundation only: Program CRUD + compute-live rollup reporting. Child sites are
# ordinary Projects linked via Project.program. This section touches NO existing
# project list / dashboard view. OPEX site-creation plumbing lives in its own
# section (opex_site_create) — see Prompt 2.
# ---------------------------------------------------------------------------

def _can_access_program(request, program):
    """View-layer access gate for a SPECIFIC Program (mirrors _pm_owns_project's role
    for projects). Admin / CEO reach every Program (role_required already limits who
    gets here). A PM reaches a Program they manage — canonical user_can_manage_program
    (authority over any child site) — OR one they created (covers a brand-new empty
    Program whose creator has no site to derive authority from yet). The created_by
    fallback is a creator check, NOT a re-implementation of PM/coordinator comparison."""
    role = _get_user_role(request)
    if role in ('Admin', 'CEO'):
        return True
    if user_can_manage_program(request.user, program):
        return True
    return program.created_by_id == request.user.id


@login_required
@role_required(['Admin', 'PM', 'CEO'])
def program_list(request):
    """Role-scoped list of Programs with a live site-count-by-stage rollup.

    Admin / CEO see every non-deleted Program; a PM sees only Programs they can access
    (manage a child site, or created it). Rollup counts are annotated on the queryset
    (single query — no N+1 across Programs), computed live from child-site status.
    """
    programs = (
        Program.objects.filter(is_deleted=False)
        .annotate(**program_rollup_annotations())
        .prefetch_related('sites', 'sites__coordinators', 'sites__assigned_pm')
    )
    role = _get_user_role(request)
    if role == 'PM':
        programs = [p for p in programs if _can_access_program(request, p)]
    else:
        programs = list(programs)

    # Attach a display-ready stage breakdown to each program from the annotated
    # counts (dynamic attr access isn't clean in templates). Only non-zero stages.
    status_keys = [
        (status, 'site_' + status.lower().replace(' ', '_'))
        for status, _ in Project.STATUS_CHOICES
    ]
    for p in programs:
        p.rollup_badges = [
            (status, getattr(p, key)) for status, key in status_keys if getattr(p, key)
        ]

    return render(request, 'projects/program_list.html', {
        'programs':   programs,
        'can_create': role in ('Admin', 'PM'),
    })


@login_required
@role_required(['Admin', 'PM', 'CEO'])
def program_detail(request, pk):
    """Program detail: live aggregate rollup vs. planned targets, plus the child-site
    list (each linking to the existing project_overview). Read-only for CEO."""
    program = get_object_or_404(Program, pk=pk, is_deleted=False)
    if not _can_access_program(request, program):
        raise Http404

    rollup = get_program_rollup(program)
    sites = (
        program.sites.filter(is_deleted=False)
        .select_related('assigned_pm', 'assigned_pm__user')
        .order_by('project_id')
    )
    role = _get_user_role(request)
    return render(request, 'projects/program_detail.html', {
        'program':   program,
        'rollup':    rollup,
        'sites':     sites,
        'can_edit':  role in ('Admin', 'PM'),
        'can_delete': role == 'Admin',
    })


@login_required
@role_required(['Admin', 'PM'])
def program_create(request):
    """Create a Program (OPEX tender or multi-site CAPEX contract). Admin/PM only."""
    if request.method == 'POST':
        form = ProgramForm(request.POST)
        if form.is_valid():
            program = form.save(commit=False)
            program.created_by = request.user
            program.save()
            log_activity(
                None, getattr(request.user, 'profile', None),
                f"Created {program.program_type} Program: {program.name}",
                entity_type='Program', entity_id=program.pk, action_code='program_created',
            )
            messages.success(request, f"Program '{program.name}' created successfully.")
            return redirect('program_detail', pk=program.pk)
    else:
        form = ProgramForm()
    return render(request, 'projects/program_form.html', {'form': form, 'action': 'Create'})


@login_required
@role_required(['Admin', 'PM'])
def program_edit(request, pk):
    """Edit a Program. Access mirrors program_detail. Each CHANGED field is audit-logged
    (one ActivityLog row per field) following the post-activation contract-edit pattern —
    editing a Program must never silently orphan/rewrite data. project_id values on
    existing child sites are immutable and are NOT touched by any rename here."""
    program = get_object_or_404(Program, pk=pk, is_deleted=False)
    if not _can_access_program(request, program):
        raise Http404

    if request.method == 'POST':
        form = ProgramForm(request.POST, instance=program)
        if form.is_valid():
            changed = form.changed_data
            actor = getattr(request.user, 'profile', None)
            with transaction.atomic():
                program = form.save()
                for field in changed:
                    log_activity(
                        None, actor,
                        f"Edited Program '{program.name}' field: {field}",
                        entity_type='Program', entity_id=program.pk, action_code='program_edited',
                    )
            messages.success(request, f"Program '{program.name}' updated successfully.")
            return redirect('program_detail', pk=program.pk)
    else:
        form = ProgramForm(instance=program)
    return render(request, 'projects/program_form.html', {
        'form': form, 'action': 'Edit', 'program': program,
    })


@login_required
@role_required(['Admin'])
def program_delete(request, pk):
    """Soft-delete a Program. Admin only, POST only. BLOCKED while it still has any
    non-deleted child site — a Program must never cascade-delete or orphan its sites."""
    if request.method != 'POST':
        return redirect('program_detail', pk=pk)
    program = get_object_or_404(Program, pk=pk, is_deleted=False)

    active_sites = program.sites.filter(is_deleted=False).count()
    if active_sites:
        messages.error(
            request,
            f"Cannot delete Program '{program.name}': it still has {active_sites} "
            f"active site(s). Remove or reassign them first."
        )
        return redirect('program_detail', pk=program.pk)

    program.is_deleted = True
    program.deleted_at = timezone.now()
    program.save(update_fields=['is_deleted', 'deleted_at'])
    log_activity(
        None, getattr(request.user, 'profile', None),
        f"Deleted Program: {program.name}",
        entity_type='Program', entity_id=program.pk, action_code='program_deleted',
    )
    messages.success(request, f"Program '{program.name}' has been deleted.")
    return redirect('program_list')


def create_opex_site(program, data, creator, profile=None):
    """Create ONE OPEX site under an OPEX Program — request-independent core, extracted
    from opex_site_create so a future bulk-upload path can call it per-row without
    duplicating the field-setting/validation logic.

    Callers own access control: `program` MUST already be loaded, OPEX-typed, and
    access-checked (no role / _can_access_program checks happen here).

    Args:
        program: parent OPEX Program (already validated + access-checked by the caller).
        data:    form data (a QueryDict like request.POST, or a plain dict for bulk) whose
                 keys match OpexSiteForm's fields.
        creator: the User creating the site (stored as created_by).
        profile: creator's UserProfile, used for PM auto-assignment; a non-PM (or None)
                 profile leaves assigned_pm=None, matching the single-add behavior.

    Returns:
        (site, form). On success: (the saved Project, the valid bound form). On failure:
        (None, the bound form carrying validation errors) — the caller re-renders that
        same form instance so entered values and per-field errors are preserved. Bulk
        callers can read `form.errors` (a dict) off the returned form.

    Composes the site's globally-unique project_id `{short_tender_code}-{site_code}` and
    sets it EXPLICITLY before save() so generate_project_id()'s suffix-parser is bypassed
    entirely. The ID is STORED (not derived at read time), which is what makes it immutable
    across a later Program rename.
    """
    form = OpexSiteForm(data, program=program)
    if not form.is_valid():
        return None, form

    with transaction.atomic():
        site = form.save(commit=False)
        site.program = program
        site.project_type = 'OPEX'          # forced — never user-selectable here
        # customer_name is not a form field for OPEX — freeze it to the parent
        # tender's client_name at creation (store, don't compute; a later Program
        # rename must not retroactively change existing sites). Model width was
        # widened to 200 to match client_name so this can never overflow.
        site.customer_name = program.client_name
        site.status = 'Draft'
        # A PM creator owns the site; an Admin creator leaves it unassigned
        # (assign later via admin_assign_pm) — assigned_pm requires a PM profile.
        site.assigned_pm = profile if (profile and profile.role == 'PM') else None
        site.created_by = creator
        site.project_id = form.composed_project_id   # explicit → skips generate_project_id()
        site.save()
    log_activity(
        site, profile,
        f"Created OPEX site {site.project_id} under program {program.name}",
        entity_type='Project', entity_id=site.pk, action_code='opex_site_created',
    )
    return site, form


@login_required
@role_required(['Admin', 'PM'])
def opex_site_create(request, pk):
    """Create ONE OPEX site under an OPEX Program (Prompt 2 — dedicated plumbing).

    This is the ONLY forward path to a new OPEX project. Access mirrors the Program:
    Admin, or a PM who can access it. The creation core lives in create_opex_site() so
    a future bulk-upload path can reuse it.
    """
    program = get_object_or_404(Program, pk=pk, is_deleted=False, program_type='OPEX')
    if not _can_access_program(request, program):
        raise Http404

    if request.method == 'POST':
        profile = getattr(request.user, 'profile', None)
        site, form = create_opex_site(program, request.POST, request.user, profile=profile)
        if site is not None:
            messages.success(request, f"Site {site.project_id} created under {program.name}.")
            return redirect('program_detail', pk=program.pk)
    else:
        form = OpexSiteForm(program=program)

    return render(request, 'projects/opex_site_form.html', {'form': form, 'program': program})


# ===========================================================================
# OPEX bulk site upload (Prompt 3)
#
# Upload an Excel file of sites under one OPEX Program and create them in one
# ALL-OR-NOTHING batch, reusing create_opex_site() (and therefore OpexSiteForm)
# for every row so validation NEVER drifts from the single-add path.
#
# Flow is two requests: (1) upload -> parse + dry-run validate -> preview; the
# validated rows ride back to the browser as JSON in a hidden field. (2) confirm
# -> the exact previewed rows are re-validated and committed for real inside one
# outer transaction. A file only ever adds sites; multiple independent batches
# per Program are expected (running total vs planned_site_count is informational).
# ===========================================================================

# Excel column header (normalized: .strip().lower()) -> OpexSiteForm data key.
# Order here is the order columns are emitted into the downloadable template.
_BULK_COLUMNS = [
    ('Site Code',            'site_code',                True),
    ('Site In-Charge Name',  'customer_contact_person',  True),
    ('Site In-Charge Phone', 'customer_phone',           True),
    ('Site In-Charge Email', 'customer_email',           False),
    ('Site Address',         'site_address',             True),
    ('City',                 'city',                     True),
    ('State',                'state',                    False),
    ('Capacity (kW)',        'capacity_kw',              False),
]
_BULK_HEADER_TO_KEY = {h.strip().lower(): key for h, key, _req in _BULK_COLUMNS}
_BULK_REQUIRED_HEADERS = [h for h, _key, req in _BULK_COLUMNS if req]
_BULK_MAX_ROWS = 500   # soft guard against a runaway file timing out the request


class _DryRunRollback(Exception):
    """Sentinel used ONLY by _validate_site_row_dry_run to unwind a deliberately
    rolled-back transaction. Dedicated (never a bare Exception) so it is obvious in
    a traceback and can never be confused with a real failure."""
    pass


class _CommitAbort(Exception):
    """Sentinel raised inside the real commit loop when a row that passed dry-run
    fails at commit time (race / tampered payload). Rolls back the WHOLE batch so
    the all-or-nothing guarantee holds end-to-end, not just at dry-run."""
    def __init__(self, index, errors):
        self.index = index          # 1-based row number for the report
        self.errors = errors        # form.errors dict
        super().__init__()


def _validate_site_row_dry_run(program, data, creator, profile):
    """Validate ONE would-be site with the REAL create_opex_site path, then throw the
    result away — a dry run that reuses production logic instead of re-checking the
    form by hand (which would silently drift the moment create_opex_site changes).

    HOW IT WORKS: create_opex_site() only reaches the DB for a *valid* row — an invalid
    row returns (None, form) before its atomic block ever opens. So we wrap the call in
    our own transaction.atomic(); for a valid row create_opex_site performs the real
    INSERT plus its ActivityLog write inside that savepoint, and we immediately raise
    _DryRunRollback to force Django to roll every bit of it back. Nothing is persisted,
    yet the exact production validation + project_id composition ran.

    WHY THE EXCEPTION IS LOAD-BEARING: if a future "simplification" deletes the raise,
    the transaction commits and this preview quietly starts creating real sites on every
    keystroke of a preview. The raise is the whole rollback mechanism — do not remove it.
    Safe only because log_activity is pure-DB (no email/WhatsApp side-effect), so the
    rollback leaves no trace outside the transaction.

    Returns (form, is_valid). `form` carries form.errors for the preview when invalid.
    """
    captured = {}
    try:
        with transaction.atomic():
            site, form = create_opex_site(program, data, creator, profile=profile)
            captured['form'] = form
            captured['valid'] = site is not None
            if site is not None:
                raise _DryRunRollback()   # load-bearing: forces rollback of the real INSERT
    except _DryRunRollback:
        pass
    return captured['form'], captured['valid']


def _bulk_cell_to_str(value):
    """Excel cell -> trimmed string WITHOUT lossy coercion. openpyxl hands back numbers
    as int/float; a phone/site-code typed as a number must render as plain digits (no
    '.0'), but we never strip or reshape characters a user actually typed as text — bad
    values stay bad so OpexSiteForm can reject them loudly (spec §5: no silent coercion)."""
    if value is None:
        return ''
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _parse_bulk_workbook(uploaded_file):
    """Parse the uploaded .xlsx into (rows, extra_headers, error). Reads ONLY the 'Sites'
    sheet (the downloadable template keeps all guidance on a separate 'Instructions' sheet,
    so helper text can never be mistaken for a data row). Returns error!=None for a
    whole-file rejection (bad file, missing required column, empty, too many rows)."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception:
        return None, None, "Could not read the file — please upload the .xlsx template unchanged."

    ws = wb['Sites'] if 'Sites' in wb.sheetnames else wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not all_rows:
        return None, None, "The file is empty."

    header_cells = [_bulk_cell_to_str(c) for c in all_rows[0]]
    header_norm = [h.strip().lower() for h in header_cells]
    missing = [h for h in _BULK_REQUIRED_HEADERS if h.strip().lower() not in header_norm]
    if missing:
        return None, None, "Missing required column(s): " + ", ".join(missing) + "."

    # Columns present in the file but not part of the template — ignored, warned (not rejected).
    extra_headers = [header_cells[i] for i, hn in enumerate(header_norm)
                     if hn and hn not in _BULK_HEADER_TO_KEY]

    data_rows = all_rows[1:]
    # Drop wholly-blank trailing rows Excel loves to include.
    data_rows = [r for r in data_rows if any(_bulk_cell_to_str(c) for c in r)]
    if not data_rows:
        return None, None, "The file has headers but no site rows to import."
    if len(data_rows) > _BULK_MAX_ROWS:
        return None, None, (f"This file has {len(data_rows)} rows — the per-upload limit is "
                            f"{_BULK_MAX_ROWS}. Split it into smaller batches.")

    rows = []
    for raw in data_rows:
        data = {}
        for i, hn in enumerate(header_norm):
            key = _BULK_HEADER_TO_KEY.get(hn)
            if key and i < len(raw):
                data[key] = _bulk_cell_to_str(raw[i])
        rows.append(data)
    return rows, extra_headers, None


def _bulk_infile_duplicate_indices(rows):
    """1-based row numbers whose site_code collides with another row IN THE SAME FILE.
    Normalizes with the SAME normalize_program_code the form uses, so 's045' / 'S-045'
    are seen as the duplicates they are. This is the ONE check dry-run cannot make (each
    row rolls back before the next runs), so it must happen here (spec §4 / §5)."""
    seen = {}
    for idx, data in enumerate(rows, start=1):
        code = normalize_program_code(data.get('site_code'))
        if code:
            seen.setdefault(code, []).append(idx)
    dupes = set()
    for code, idxs in seen.items():
        if len(idxs) > 1:
            dupes.update(idxs)
    return dupes


@login_required
@role_required(['Admin', 'PM'])
def opex_site_bulk_upload(request, pk):
    """Bulk-create OPEX sites under one Program from an uploaded .xlsx (Admin/PM, OPEX
    only — same gate as opex_site_create). Two POST phases keyed by the 'phase' field:
    'preview' (a file upload -> validate + preview) and 'commit' (confirm the previewed
    JSON -> create for real, all-or-nothing)."""
    program = get_object_or_404(Program, pk=pk, is_deleted=False, program_type='OPEX')
    if not _can_access_program(request, program):
        raise Http404

    profile = getattr(request.user, 'profile', None)
    rollup = get_program_rollup(program)
    existing_count = rollup['total']
    planned = program.planned_site_count

    ctx = {
        'program': program,
        'existing_count': existing_count,
        'planned': planned,
        'stage': 'upload',
    }

    phase = request.POST.get('phase') if request.method == 'POST' else None

    # ---- Phase: PREVIEW (parse + dry-run validate the uploaded file) ----
    if phase == 'preview':
        uploaded = request.FILES.get('file')
        if not uploaded:
            ctx['file_error'] = "Please choose a file to upload."
            return render(request, 'projects/opex_site_bulk_upload.html', ctx)

        rows, extra_headers, file_error = _parse_bulk_workbook(uploaded)
        if file_error:
            ctx['file_error'] = file_error
            return render(request, 'projects/opex_site_bulk_upload.html', ctx)

        dup_indices = _bulk_infile_duplicate_indices(rows)
        results = []
        all_valid = True
        for idx, data in enumerate(rows, start=1):
            form, valid = _validate_site_row_dry_run(program, data, request.user, profile)
            errors = {field: list(msgs) for field, msgs in form.errors.items()}
            if idx in dup_indices:
                errors.setdefault('site_code', [])
                errors['site_code'].append("Duplicate site code within this file.")
                valid = False
            if not valid:
                all_valid = False
            results.append({
                'index': idx,
                'data': data,
                'site_code': data.get('site_code', ''),
                'name': data.get('customer_contact_person', ''),
                'city': data.get('city', ''),
                'errors': errors,
                'valid': valid,
            })

        after_count = existing_count + len(rows)
        ctx.update({
            'stage': 'preview',
            'results': results,
            'total_rows': len(rows),
            'valid_count': sum(1 for r in results if r['valid']),
            'invalid_count': sum(1 for r in results if not r['valid']),
            'all_valid': all_valid,
            'extra_headers': extra_headers,
            'after_count': after_count,
            'exceeds_planned': bool(planned) and after_count > planned,
            # Only valid-batch rows ride forward; all_valid is required to show Confirm.
            'rows_json': json.dumps([r['data'] for r in results]) if all_valid else '',
        })
        return render(request, 'projects/opex_site_bulk_upload.html', ctx)

    # ---- Phase: COMMIT (create the previewed rows for real, all-or-nothing) ----
    if phase == 'commit':
        try:
            rows = json.loads(request.POST.get('rows_json') or '[]')
        except (ValueError, TypeError):
            rows = None
        if not rows or not isinstance(rows, list):
            ctx['file_error'] = "Nothing to create — please upload and preview a file first."
            return render(request, 'projects/opex_site_bulk_upload.html', ctx)

        # Re-check in-file duplicates on the confirmed payload (defends against a tampered
        # hidden field); the per-row create below re-runs full validation for real.
        dup_indices = _bulk_infile_duplicate_indices(rows)

        created = []
        commit_error = None
        try:
            with transaction.atomic():
                if dup_indices:
                    raise _CommitAbort(min(dup_indices),
                                       {'site_code': ['Duplicate site code within this file.']})
                for idx, data in enumerate(rows, start=1):
                    site, form = create_opex_site(program, data, request.user, profile=profile)
                    if site is None:
                        raise _CommitAbort(idx, {f: list(m) for f, m in form.errors.items()})
                    created.append(site.project_id)
                # Batch-level audit entry (in addition to create_opex_site's per-site logs).
                # Inside the atomic, so a failed batch rolls this back too.
                log_activity(
                    None, profile,
                    f"Bulk uploaded {len(created)} sites to program {program.name}",
                    entity_type='Program', entity_id=program.pk,
                    action_code='opex_sites_bulk_created',
                )
        except _CommitAbort as abort:
            created = []
            commit_error = {'index': abort.index, 'errors': abort.errors}
        except IntegrityError:
            created = []
            commit_error = {'index': None, 'errors': {
                '__all__': ['A site code was taken by another user between preview and confirm. '
                            'Nothing was created — please re-upload to refresh the check.']}}

        if commit_error is not None:
            ctx.update({'stage': 'result', 'success': False, 'commit_error': commit_error})
            return render(request, 'projects/opex_site_bulk_upload.html', ctx)

        # Success — recompute the running total from the DB post-commit.
        new_total = get_program_rollup(program)['total']
        messages.success(request, f"{len(created)} sites created under {program.name}.")
        ctx.update({
            'stage': 'result', 'success': True,
            'created': created, 'created_count': len(created),
            'new_total': new_total, 'planned': planned,
        })
        return render(request, 'projects/opex_site_bulk_upload.html', ctx)

    # GET (or unknown phase) — the initial upload screen.
    return render(request, 'projects/opex_site_bulk_upload.html', ctx)


@login_required
@role_required(['Admin', 'PM'])
def opex_site_bulk_template(request, pk):
    """Download the .xlsx template for this Program's bulk upload: a 'Sites' sheet with
    just the header row (what the parser reads), plus an 'Instructions' sheet holding the
    field guide, phone-format rule, and a worked example — kept OFF the data sheet so no
    guidance row can ever be imported as a real site."""
    program = get_object_or_404(Program, pk=pk, is_deleted=False, program_type='OPEX')
    if not _can_access_program(request, program):
        raise Http404

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sites'
    ws.append([h for h, _key, _req in _BULK_COLUMNS])

    info = wb.create_sheet('Instructions')
    info.append(['OPEX Bulk Site Upload — Instructions'])
    info.append([])
    info.append([f'Program: {program.name}  (tender code {program.short_tender_code})'])
    info.append(['Each site ID is composed as  <tender code>-<Site Code>.'])
    info.append([])
    info.append(['Fill one row per site on the "Sites" sheet. Do not rename the header row.'])
    info.append([])
    info.append(['Column', 'Required?', 'Notes'])
    _notes = {
        'site_code': 'Uppercase letters/digits (e.g. S045). Unique within this tender.',
        'customer_contact_person': 'Site In-Charge full name.',
        'customer_phone': '10 digits, no country code, must start with 6, 7, 8, or 9.',
        'customer_email': 'Optional.',
        'site_address': 'Full site address.',
        'city': 'City.',
        'state': 'Optional — defaults to "Uttar Pradesh" if left blank.',
        'capacity_kw': 'Optional. Number, e.g. 150.00.',
    }
    for header, key, req in _BULK_COLUMNS:
        info.append([header, 'Yes' if req else 'No', _notes.get(key, '')])
    info.append([])
    info.append(['Example (do NOT copy this onto the Sites sheet as-is):'])
    info.append([h for h, _key, _req in _BULK_COLUMNS])
    info.append(['S045', 'Ravi Kumar', '9876543210', 'ravi@example.com',
                 '12 Grid Lane, Sector 5', 'Delhi', 'Delhi', '150.00'])

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = (
        f'attachment; filename="opex_sites_{program.short_tender_code or program.pk}.xlsx"')
    wb.save(resp)
    return resp


# ---------------------------------------------------------------------------
# HTMX partial-response helpers (reload elimination — Prompt B)
#
# Each helper re-derives exactly the per-row / per-section context the full-page
# render uses, so a single swapped fragment is byte-identical to what the page
# would have produced. None of these change server-side behaviour for a normal
# (non-HTMX) request — the callers branch on _is_hx() and fall through to the
# original redirect/full-render path when HX-Request is absent.
# ---------------------------------------------------------------------------

def _is_hx(request):
    """True when the request was issued by HTMX (so we return a partial, not a redirect)."""
    return request.headers.get('HX-Request') == 'true'


def _gate_task_pk(project):
    """PK of the first task of the first phase — the only row that can carry the
    BD milestone-amount gate. Reproduces the template's
    `forloop.parentloop.first and forloop.first` check when a single row is
    re-rendered outside the phase loop."""
    first_phase = (
        ProjectPhase.objects.filter(project=project).order_by('phase_order').first()
    )
    if not first_phase:
        return None
    first_task = Task.objects.filter(phase=first_phase).order_by('task_order').first()
    return first_task.pk if first_task else None


def _render_task_row_hx(request, project, task, oob_tasks=None):
    """Render the HTMX task-row response for project_overview (#1/#3/#5):
    the primary row (swapped into #task-row-<pk>), optional out-of-band cascade
    rows, and the OOB flash-message fragment. Recomputes the same per-row
    permission context the page uses so role-gating is identical to a full render."""
    profile = getattr(request.user, 'profile', None)
    role    = getattr(profile, 'role', None)
    _PROFILE_TO_TASK_ROLE = {'BD': 'BD / Sales'}
    user_task_role = _PROFILE_TO_TASK_ROLE.get(role, role)
    is_assigned_pm = _pm_owns_project(request, project)
    gate_pk        = _gate_task_pk(project)
    return render(request, 'projects/partials/_task_row_response.html', {
        'project':             project,
        'row_task':            task,
        'oob_tasks':           oob_tasks or [],
        'primary_is_gate':     task.pk == gate_pk,
        'is_assigned_pm':      is_assigned_pm,
        'user_task_role':      user_task_role,
        'role':                role,
        'task_status_choices': Task.STATUS_CHOICES,
    })


def _render_task_status_hx(request, project, task):
    """Render the HTMX status-block response for the task-detail page (#2)."""
    profile = getattr(request.user, 'profile', None)
    return render(request, 'projects/partials/_task_detail_status_response.html', {
        'project':             project,
        'task':                task,
        'is_assignee':         task.assigned_to is not None and task.assigned_to == profile,
        'task_status_choices': Task.STATUS_CHOICES,
    })


def _render_attachments_hx(request, project, task):
    """Render the HTMX attachment-list response (#7 upload / #8 delete). Uses the
    same is_deleted=False filter the task-detail page uses so the swapped list is
    identical to a full render."""
    profile = getattr(request.user, 'profile', None)
    return render(request, 'projects/partials/_task_attachments_response.html', {
        'project':      project,
        'task':         task,
        'attachments':  task.attachments.filter(is_deleted=False),
        'user_profile': profile,
    })


def _render_comments_hx(request, project, task):
    """Render the HTMX comment-thread response (#9). Mirrors the task-detail
    queryset (top-level comments with prefetched replies)."""
    profile = getattr(request.user, 'profile', None)
    task_comments = (
        Comment.objects.filter(task=task, parent=None)
        .select_related('author__user')
        .prefetch_related(
            Prefetch('replies', queryset=Comment.objects.select_related('author__user'))
        )
    )
    return render(request, 'projects/partials/_task_comments_response.html', {
        'project':       project,
        'task':          task,
        'task_comments': task_comments,
        'user_profile':  profile,
    })


def _checklist_for_task(task, project):
    """Resolve the active Checklist assigned to this task via ChecklistTaskLink
    (task_name + project_type), or None. Inactive checklists are treated as unassigned."""
    link = (ChecklistTaskLink.objects
            .select_related('checklist')
            .filter(task_name=task.task_name, project_type=project.project_type)
            .first())
    if link is None or not link.checklist.is_active:
        return None
    return link.checklist


def _checklist_context(request, project, task):
    """Build the shared context for the checklist section — used by the full task-detail
    render, the HTMX response partial, and any view that swaps #checklistSection. Items come
    from the Checklist linked to this (task_name, project_type); per-item completion state is
    looked up per (item, task) so each task instance completes independently. Keeps the
    permission flag computed in exactly one place."""
    profile = getattr(request.user, 'profile', None)
    checklist = _checklist_for_task(task, project)
    items = list(checklist.items.all()) if checklist else []
    completions = {}
    if items:
        completions = {
            c.item_id: c
            for c in ChecklistItemCompletion.objects.filter(task=task, item__in=items)
                                                    .select_related('checked_by')
        }
    rows = [{'item': it, 'completion': completions.get(it.id)} for it in items]
    return {
        'project':            project,
        'task':               task,
        'checklist':          checklist,
        'checklist_rows':     rows,
        'checklist_items':    items,   # truthiness + count badge
        'user_profile':       profile,
        'can_complete_items': _user_can_complete_checklist_item(request.user, task, project),
    }


def _render_checklist_hx(request, project, task):
    """Render the HTMX checklist response — swaps the item list into #checklistSection,
    updates the header count badge out-of-band, and surfaces flash messages inline."""
    return render(request, 'projects/partials/_checklist_response.html',
                  _checklist_context(request, project, task))


def _render_task_assign_design_success_hx(request, project, task):
    """#4 success: OOB-swap the updated row into its existing #task-row-<pk> and
    close the modal (taskFormDone trigger). Recomputes the same per-row context
    the page uses, from the requesting user's perspective (a Design Head need not
    be the PM)."""
    profile = getattr(request.user, 'profile', None)
    role    = getattr(profile, 'role', None)
    _PROFILE_TO_TASK_ROLE = {'BD': 'BD / Sales'}
    resp = render(request, 'projects/partials/_task_row_modal_success.html', {
        'project':             project,
        'row_task':            task,
        'primary_is_gate':     task.pk == _gate_task_pk(project),
        'is_assigned_pm':      _pm_owns_project(request, project),
        'user_task_role':      _PROFILE_TO_TASK_ROLE.get(role, role),
        'role':                role,
        'task_status_choices': Task.STATUS_CHOICES,
    })
    resp['HX-Trigger'] = 'taskFormDone'
    return resp


def _render_task_add_success_hx(request, project, phase):
    """#6 success: OOB-swap the chosen phase's <tbody> and task-count header from
    freshly-queried server truth, and close the modal (taskFormDone trigger).
    phase.tasks uses Task Meta ordering ['task_order'] so the re-render matches the
    page order and correctly places the new row."""
    profile = getattr(request.user, 'profile', None)
    role    = getattr(profile, 'role', None)
    _PROFILE_TO_TASK_ROLE = {'BD': 'BD / Sales'}
    phase_tasks = list(phase.tasks.all())
    resp = render(request, 'projects/partials/_task_add_success.html', {
        'project':             project,
        'phase':               phase,
        'phase_tasks':         phase_tasks,
        'phase_count':         len(phase_tasks),
        'gate_task_pk':        _gate_task_pk(project),
        'is_assigned_pm':      _pm_owns_project(request, project),
        'user_task_role':      _PROFILE_TO_TASK_ROLE.get(role, role),
        'role':                role,
        'task_status_choices': Task.STATUS_CHOICES,
    })
    resp['HX-Trigger'] = 'taskFormDone'
    return resp


@login_required
def task_status_update(request, project_id, task_id):
    """
    Update a task's status. Enforces a transition table so invalid moves are rejected.
    The assigned role or the project PM may update. Blocking a task requires a title
    for a new Issue — the issue is auto-created and linked to the task.
    filter().update() used instead of .save() to prevent race condition on concurrent
    status changes from two users.
    Access: task's assigned_role or project PM. POST only.
    """
    if request.method != 'POST':
        return redirect('project_overview', project_id=project_id)

    project = get_object_or_404(Project, project_id=project_id)
    task = get_object_or_404(Task, pk=task_id, phase__project=project)

    if task.assigned_to is None:
        # HTMX: revert the (disabled-select edge) change and surface the error inline
        # instead of rendering raw JSON into the row target.
        if _is_hx(request):
            messages.error(request, 'Task is unassigned. Assign it before changing status.')
            return _render_task_row_hx(request, project, task)
        return JsonResponse({
            'success': False,
            'error': 'Task is unassigned. Assign it before changing status.'
        }, status=400)

    # Permission check before any DB write
    try:
        user_role = request.user.profile.role
    except Exception:
        user_role = None

    is_pm = _pm_owns_project(request, project)

    # Task.BD = 'BD / Sales' but UserProfile stores 'BD' — normalise before comparison
    _PROFILE_TO_TASK_ROLE = {'BD': 'BD / Sales'}
    normalised_user_role = _PROFILE_TO_TASK_ROLE.get(user_role, user_role)

    if normalised_user_role != task.assigned_role and not is_pm:
        if _is_hx(request):
            messages.error(request, 'You do not have permission to change this task.')
            return _render_task_row_hx(request, project, task)
        return HttpResponseForbidden()

    new_status = request.POST.get('status', '').strip()
    valid_statuses = {s[0] for s in Task.STATUS_CHOICES}
    if new_status not in valid_statuses:
        messages.error(request, 'Invalid status value.')
        if _is_hx(request):
            return _render_task_row_hx(request, project, task)
        return redirect('project_overview', project_id=project.project_id)

    # State machine: defines allowed next states for each current state.
    # DONE can only go to BLOCKED (not back to In Progress) — prevents gaming completion.
    VALID_TRANSITIONS = {
        Task.NOT_STARTED: {Task.IN_PROGRESS, Task.BLOCKED, Task.DONE},
        Task.IN_PROGRESS: {Task.DONE, Task.BLOCKED},
        Task.BLOCKED:     {Task.IN_PROGRESS, Task.BLOCKED},
        Task.DONE:        {Task.BLOCKED},
    }

    allowed = VALID_TRANSITIONS.get(task.status, set())
    if new_status not in allowed:
        messages.error(
            request,
            f"Cannot move task from '{task.status}' to '{new_status}'."
        )
        if _is_hx(request):
            return _render_task_row_hx(request, project, task)
        return redirect('project_overview', project_id=project.project_id)

    # Finance can supply due_date inline when switching to In Progress — save before the guard
    _due_date_str = request.POST.get('due_date', '').strip()
    if _due_date_str and new_status == Task.IN_PROGRESS and not task.due_date:
        try:
            _parsed_due = date.fromisoformat(_due_date_str)
            Task.objects.filter(pk=task.pk).update(due_date=_parsed_due)
            task.due_date = _parsed_due
        except ValueError:
            pass

    # Server-side guard: In Progress requires a due date
    if new_status == Task.IN_PROGRESS and not task.due_date:
        messages.warning(request, 'Please set a due date before marking this task as In Progress.')
        if _is_hx(request):
            return _render_task_row_hx(request, project, task)
        return redirect('project_overview', project_id=project.project_id)

    update_kwargs = {'status': new_status}
    if new_status == Task.DONE:
        update_kwargs['completed_at'] = timezone.now()
    # Track when a task first becomes blocked so the CEO aged-KPI can measure how long it has been stuck
    if new_status == Task.BLOCKED and task.status != Task.BLOCKED:
        update_kwargs['blocked_since'] = timezone.now()
    # Clear on un-block so any future re-block re-ages from zero, not from the original block date
    elif new_status != Task.BLOCKED and task.status == Task.BLOCKED:
        update_kwargs['blocked_since'] = None

    # Blocked requires a stated blocking issue (fresh transition only)
    if new_status == Task.BLOCKED and task.status != Task.BLOCKED:
        block_issue_title = request.POST.get('block_issue_title', '').strip()
        if not block_issue_title:
            messages.error(request, 'Please state the blocking issue before marking this task as Blocked.')
            if _is_hx(request):
                return _render_task_row_hx(request, project, task)
            next_url = request.POST.get('next', '')
            if next_url and not _urlparse(next_url).netloc:
                return redirect(next_url)
            return redirect('project_overview', project_id=project.project_id)

        Task.objects.filter(pk=task.pk).update(**update_kwargs)

        block_severity = request.POST.get('block_issue_severity', Issue.HIGH)
        if block_severity not in dict(Issue.SEVERITY_CHOICES):
            block_severity = Issue.HIGH
        block_assignee = None
        block_assignee_id = request.POST.get('block_issue_assigned_to', '').strip()
        if block_assignee_id:
            try:
                block_assignee = UserProfile.objects.get(pk=block_assignee_id)
            except UserProfile.DoesNotExist:
                pass
        issue = Issue.objects.create(
            project=project,
            task=task,
            title=block_issue_title,
            description=request.POST.get('block_issue_description', '').strip(),
            severity=block_severity,
            status=Issue.OPEN,
            raised_by=request.user.profile,
            assigned_to=block_assignee,
        )
        log_activity(
            project, request.user.profile,
            f"Blocked task '{task.task_name}' — issue: {block_issue_title}",
            entity_type='Issue', entity_id=issue.pk, action_code='issue_created',
        )
        messages.success(request, f'Task blocked. Issue "{block_issue_title}" created.')
    else:
        Task.objects.filter(pk=task.pk).update(**update_kwargs)
        # Log status changes for all non-blocked transitions (blocked has its own log above)
        log_activity(project, request.user.profile, f"Changed task status to {new_status}: {task.task_name}", entity_type='Task', entity_id=task.pk,
                     action_code=f"task_status_{new_status.lower().replace(' ', '_')}")

        # Bidirectional sync: Finance confirmation tasks → PaymentMilestone Received.
        # Mapping by task name — names are fixed in the residential template.
        _FINANCE_TASK_TO_MILESTONE = {
            'Advance Payment Confirmation':      'M1',
            'Pre Dispatch Payment Confirmation': 'M2',
            '100% Payment Confirmation':         'M3',
        }
        if new_status == Task.DONE and task.task_name in _FINANCE_TASK_TO_MILESTONE:
            _ms_label  = _FINANCE_TASK_TO_MILESTONE[task.task_name]
            _ms_update = {'status': 'Received', 'received_date': date.today()}
            _ar_str    = request.POST.get('amount_received', '').strip()
            _vr_str    = request.POST.get('variance_reason', '').strip()
            if _ar_str:
                try:
                    _ms_update['amount_received'] = Decimal(_ar_str)
                except InvalidOperation:
                    pass
            if _vr_str:
                _ms_update['variance_reason'] = _vr_str
            try:
                _ms_updated = PaymentMilestone.objects.filter(
                    project=project,
                    milestone_name=_ms_label,
                    status__in=['Pending', 'Invoiced'],
                ).update(**_ms_update)
                # Attribution: this Finance-confirmation task can be completed by the PM
                # OR a Project Coordinator (drizzle-down authority), which auto-flips the
                # milestone to Received. Log WHO did it — by name and role — so Finance can
                # identify the specific person, never attributed generically to "PM".
                if _ms_updated:
                    _actor = request.user.profile
                    _actor_name = _actor.user.get_full_name() or _actor.user.username
                    log_activity(
                        project, _actor,
                        f"Milestone {_ms_label} auto-marked Received via completion of "
                        f"Finance task '{task.task_name}' by {_actor_name} ({_actor.role})",
                        entity_type='Milestone',
                    )
            except Exception:
                pass  # Non-critical — never block the task update

        # Payment milestone notification: task.is_payment_milestone is read from the pre-update
        # object — the flag never changes during a status update, so this is safe.
        if new_status == Task.DONE and task.is_payment_milestone:
            seen_pks = set()
            milestone_recipients = list(UserProfile.objects.filter(role='Finance', is_active=True))
            milestone_recipients += project_managers(project)
            milestone_recipients += list(UserProfile.objects.filter(role__in=['BD', 'CEO'], is_active=True))
            for recipient in milestone_recipients:
                if recipient.pk in seen_pks:
                    continue
                seen_pks.add(recipient.pk)
                _pm_link = f'/projects/{project.project_id}/'
                _pm_message = (
                    f'Project {project.project_id} ({project.customer_name}) has reached '
                    f'payment milestone: "{task.task_name}".\n\n'
                    f'Please initiate collection at the earliest.'
                )
                _pm_email_message = (
                    f'{_pm_message}\n\nView in Horizon Solar PMS:\n'
                    f'https://horizon-solar-pms-production.up.railway.app{_pm_link}'
                )
                send_notification(
                    recipient=recipient,
                    message=_pm_email_message,
                    channels=['in_app', 'whatsapp', 'email'],
                    link=_pm_link,
                    subject=f'Payment Milestone Reached — {project.customer_name}',
                    template='payment_notification',
                    template_params=[project.customer_name, task.task_name, project.customer_name],
                    related_project=project,
                    actor=request.user.profile,
                )

    # HTMX success: swap just this row with its new status/completed date. task was
    # updated via filter().update() so refresh the in-memory copy before rendering.
    if _is_hx(request):
        task.refresh_from_db()
        return _render_task_row_hx(request, project, task)

    # Honour the ?next= redirect if it's a local URL (netloc empty = same domain)
    next_url = request.POST.get('next', None)
    if next_url:
        from urllib.parse import urlparse
        if urlparse(next_url).netloc == '':
            return redirect(next_url)
    return redirect('project_overview', project_id=project.project_id)


@login_required
def task_detail_status_update(request, project_id, task_id):
    """
    Status update submitted from the task detail page.
    Only the user specifically assigned to the task (task.assigned_to) may change status.
    Uses the same transition table and notification flows as task_status_update, but
    permission is user-level (not role-level) and the redirect returns to the task detail page.
    POST only.
    """
    if request.method != 'POST':
        return redirect('task_detail', project_id=project_id, task_id=task_id)

    project = get_object_or_404(Project, project_id=project_id)
    task    = get_object_or_404(Task, pk=task_id, phase__project=project)

    try:
        profile = request.user.profile
    except Exception:
        return HttpResponseForbidden()

    # User-level check: only the specific assigned user, evaluated fresh per request
    if task.assigned_to is None or task.assigned_to != profile:
        return HttpResponseForbidden()

    new_status = request.POST.get('status', '').strip()
    valid_statuses = {s[0] for s in Task.STATUS_CHOICES}
    if new_status not in valid_statuses:
        messages.error(request, 'Invalid status value.')
        if _is_hx(request):
            return _render_task_status_hx(request, project, task)
        return redirect('task_detail', project_id=project.project_id, task_id=task.pk)

    VALID_TRANSITIONS = {
        Task.NOT_STARTED: {Task.IN_PROGRESS, Task.BLOCKED, Task.DONE},
        Task.IN_PROGRESS: {Task.DONE, Task.BLOCKED},
        Task.BLOCKED:     {Task.IN_PROGRESS, Task.BLOCKED},
        Task.DONE:        {Task.BLOCKED},
    }

    allowed = VALID_TRANSITIONS.get(task.status, set())
    if new_status not in allowed:
        messages.error(request, f"Cannot move task from '{task.status}' to '{new_status}'.")
        if _is_hx(request):
            return _render_task_status_hx(request, project, task)
        return redirect('task_detail', project_id=project.project_id, task_id=task.pk)

    # In Progress requires a due date — Finance users may supply one inline
    _due_date_str = request.POST.get('due_date', '').strip()
    if _due_date_str and new_status == Task.IN_PROGRESS and not task.due_date:
        try:
            _parsed_due = date.fromisoformat(_due_date_str)
            Task.objects.filter(pk=task.pk).update(due_date=_parsed_due)
            task.due_date = _parsed_due
        except ValueError:
            pass

    if new_status == Task.IN_PROGRESS and not task.due_date:
        messages.warning(request, 'Please set a due date before marking this task as In Progress.')
        if _is_hx(request):
            return _render_task_status_hx(request, project, task)
        return redirect('task_detail', project_id=project.project_id, task_id=task.pk)

    update_kwargs = {'status': new_status}
    if new_status == Task.DONE:
        update_kwargs['completed_at'] = timezone.now()
    if new_status == Task.BLOCKED and task.status != Task.BLOCKED:
        update_kwargs['blocked_since'] = timezone.now()
    elif new_status != Task.BLOCKED and task.status == Task.BLOCKED:
        update_kwargs['blocked_since'] = None

    if new_status == Task.BLOCKED and task.status != Task.BLOCKED:
        block_issue_title = request.POST.get('block_issue_title', '').strip()
        if not block_issue_title:
            messages.error(request, 'Please state the blocking issue before marking this task as Blocked.')
            if _is_hx(request):
                return _render_task_status_hx(request, project, task)
            return redirect('task_detail', project_id=project.project_id, task_id=task.pk)

        Task.objects.filter(pk=task.pk).update(**update_kwargs)

        block_severity = request.POST.get('block_issue_severity', Issue.HIGH)
        if block_severity not in dict(Issue.SEVERITY_CHOICES):
            block_severity = Issue.HIGH
        block_assignee = None
        block_assignee_id = request.POST.get('block_issue_assigned_to', '').strip()
        if block_assignee_id:
            try:
                block_assignee = UserProfile.objects.get(pk=block_assignee_id)
            except UserProfile.DoesNotExist:
                pass
        issue = Issue.objects.create(
            project=project,
            task=task,
            title=block_issue_title,
            description=request.POST.get('block_issue_description', '').strip(),
            severity=block_severity,
            status=Issue.OPEN,
            raised_by=profile,
            assigned_to=block_assignee,
        )
        log_activity(
            project, profile,
            f"Blocked task '{task.task_name}' — issue: {block_issue_title}",
            entity_type='Issue', entity_id=issue.pk, action_code='issue_created',
        )
        messages.success(request, f'Task blocked. Issue "{block_issue_title}" created.')
    else:
        Task.objects.filter(pk=task.pk).update(**update_kwargs)
        log_activity(project, profile, f"Changed task status to {new_status}: {task.task_name}", entity_type='Task', entity_id=task.pk,
                     action_code=f"task_status_{new_status.lower().replace(' ', '_')}")

        # Bidirectional sync: Finance confirmation tasks → PaymentMilestone Received
        _FINANCE_TASK_TO_MILESTONE = {
            'Advance Payment Confirmation': 'M1',
            'Finance Confirmation':         'M2',
            '100% Payment Confirmation':    'M3',
        }
        if new_status == Task.DONE and task.task_name in _FINANCE_TASK_TO_MILESTONE:
            _ms_label  = _FINANCE_TASK_TO_MILESTONE[task.task_name]
            _ms_update = {'status': 'Received', 'received_date': date.today()}
            _ar_str    = request.POST.get('amount_received', '').strip()
            _vr_str    = request.POST.get('variance_reason', '').strip()
            if _ar_str:
                try:
                    _ms_update['amount_received'] = Decimal(_ar_str)
                except InvalidOperation:
                    pass
            if _vr_str:
                _ms_update['variance_reason'] = _vr_str
            try:
                _ms_updated = PaymentMilestone.objects.filter(
                    project=project,
                    milestone_name=_ms_label,
                    status__in=['Pending', 'Invoiced'],
                ).update(**_ms_update)
                # Attribution (see task_status_update): record the specific person who
                # completed the Finance-confirmation task and thereby flipped the milestone.
                if _ms_updated:
                    _actor_name = profile.user.get_full_name() or profile.user.username
                    log_activity(
                        project, profile,
                        f"Milestone {_ms_label} auto-marked Received via completion of "
                        f"Finance task '{task.task_name}' by {_actor_name} ({profile.role})",
                        entity_type='Milestone',
                    )
            except Exception:
                pass

        if new_status == Task.DONE and task.is_payment_milestone:
            seen_pks = set()
            milestone_recipients = list(UserProfile.objects.filter(role='Finance', is_active=True))
            milestone_recipients += project_managers(project)
            milestone_recipients += list(UserProfile.objects.filter(role__in=['BD', 'CEO'], is_active=True))
            for recipient in milestone_recipients:
                if recipient.pk in seen_pks:
                    continue
                seen_pks.add(recipient.pk)
                _pm_link = f'/projects/{project.project_id}/'
                _pm_message = (
                    f'Project {project.project_id} ({project.customer_name}) has reached '
                    f'payment milestone: "{task.task_name}".\n\n'
                    f'Please initiate collection at the earliest.'
                )
                _pm_email_message = (
                    f'{_pm_message}\n\nView in Horizon Solar PMS:\n'
                    f'https://horizon-solar-pms-production.up.railway.app{_pm_link}'
                )
                send_notification(
                    recipient=recipient,
                    message=_pm_email_message,
                    channels=['in_app', 'whatsapp', 'email'],
                    link=_pm_link,
                    subject=f'Payment Milestone Reached — {project.customer_name}',
                    template='payment_notification',
                    template_params=[project.customer_name, task.task_name, project.customer_name],
                    related_project=project,
                    actor=profile,
                )

    if _is_hx(request):
        task.refresh_from_db()
        return _render_task_status_hx(request, project, task)

    return redirect('task_detail', project_id=project.project_id, task_id=task.pk)


def _log_task_assignment(project, actor, task, prev_assignee, new_assignee):
    """Write the correct ActivityLog line for a change to Task.assigned_to.

    Three shapes, distinguished by prev/new: positive assignment, reassignment,
    and unassign. `prev_assignee` is the value on the task BEFORE the update;
    `new_assignee` is the value after (or None). No-op (prev == new) writes nothing.
    """
    if prev_assignee == new_assignee:
        return

    def _name(p):
        return p.user.get_full_name() or p.user.username

    if new_assignee is None:
        action, code = f"Unassigned from {_name(prev_assignee)}", 'task_unassigned'
    elif prev_assignee is None:
        action, code = f"Assigned to {_name(new_assignee)}", 'task_assigned'
    else:
        action, code = (f"Reassigned from {_name(prev_assignee)} to {_name(new_assignee)}",
                        'task_reassigned')

    log_activity(project, actor, action, entity_type='Task', entity_id=task.pk, action_code=code)


@login_required
@role_required(['PM', 'Project Coordinator'])
def task_assign(request, project_id, task_id):
    """
    Assign (or clear) the individual user on a task. Candidate list is filtered
    to the task's assigned_role so a PM-role task can only be given to a PM user.
    filter().update() used to avoid overwriting other task fields via .save().
    Access: assigned PM only.
    """
    project = get_object_or_404(Project, project_id=project_id)

    if not _pm_owns_project(request, project):
        raise Http404

    task = get_object_or_404(Task, pk=task_id, phase__project=project)

    # Task.ROLE_CHOICES uses 'BD / Sales' but UserProfile stores 'BD'
    _TASK_TO_PROFILE_ROLE = {'BD / Sales': 'BD'}
    profile_role = _TASK_TO_PROFILE_ROLE.get(task.assigned_role, task.assigned_role)

    # Candidates scoped to the task's role — prevents assigning a Finance user to a PM task
    candidates = UserProfile.objects.filter(role=profile_role, is_active=True)

    if request.method == 'POST':
        prev_assignee = task.assigned_to  # captured before the update, for assignment logging
        assigned_to_id = request.POST.get('assigned_to', '').strip()
        if assigned_to_id:
            assignee = get_object_or_404(UserProfile, pk=assigned_to_id, role=profile_role, is_active=True)
            Task.objects.filter(pk=task.pk).update(assigned_to=assignee)
            _log_task_assignment(project, request.user.profile, task, prev_assignee, assignee)
            recipient_name = assignee.user.get_full_name() or assignee.user.username
            task_url = f'/projects/{project.project_id}/tasks/{task.pk}/'
            task_url_abs = request.build_absolute_uri(task_url)
            _at_message = (
                f'Hi {recipient_name},\n\n'
                f'The task "{task.task_name}" on project {project.customer_name} has been assigned to you.\n\n'
                f'Please login to review details and update progress.'
            )
            _at_email_message = (
                f'{_at_message}\n\nView in Horizon Solar PMS:\n'
                f'https://horizon-solar-pms-production.up.railway.app{task_url}'
            )
            # Guard: skip if an identical assign_task notification was logged in
            # the last 10 seconds — catches browser double-submit before the
            # redirect completes. 10 s is generous for a redirect but tight enough
            # to allow a legitimate reassign-to-same-person seconds later.
            _recent_cutoff = timezone.now() - timedelta(seconds=10)
            _already_sent = NotificationLog.objects.filter(
                recipient=assignee,
                related_project=project,
                template_name='assign_task',
                created_at__gte=_recent_cutoff,
            ).exists()
            if not _already_sent:
                send_notification(
                    recipient=assignee,
                    message=_at_email_message,
                    channels=['in_app', 'whatsapp', 'email'],
                    link=task_url,
                    subject=f'Task Assigned: {task.task_name} — {project.customer_name}',
                    template='assign_task',
                    template_params=[project.customer_name, recipient_name, task.task_name, project.customer_name, task_url_abs],
                    related_project=project,
                    actor=request.user.profile,
                )
        else:
            Task.objects.filter(pk=task.pk).update(assigned_to=None)
            _log_task_assignment(project, request.user.profile, task, prev_assignee, None)

        if _is_hx(request):
            task.refresh_from_db()
            resp = _render_task_row_hx(request, project, task)
            # Fires the "taskAssigned" event so the shared Assign modal closes.
            resp['HX-Trigger'] = 'taskAssigned'
            return resp

        return redirect('project_overview', project_id=project.project_id)

    return render(request, 'projects/task_assign_form.html', {
        'project':    project,
        'task':       task,
        'candidates': candidates,
    })


@login_required
def task_assign_design_head(request, project_id, task_id):
    """
    Assign (or clear) the individual user on a Design-role task. Separate from
    task_assign — gated on UserProfile.is_design_head instead of the PM-owns-
    project rule, so a Design Head can reassign Design tasks without being the
    project's PM. Candidate list is always Design-role, active users.
    Access: is_design_head flag only, and only for tasks where assigned_role == 'Design'.
    """
    if not request.user.profile.is_design_head:
        raise Http404

    project = get_object_or_404(Project, project_id=project_id, is_deleted=False)
    task = get_object_or_404(Task, pk=task_id, phase__project=project)

    if task.assigned_role != 'Design':
        raise Http404

    candidates = UserProfile.objects.filter(role='Design', is_active=True)

    if request.method == 'POST':
        prev_assignee = task.assigned_to  # captured before the update, for assignment logging
        assigned_to_id = request.POST.get('assigned_to', '').strip()
        if assigned_to_id:
            assignee = get_object_or_404(UserProfile, pk=assigned_to_id, role='Design', is_active=True)
            Task.objects.filter(pk=task.pk).update(assigned_to=assignee)
            _log_task_assignment(project, request.user.profile, task, prev_assignee, assignee)
            recipient_name = assignee.user.get_full_name() or assignee.user.username
            task_url = f'/projects/{project.project_id}/tasks/{task.pk}/'
            task_url_abs = request.build_absolute_uri(task_url)
            _at_message = (
                f'Hi {recipient_name},\n\n'
                f'The task "{task.task_name}" on project {project.customer_name} has been assigned to you.\n\n'
                f'Please login to review details and update progress.'
            )
            _at_email_message = (
                f'{_at_message}\n\nView in Horizon Solar PMS:\n'
                f'https://horizon-solar-pms-production.up.railway.app{task_url}'
            )
            # Guard: skip if an identical assign_task notification was logged in
            # the last 10 seconds — catches browser double-submit before the
            # redirect completes.
            _recent_cutoff = timezone.now() - timedelta(seconds=10)
            _already_sent = NotificationLog.objects.filter(
                recipient=assignee,
                related_project=project,
                template_name='assign_task',
                created_at__gte=_recent_cutoff,
            ).exists()
            if not _already_sent:
                send_notification(
                    recipient=assignee,
                    message=_at_email_message,
                    channels=['in_app', 'whatsapp', 'email'],
                    link=task_url,
                    subject=f'Task Assigned: {task.task_name} — {project.customer_name}',
                    template='assign_task',
                    template_params=[project.customer_name, recipient_name, task.task_name, project.customer_name, task_url_abs],
                    related_project=project,
                    actor=request.user.profile,
                )
        else:
            Task.objects.filter(pk=task.pk).update(assigned_to=None)
            _log_task_assignment(project, request.user.profile, task, prev_assignee, None)

        if _is_hx(request):
            task.refresh_from_db()
            return _render_task_assign_design_success_hx(request, project, task)

        return redirect('project_overview', project_id=project.project_id)

    if _is_hx(request):
        # hx-get → load the assign form into the shared modal instead of the standalone page.
        return render(request, 'projects/task_assign_design_head_modal.html', {
            'project':    project,
            'task':       task,
            'candidates': candidates,
        })

    return render(request, 'projects/task_assign_form.html', {
        'project':    project,
        'task':       task,
        'candidates': candidates,
    })


@login_required
def task_set_due_date(request, project_id, task_id):
    """
    Set the due date on a task.
    PM: can edit any task; cascade-recalculates subsequent tasks.
    Other roles: can only edit tasks assigned to their role, and only when cascade is OFF.
    POST only.
    """
    if request.method != 'POST':
        return redirect('project_overview', project_id=project_id)

    project = get_object_or_404(Project, project_id=project_id)
    profile = request.user.profile
    is_pm = _pm_owns_project(request, project)
    task = get_object_or_404(Task, pk=task_id, phase__project=project)

    if not is_pm:
        # Map UserProfile.role to Task.assigned_role (BD → BD / Sales)
        _PROFILE_TO_TASK_ROLE = {'BD': 'BD / Sales'}
        user_task_role = _PROFILE_TO_TASK_ROLE.get(profile.role, profile.role)

        if task.assigned_role != user_task_role:
            raise PermissionDenied

        if project.cascade_scheduling:
            messages.error(request, 'Due dates are managed automatically by cascading scheduling.')
            if _is_hx(request):
                return _render_task_row_hx(request, project, task)
            return redirect('project_overview', project_id=project.project_id)

        # Non-PM: save this task's date only, no cascade ripple
        date_str = request.POST.get('due_date', '').strip()
        if date_str:
            try:
                task.due_date = date.fromisoformat(date_str)
                task.save()
                log_activity(project, profile, f"Updated due date for task: {task.task_name}", entity_type='Task', entity_id=task.pk)
                messages.success(request, 'Due date updated.')
            except ValueError:
                messages.error(request, 'Invalid date.')
        else:
            task.due_date = None
            task.save()
            log_activity(project, profile, f"Cleared due date for task: {task.task_name}", entity_type='Task', entity_id=task.pk)
            messages.success(request, 'Due date cleared.')
        if _is_hx(request):
            return _render_task_row_hx(request, project, task)
        return redirect('project_overview', project_id=project.project_id)

    # PM path — ripple only when cascade is ON
    changed_tasks = []
    date_str = request.POST.get('due_date', '').strip()
    if date_str:
        try:
            new_date = date.fromisoformat(date_str)
            if project.cascade_scheduling:
                count, changed_tasks = recalculate_from_task(project, task, new_date, user=request.user)
                messages.success(request, f'Due date updated. {count} task(s) recalculated.')
            else:
                task.due_date = new_date
                task.save()
                messages.success(request, 'Due date updated.')
            log_activity(project, profile, f"Updated due date for task: {task.task_name}", entity_type='Task', entity_id=task.pk)
        except ValueError:
            messages.error(request, 'Invalid date.')
    else:
        task.due_date = None
        task.save()
        messages.success(request, 'Due date cleared.')
        log_activity(project, profile, f"Updated due date for task: {task.task_name}", entity_type='Task', entity_id=task.pk)

    if _is_hx(request):
        # Swap the edited row; every other task the cascade actually moved is
        # swapped out-of-band so the whole ripple updates without a reload.
        task.refresh_from_db()
        oob = [t for t in changed_tasks if t.pk != task.pk]
        return _render_task_row_hx(request, project, task, oob_tasks=oob)

    return redirect('project_overview', project_id=project.project_id)


@login_required
def assign_coordinators(request, project_id):
    """
    Manage the Project Coordinators on a project (multi-select).

    Coordinators share the PM's execution authority — additive-only, see
    permissions.user_can_manage_project(). Access is gated on that same capability:
    a PM (or an existing coordinator) on THIS project may edit the coordinator set.

    Uses M2M .add()/.remove() ONLY — this endpoint never assigns or reads
    assigned_pm, so PM ownership can never be transferred or lost through it
    (the Overwrite bug the invariant warns against is impossible here by construction).
    """
    project = get_object_or_404(Project, project_id=project_id)

    # Only someone who already manages this project (PM or a coordinator on it) may edit.
    if not user_can_manage_project(request.user, project):
        raise Http404

    candidates = (
        UserProfile.objects
        .filter(role='Project Coordinator', is_active=True)
        .select_related('user')
        .order_by('user__first_name')
    )

    if request.method == 'POST':
        selected_ids = {int(v) for v in request.POST.getlist('coordinator_ids') if v.isdigit()}
        valid_ids    = set(candidates.values_list('pk', flat=True))
        desired      = selected_ids & valid_ids   # ignore anything not an active PC candidate

        current   = set(project.coordinators.values_list('pk', flat=True))
        to_add    = desired - current
        to_remove = current - desired

        # M2M operations ONLY — assigned_pm is never touched.
        if to_add:
            project.coordinators.add(*to_add)
        if to_remove:
            project.coordinators.remove(*to_remove)

        if to_add or to_remove:
            actor = request.user.profile
            log_activity(
                project, actor,
                f"Updated project coordinators (added {len(to_add)}, removed {len(to_remove)})",
                entity_type='Project', entity_id=project.pk,
            )
            # In-app notify newly-added coordinators (no WhatsApp template dependency).
            for _prof in UserProfile.objects.filter(pk__in=to_add).select_related('user'):
                _name = _prof.user.get_full_name() or _prof.user.username
                _link = f'/projects/{project.project_id}/overview/'
                send_notification(
                    recipient=_prof,
                    message=(
                        f'Hi {_name}, you have been added as a Project Coordinator for '
                        f'{project.customer_name} ({project.project_id}). You now have '
                        f'execution access to this project.'
                    ),
                    channels=['in_app'],
                    link=_link,
                    related_project=project,
                    actor=actor,
                )
            messages.success(request, 'Project coordinators updated.')
        else:
            messages.info(request, 'No changes to project coordinators.')
        return redirect('project_overview', project_id=project.project_id)

    current_ids = set(project.coordinators.values_list('pk', flat=True))
    return render(request, 'projects/assign_coordinators_form.html', {
        'project':     project,
        'candidates':  candidates,
        'current_ids': current_ids,
    })


# ---------------------------------------------------------------------------
# Vendor Master
# ---------------------------------------------------------------------------

@login_required
def vendor_list(request):
    """
    List all vendors with optional category and active/inactive filters.
    Access: SCM and Admin only (inline role check — not via decorator).
    """
    profile = request.user.profile
    # Inline role check used here (not @role_required) because vendor views
    # are shared between SCM and Admin with identical permission logic
    if profile.role not in ('SCM', 'Admin'):
        return HttpResponseForbidden()

    vendors = Vendor.objects.prefetch_related('categories').order_by('-is_active', 'name')

    category_filter = request.GET.get('category', '')
    active_filter   = request.GET.get('active', '')

    if category_filter:
        vendors = vendors.filter(categories__id=category_filter)
    if active_filter == '1':
        vendors = vendors.filter(is_active=True)
    elif active_filter == '0':
        vendors = vendors.filter(is_active=False)

    return render(request, 'vendors/vendor_list.html', {
        'vendors':          vendors,
        'all_categories':   VendorCategory.objects.all(),
        'category_filter':  category_filter,
        'active_filter':    active_filter,
    })


def _save_vendor_brands(vendor, post_data):
    """
    Replace all VendorBrand entries for vendor with the brand rows from the POST data.
    Expects parallel lists: make_brand[] and brand_category[] from the dynamic form rows.
    Empty brand names are silently skipped.
    """
    brand_names = post_data.getlist('make_brand')
    brand_cats  = post_data.getlist('brand_category')

    vendor.brands.all().delete()
    for name, cat_id in zip(brand_names, brand_cats):
        name = name.strip()
        if not name:
            continue
        VendorBrand.objects.create(
            vendor=vendor,
            make_brand=name,
            category_id=int(cat_id) if cat_id else None,
        )


@login_required
def vendor_add(request):
    """
    Add a new vendor. Warns (but does not block) if a vendor with the same name
    already exists — duplicate names are allowed to handle trading variants.
    VendorBrand entries (make_brand + optional category) are saved from the
    dynamic brand rows submitted alongside the main form.
    Access: SCM and Admin only.
    """
    profile = request.user.profile
    if profile.role not in ('SCM', 'Admin'):
        return HttpResponseForbidden()

    if request.method == 'POST':
        form = VendorForm(request.POST)
        if form.is_valid():
            vendor = form.save(commit=False)
            vendor.created_by = profile
            vendor.save()
            form.save_m2m()

            # Save brand rows — each non-empty make_brand value becomes one VendorBrand entry
            _save_vendor_brands(vendor, request.POST)

            duplicate_exists = Vendor.objects.filter(
                name__iexact=vendor.name
            ).exclude(pk=vendor.pk).exists()
            if duplicate_exists:
                messages.warning(
                    request,
                    f'A vendor named "{vendor.name}" already exists. Saved anyway.'
                )
            else:
                messages.success(request, f'Vendor "{vendor.name}" added successfully.')
            return redirect('vendor_list')
    else:
        form = VendorForm()

    return render(request, 'vendors/vendor_form.html', {
        'form':             form,
        'title':            'Add Vendor',
        'vendor_brands':    [],  # no existing brands on a new vendor
        'vendor_categories': VendorCategory.objects.all(),
    })


@login_required
def vendor_edit(request, vendor_id):
    """
    Edit an existing vendor's details. Replaces all VendorBrand entries with
    the brand rows submitted in the form.
    Access: SCM and Admin only.
    """
    profile = request.user.profile
    if profile.role not in ('SCM', 'Admin'):
        return HttpResponseForbidden()

    vendor = get_object_or_404(Vendor, pk=vendor_id)

    if request.method == 'POST':
        form = VendorForm(request.POST, instance=vendor)
        if form.is_valid():
            form.save()

            # Replace all brand entries with whatever was submitted in the form
            _save_vendor_brands(vendor, request.POST)

            duplicate_exists = Vendor.objects.filter(
                name__iexact=vendor.name
            ).exclude(pk=vendor.pk).exists()
            if duplicate_exists:
                messages.warning(
                    request,
                    f'A vendor named "{vendor.name}" already exists. Saved anyway.'
                )
            else:
                messages.success(request, f'Vendor "{vendor.name}" updated successfully.')
            return redirect('vendor_list')
    else:
        form = VendorForm(instance=vendor)

    return render(request, 'vendors/vendor_form.html', {
        'form':              form,
        'vendor':            vendor,
        'title':             f'Edit Vendor — {vendor.name}',
        'vendor_brands':     list(vendor.brands.select_related('category').all()),
        'vendor_categories': VendorCategory.objects.all(),
    })


@login_required
def vendor_toggle_status(request, vendor_id):
    """
    Toggle a vendor's active/inactive status. Returns JSON so the UI can update
    the toggle button without a full page reload.
    Access: SCM and Admin only. POST only.
    """
    profile = request.user.profile
    if profile.role not in ('SCM', 'Admin'):
        return HttpResponseForbidden()
    if request.method != 'POST':
        return HttpResponseForbidden()

    vendor = get_object_or_404(Vendor, pk=vendor_id)
    vendor.is_active = not vendor.is_active
    vendor.save()
    return JsonResponse({'is_active': vendor.is_active})


# ---------------------------------------------------------------------------
# BOQ Submission
# ---------------------------------------------------------------------------

def _boq_snapshot(boq):
    """
    Return a JSON-safe snapshot of all BOQ items at a workflow transition (Decimal → float).

    make_brand_label is resolved from VendorBrand at snapshot time (vendor × category match)
    and stored directly in the JSON row so the history view is self-contained and does not
    need to re-query VendorBrand for old revisions. Falls back to the vendor company name.
    """
    import decimal as _decimal
    rows = list(boq.items.values(
        'serial_no', 'category', 'description', 'uom',
        'boq_quantity', 'ordered_quantity',
        'make_preference_id', 'make_preference__name', 'ordered_vendor__name',
    ))

    # Resolve brand labels in one query: (vendor_id, category_name) → make_brand
    vendor_ids = {r['make_preference_id'] for r in rows if r['make_preference_id']}
    brand_map  = {}
    if vendor_ids:
        for vb in VendorBrand.objects.filter(vendor_id__in=vendor_ids).select_related('category'):
            brand_map[(vb.vendor_id, vb.category.name if vb.category else None)] = vb.make_brand

    for row in rows:
        vid = row.pop('make_preference_id')  # internal FK — not needed in the stored JSON
        if vid:
            cat = row.get('category')
            row['make_brand_label'] = (
                brand_map.get((vid, cat)) or       # category-specific brand
                brand_map.get((vid, None)) or      # unscoped brand for this vendor
                row.get('make_preference__name')   # company name fallback
            )
        else:
            row['make_brand_label'] = None
        for k, v in list(row.items()):
            if isinstance(v, _decimal.Decimal):
                row[k] = float(v)
    return rows


def _notify_boq_acknowledged(boq, acknowledging_profile, request):
    """Notify PM and the acknowledging SCM user that the BOQ has been acknowledged."""
    items = boq.items.all()
    has_changes = any(
        item.ordered_quantity is not None and item.boq_quantity is not None
        and item.ordered_quantity != item.boq_quantity
        for item in items
    )
    suffix  = ' with changes to ordered quantities' if has_changes else ''
    message = (
        f'BOQ for {boq.project.project_id} ({boq.project.customer_name}) '
        f'has been acknowledged by SCM{suffix}.'
    )
    design_user_name = boq.submitted_by.user.get_full_name() or boq.submitted_by.user.username
    boq_link = f'/projects/{boq.project.project_id}/boq/'
    boq_link_abs = request.build_absolute_uri(boq_link)
    recipients = list(project_managers(boq.project))
    if acknowledging_profile not in recipients:
        recipients.append(acknowledging_profile)
    scm_name = request.user.get_full_name() or request.user.username
    _boq_message = (
        f'The Bill of Quantities for project {boq.project.customer_name} has been acknowledged '
        f'by {scm_name}.\n\n'
        f'Material procurement can now proceed.'
    )
    _boq_email_message = (
        f'{_boq_message}\n\nView in Horizon Solar PMS:\n'
        f'https://horizon-solar-pms-production.up.railway.app{boq_link}'
    )
    for recipient in recipients:
        send_notification(
            recipient=recipient,
            message=_boq_email_message,
            channels=['in_app', 'whatsapp', 'email'],
            link=boq_link,
            subject=f'BOQ Acknowledged — {boq.project.customer_name}',
            template='boq_acknowledged',
            template_params=[
                boq.project.customer_name,   # [0] header
                scm_name,                    # [1] body[0] — scm_name
            ],
            related_project=boq.project,
            actor=acknowledging_profile,
        )


def _build_vendors_by_category():
    """
    Return dict mapping category name → list of {id, name, make_brand} for active vendors.

    Vendors with VendorBrand entries appear once per brand per relevant category —
    make_brand is shown in the BOQ dropdown; the stored value is still the vendor PK.
    Brands with no category set appear in every supply category for that vendor.
    Vendors with no brands at all fall back to showing the company name.
    """
    result = {}

    # Pre-build vendor → supply categories map (from the M2M on Vendor)
    vendor_supply_cats = {}
    for v in Vendor.objects.filter(is_active=True).prefetch_related('categories'):
        vendor_supply_cats[v.pk] = [cat.name for cat in v.categories.all()]

    # Vendors that have at least one brand entry
    vendors_with_brands = set()
    for vb in (VendorBrand.objects
               .filter(vendor__is_active=True)
               .select_related('vendor', 'category')
               .order_by('vendor__name', 'make_brand')):
        vendors_with_brands.add(vb.vendor_id)
        if vb.category:
            cats = [vb.category.name]
        else:
            # Not scoped — show this brand in every category the vendor supplies
            cats = vendor_supply_cats.get(vb.vendor_id, [])
        for cat in cats:
            result.setdefault(cat, []).append({
                'id':         vb.vendor_id,
                'name':       vb.vendor.name,
                'make_brand': vb.make_brand,
            })

    # Fallback: vendors with no brands appear using their company name
    for v in (Vendor.objects
              .filter(is_active=True)
              .exclude(pk__in=vendors_with_brands)
              .prefetch_related('categories')
              .order_by('name')):
        for cat in v.categories.all():
            result.setdefault(cat.name, []).append({
                'id':         v.pk,
                'name':       v.name,
                'make_brand': '',
            })

    return result


@login_required
def boq_detail(request, project_id):
    """
    BOQ detail page. Handles all BOQ POST actions in a single view:
      Design: save_design, submit_design, add_item, delete_item
      SCM:    save_scm, acknowledge_scm
    BOQ is auto-created for Design users if it doesn't exist yet.

    Access is project-scoped, not role-scoped: read requires a relationship to THIS project
    (or a portfolio-wide read role), and each write branch below re-checks edit authority
    separately. The read gate deliberately does not stand in for the write gate — they are
    different questions with different answers, so do not hoist a single check to the top.

    PART 6 — GROUP LOCK. When this site sits in a locked SiteGroup its quantities have been
    committed to a purchase and may not move. That is enforced HERE, at the caller, as a
    second AND term beside user_can_edit_project_boq() — see project_boq_is_group_locked()
    for why the two are separate predicates rather than one. Read is unaffected: a locked
    BOQ is still fully visible, it is just no longer writable by Design.
    """
    project = get_object_or_404(Project, project_id=project_id)
    profile = request.user.profile
    role    = profile.role

    if not user_can_view_project_boq(request.user, project):
        return HttpResponseForbidden()

    # Computed once and reused by every Design write gate on this view, so the lock can
    # never be applied to one branch and forgotten on another.
    boq_group_locked = project_boq_is_group_locked(project)

    # Get or auto-create BOQ for users who may author one (others see 'no BOQ yet' state)
    try:
        boq = project.boq
    except BOQ.DoesNotExist:
        boq = None

    if boq is None:
        # Seeding 53 catalogue rows is a WRITE, so it takes the write gate even though we
        # are on a GET. A reader with no authorship relationship to this project must not
        # be able to bring a BOQ into existence just by loading the page — nor may one
        # appear on a site whose group has already been locked.
        if user_can_edit_project_boq(request.user, project) and not boq_group_locked:
            boq = BOQ.objects.create(project=project)
            # description is copied as a point-in-time snapshot; item_master carries the
            # stable catalogue link that quantity aggregation across sites joins on.
            masters = {
                m.description: m
                for m in BOQItemMaster.objects.filter(is_active=True)
            }
            BOQItem.objects.bulk_create([
                BOQItem(boq=boq, item_master=masters.get(item_data['description']), **item_data)
                for item_data in get_standard_boq_items()
            ])
        else:
            return render(request, 'projects/boq_detail.html', {
                'project': project,
                'boq':     None,
                'role':    role,
            })

    # Handle POST actions
    if request.method == 'POST':
        action = request.POST.get('action', '')

        # Design can edit in these statuses; Submitted is locked until SCM acts
        _DESIGN_EDITABLE = ('Draft', 'Revision Requested', 'Acknowledged')

        # Computed once for the three Design write branches below. Replaces the bare
        # `role == 'Design'` test each of them used to carry: role alone let any Design user
        # write the BOQ of any project in the system.
        #
        # The group-lock term is ANDed in here rather than inside user_can_edit_project_boq()
        # — that Part 0.6 helper answers "is this person the designer" and is not modified.
        # The SCM branch below deliberately does NOT take this term: it writes
        # `ordered_quantity`, and locking the group is the signal to start ordering.
        _can_edit = user_can_edit_project_boq(request.user, project) and not boq_group_locked

        # Say WHY, rather than falling through to the bare redirect at the bottom. Without
        # this the lock is enforced but silent: the POST no-ops and the page comes back
        # looking identical, which reads as "my save did nothing" rather than "this BOQ is
        # final". Only the four Design actions are answered — the SCM branch below is
        # untouched by the lock and must not be intercepted here.
        if boq_group_locked and action in ('save_design', 'submit_design',
                                           'add_item', 'delete_item'):
            messages.error(request, 'This site is in a locked procurement group — its BOQ '
                                    'quantities are final and can no longer be changed. A '
                                    'correction now needs a variance against the order.')
            return redirect('boq_detail', project_id=project_id)

        if action in ('save_design', 'submit_design') and _can_edit and boq.status in _DESIGN_EDITABLE:
            boq.notes = request.POST.get('notes', '').strip() or None
            boq.save(update_fields=['notes'])
            for item in boq.items.all():
                qty_str    = request.POST.get(f'boq_qty_{item.pk}', '').strip()
                vendor_str = request.POST.get(f'make_pref_{item.pk}', '').strip()
                try:
                    item.boq_quantity = Decimal(qty_str) if qty_str else None
                except InvalidOperation:
                    item.boq_quantity = None
                item.make_preference_id = int(vendor_str) if vendor_str.isdigit() else None
                item.save(update_fields=['boq_quantity', 'make_preference'])

            if action == 'save_design':
                messages.success(request, 'BOQ saved.')
                return redirect('boq_detail', project_id=project_id)

            # submit_design: validate then snapshot and submit
            if not boq.items.filter(boq_quantity__gt=0).exists():
                messages.error(request, 'Enter a quantity for at least one item before submitting.')
                return redirect('boq_detail', project_id=project_id)

            try:
                is_revision  = boq.status in ('Revision Requested', 'Acknowledged')
                new_version  = boq.version + 1 if is_revision else boq.version
                reason       = f'Revision v{new_version}' if is_revision else 'Initial submission'
                snapshot     = _boq_snapshot(boq)
                BOQRevision.objects.create(
                    boq=boq, revised_by=profile,
                    version=new_version, reason=reason, snapshot=snapshot,
                )
                boq.status       = 'Submitted'
                boq.submitted_by = profile
                boq.submitted_at = timezone.now()
                if is_revision:
                    boq.version = new_version
                boq.save(update_fields=['status', 'submitted_by', 'submitted_at', 'version'])
                messages.success(request, 'BOQ submitted to SCM for review.')
                return redirect('boq_detail', project_id=project_id)
            except Exception as exc:
                logger.exception('BOQ submit_design failed for %s', project_id)
                messages.error(request, f'Submit failed: {exc}')
                return redirect('boq_detail', project_id=project_id)

        elif action in ('save_scm', 'acknowledge_scm') and role == 'SCM' and boq.status in ('Submitted', 'Acknowledged'):
            # Save ordered qty, make preference, and ordered vendor regardless of save vs. acknowledge
            for item in boq.items.all():
                qty_str    = request.POST.get(f'ord_qty_{item.pk}', '').strip()
                make_str   = request.POST.get(f'make_pref_{item.pk}', '').strip()
                vendor_str = request.POST.get(f'ord_vendor_{item.pk}', '').strip()
                try:
                    item.ordered_quantity = Decimal(qty_str) if qty_str else None
                except InvalidOperation:
                    item.ordered_quantity = None
                item.make_preference_id = int(make_str) if make_str.isdigit() else None
                item.ordered_vendor_id  = int(vendor_str) if vendor_str.isdigit() else None
                item.save(update_fields=['ordered_quantity', 'make_preference', 'ordered_vendor'])

            if action == 'save_scm':
                messages.success(request, 'Ordered details saved.')
                return redirect('boq_detail', project_id=project_id)

            # acknowledge_scm: snapshot then acknowledge
            snapshot = _boq_snapshot(boq)
            BOQRevision.objects.create(
                boq=boq, revised_by=profile,
                version=boq.version,
                reason=f'SCM Acknowledged v{boq.version}',
                snapshot=snapshot,
            )
            boq.status = 'Acknowledged'
            boq.save(update_fields=['status'])

            # Notify the Design user who submitted this BOQ
            if boq.submitted_by:
                _notify_boq_acknowledged(boq, profile, request)

            messages.success(request, 'BOQ acknowledged.')
            return redirect('boq_detail', project_id=project_id)

        elif action == 'add_item' and _can_edit and boq.status in _DESIGN_EDITABLE:
            category    = request.POST.get('new_category', 'Other')
            description = request.POST.get('new_description', '').strip()
            uom         = request.POST.get('new_uom', 'Nos')
            if description:
                last_serial = boq.items.aggregate(Max('serial_no'))['serial_no__max'] or 0
                BOQItem.objects.create(
                    boq=boq, serial_no=last_serial + 1,
                    category=category, description=description,
                    uom=uom, is_standard_item=False,
                )
                messages.success(request, 'Row added.')
            return redirect('boq_detail', project_id=project_id)

        elif action == 'delete_item' and _can_edit and boq.status in _DESIGN_EDITABLE:
            item_id = request.POST.get('item_id', '')
            if item_id.isdigit():
                # Object consistency: the item must belong to THIS project's BOQ. An id from
                # another project's BOQ is a 404, not a silent no-op reported as success.
                item = get_object_or_404(BOQItem, pk=int(item_id), boq=boq)
                if not item.is_standard_item:     # standard rows are undeletable, as before
                    item.delete()
                    messages.success(request, 'Row deleted.')
            return redirect('boq_detail', project_id=project_id)

        return redirect('boq_detail', project_id=project_id)

    items = list(boq.items.select_related('make_preference', 'ordered_vendor').order_by('serial_no'))

    # Annotate each item with make_brand_display: the brand label shown in the static
    # (non-editable) Make/Preference cell. Resolved from VendorBrand in one batch query.
    _vendor_ids = {i.make_preference_id for i in items if i.make_preference_id}
    _brand_map  = {}
    if _vendor_ids:
        for vb in VendorBrand.objects.filter(vendor_id__in=_vendor_ids).select_related('category'):
            _brand_map[(vb.vendor_id, vb.category.name if vb.category else None)] = vb.make_brand
    for item in items:
        if item.make_preference_id:
            item.make_brand_display = (
                _brand_map.get((item.make_preference_id, item.category)) or
                _brand_map.get((item.make_preference_id, None)) or
                item.make_preference.name
            )
        else:
            item.make_brand_display = None

    items_by_category = {}
    for item in items:
        items_by_category.setdefault(item.category, []).append(item)

    # Whether the Design half of the page renders INPUTS or static values. One flag,
    # because the template asked the same three-way status question in seven places and
    # the group lock has to reach every one of them.
    #
    # The status tuple is spelled out rather than read from `_DESIGN_EDITABLE`: that
    # constant is local to the POST branch above, and hoisting it to module scope is an
    # edit to a constant Part 6 is told to leave alone. The duplication is deliberate and
    # recorded in DESIGN_MODULE_DEFERRED.md — keep the two in step if either changes.
    design_form_open = (boq.status in ('Draft', 'Revision Requested', 'Acknowledged')
                        and not boq_group_locked)

    # The locked group itself, for the banner. Only fetched when the lock is on, so the
    # unlocked path (every Residential BOQ, always) costs no extra query.
    locked_group = None
    if boq_group_locked:
        locked_membership = (project.group_memberships
                             .filter(removed_at__isnull=True, group__status='locked')
                             .select_related('group', 'group__locked_by__user').first())
        locked_group = locked_membership.group if locked_membership else None

    return render(request, 'projects/boq_detail.html', {
        'project':             project,
        'boq':                 boq,
        'items':               items,
        'items_by_category':   items_by_category.items(),
        'role':                role,
        'boq_group_locked':    boq_group_locked,
        'locked_group':        locked_group,
        'design_form_open':    design_form_open,
        'vendors_by_category': _build_vendors_by_category(),
        'category_choices':    BOQItem.CATEGORY_CHOICES,
        'uom_choices':         BOQItem.UOM_CHOICES,
        'today':               date.today(),
    })


@login_required
def boq_submit(request, project_id):
    """
    Standalone BOQ submit endpoint (also handled inline in boq_detail).
    Validates at least one item has a quantity, snapshots the BOQ, and moves
    status to Submitted. Increments version on resubmission.
    Access: Design, on a project this user has a design relationship to. POST only.

    PART 6: the group-lock term is ANDed in alongside the Part 0.6 authority helper, the
    same way boq_detail does it, so this endpoint cannot become a way around the lock.
    """
    if request.method != 'POST':
        return redirect('boq_detail', project_id=project_id)

    project = get_object_or_404(Project, project_id=project_id)
    profile = request.user.profile

    if not user_can_edit_project_boq(request.user, project):
        return HttpResponseForbidden()

    if project_boq_is_group_locked(project):
        return HttpResponseForbidden(
            'This site is in a locked procurement group — its BOQ can no longer be changed.')

    # `project=project` scopes the BOQ to the project in the URL — a BOQ belonging to any
    # other project is unreachable through this endpoint.
    boq = get_object_or_404(BOQ, project=project)

    if boq.status not in ('Draft', 'Revision Requested'):
        messages.error(request, 'BOQ cannot be submitted in its current state.')
        return redirect('boq_detail', project_id=project_id)

    if not boq.items.filter(boq_quantity__gt=0).exists():
        messages.error(request, 'At least one item must have a BOQ quantity before submitting.')
        return redirect('boq_detail', project_id=project_id)

    is_resubmission = boq.status == 'Revision Requested'
    new_version     = boq.version + 1 if is_resubmission else boq.version
    reason          = f'Revision v{new_version}' if is_resubmission else 'Initial submission'

    snapshot = list(boq.items.values(
        'serial_no', 'category', 'description', 'uom',
        'boq_quantity', 'ordered_quantity',
        'make_preference__name', 'ordered_vendor__name',
    ))

    BOQRevision.objects.create(
        boq=boq, revised_by=profile,
        version=new_version, reason=reason, snapshot=snapshot,
    )

    boq.status       = 'Submitted'
    boq.submitted_by = profile
    boq.submitted_at = timezone.now()
    if is_resubmission:
        boq.version = new_version
    boq.save()

    # Log BOQ submission event
    log_activity(project, profile, f"Submitted BOQ for project: {project.project_id}", entity_type='BOQ', entity_id=boq.pk)
    messages.success(request, 'BOQ submitted to SCM for review.')
    return redirect('boq_detail', project_id=project_id)


@login_required
def boq_acknowledge(request, project_id):
    """
    Standalone BOQ acknowledge endpoint (also handled inline in boq_detail).
    Moves status from Submitted → Acknowledged. Access: SCM only. POST only.
    """
    if request.method != 'POST':
        return redirect('boq_detail', project_id=project_id)

    project = get_object_or_404(Project, project_id=project_id)
    profile = request.user.profile

    # DELIBERATELY role-only, with no project relationship requirement. SCM is portfolio-wide
    # by remit and acknowledges BOQs across every project; scoping this to a relationship
    # would break acknowledgement system-wide. Do not "harden" this to match the Design gates.
    if profile.role != 'SCM':
        return HttpResponseForbidden()

    # `project=project` scopes the BOQ to the project in the URL.
    boq = get_object_or_404(BOQ, project=project)

    if boq.status != 'Submitted':
        messages.error(request, 'Only submitted BOQs can be acknowledged.')
        return redirect('boq_detail', project_id=project_id)

    boq.status = 'Acknowledged'
    boq.save()

    # Log BOQ acknowledgment event
    log_activity(project, profile, f"BOQ Acknowledged for project: {project.project_id}", entity_type='BOQ', entity_id=boq.pk)
    messages.success(request, 'BOQ acknowledged.')
    return redirect('boq_detail', project_id=project_id)


@login_required
def boq_request_revision(request, project_id):
    """
    PM requests a revision on a Submitted or Acknowledged BOQ.
    Snapshots current state before moving to 'Revision Requested'.
    GET renders the reason form; POST processes the request.

    Access: PM-level management authority on THIS project — the assigned PM or a Project
    Coordinator on it. Coordinators were always in the old role tuple and that is correct;
    user_can_manage_project() treats them as PM-equivalent, so routing through it keeps the
    two in step. (The docstring previously said "PM only", which the role tuple never was.)
    """
    project = get_object_or_404(Project, project_id=project_id)
    profile = request.user.profile

    if not user_can_manage_project(request.user, project):
        return HttpResponseForbidden()

    # `project=project` scopes the BOQ to the project in the URL.
    boq = get_object_or_404(BOQ, project=project)

    if boq.status not in ('Submitted', 'Acknowledged'):
        messages.error(request, 'Revision can only be requested on Submitted or Acknowledged BOQs.')
        return redirect('boq_detail', project_id=project_id)

    if request.method == 'POST':
        reason = request.POST.get('reason', '').strip()
        if not reason:
            messages.error(request, 'Please provide a reason for the revision request.')
            return render(request, 'projects/boq_request_revision.html', {
                'project': project,
                'boq':     boq,
            })

        try:
            snapshot = _boq_snapshot(boq)
            BOQRevision.objects.create(
                boq=boq, revised_by=profile,
                version=boq.version, reason=f'Revision requested: {reason}',
                snapshot=snapshot,
            )
            boq.status = 'Revision Requested'
            boq.save(update_fields=['status'])
            # Log BOQ revision request event
            log_activity(project, profile, f"BOQ Revision Requested for project: {project.project_id}", entity_type='BOQ', entity_id=boq.pk)
            messages.success(request, 'Revision requested. Design team will be notified.')
            return redirect('boq_detail', project_id=project_id)
        except Exception as exc:
            logger.exception('boq_request_revision failed for %s', project_id)
            messages.error(request, f'Request failed: {exc}')
            return redirect('boq_detail', project_id=project_id)

    return render(request, 'projects/boq_request_revision.html', {
        'project': project,
        'boq':     boq,
    })


@login_required
def boq_history(request, project_id):
    """
    BOQ revision timeline. Annotates each BOQRevision with event_type and badge
    styling based on the reason text. Ordered oldest-first for a readable timeline.
    Access: same read gate as boq_detail — the history is the BOQ's own audit trail, so
    anyone who may read the BOQ may read how it got there, and nobody else.
    """
    project = get_object_or_404(Project, project_id=project_id)

    if not user_can_view_project_boq(request.user, project):
        return HttpResponseForbidden()

    # `project=project` scopes the BOQ to the project in the URL.
    boq = get_object_or_404(BOQ, project=project)
    # Chronological order so the timeline reads top-to-bottom oldest-first
    raw_revisions = boq.revisions.select_related('revised_by__user').order_by('revised_at')

    def _annotate(rev):
        """Tag a BOQRevision with display metadata (event_type, badge CSS class) based on its reason text."""
        r = rev.reason or ''
        if 'SCM Acknowledged' in r:
            rev.event_type   = 'acknowledged'
            rev.event_label  = 'Acknowledged'
            rev.dot_color    = 'bg-success'
            rev.badge_class  = 'bg-success'
        elif 'Revision requested' in r:
            rev.event_type   = 'revision_requested'
            rev.event_label  = 'Revision Requested'
            rev.dot_color    = 'bg-warning'
            rev.badge_class  = 'bg-warning text-dark'
        elif r.startswith('Revision v'):
            rev.event_type   = 'resubmitted'
            rev.event_label  = 'Resubmitted'
            rev.dot_color    = 'bg-primary'
            rev.badge_class  = 'bg-primary'
        else:
            rev.event_type   = 'submitted'
            rev.event_label  = 'Submitted'
            rev.dot_color    = 'bg-primary'
            rev.badge_class  = 'bg-primary'
        return rev

    revisions = [_annotate(rev) for rev in raw_revisions]

    return render(request, 'projects/boq_history.html', {
        'project':   project,
        'boq':       boq,
        'revisions': revisions,
    })


@login_required
def notifications_view(request):
    """
    List the user's notifications and mark them all as read.
    Both GET and POST mark notifications read — visiting the page clears the badge.
    Access: any authenticated user (their own notifications only).
    """
    profile       = request.user.profile
    notifications = profile.notifications.all()  # already ordered -created_at

    if request.method == 'POST':
        # Mark all as read
        profile.notifications.filter(is_read=False).update(is_read=True)
        return redirect('notifications')

    # Mark all as read on GET too (visiting the page clears the badge)
    profile.notifications.filter(is_read=False).update(is_read=True)

    return render(request, 'projects/notifications.html', {
        'notifications':      notifications,
        'user_dashboard_url': get_user_dashboard(request.user),
    })


# ---------------------------------------------------------------------------
# Finance milestone actions
# ---------------------------------------------------------------------------

@login_required
@role_required(['Finance'])
def milestone_invoice(request, project_id, milestone_pk):
    """
    Mark a Pending milestone as Invoiced and record today as invoice_date.
    Access: Finance only. POST only.
    """
    if request.method != 'POST':
        return redirect('dashboard_finance')

    milestone = get_object_or_404(
        PaymentMilestone, pk=milestone_pk, project__project_id=project_id
    )
    if milestone.status != 'Pending':
        messages.error(request, 'Only Pending milestones can be marked as Invoiced.')
        return redirect('dashboard_finance')

    milestone.status       = 'Invoiced'
    milestone.invoice_date = date.today()
    milestone.save(update_fields=['status', 'invoice_date'])
    # Log milestone invoiced — project accessed via FK to avoid extra query
    log_activity(milestone.project, request.user.profile, f"Marked {milestone.milestone_name} as Invoiced", entity_type='Milestone', entity_id=milestone.pk)
    messages.success(request, f'{milestone.milestone_name} marked as Invoiced.')
    return redirect('dashboard_finance')


@login_required
@role_required(['Finance'])
def milestone_receive(request, project_id, milestone_pk):
    """
    Mark an Invoiced milestone as Received. Records amount_received and variance_reason.
    Auto-sets variance_reason='Overpayment' when amount_received > amount and no reason given.
    Access: Finance only. POST only.
    """
    if request.method != 'POST':
        return redirect('dashboard_finance')

    milestone = get_object_or_404(
        PaymentMilestone, pk=milestone_pk, project__project_id=project_id
    )
    if milestone.status == 'Received':
        messages.error(request, 'Milestone is already marked as Received.')
        return redirect('dashboard_finance')

    amount_received_str = request.POST.get('amount_received', '').strip()
    if not amount_received_str:
        messages.error(request, 'Amount received is required.')
        return redirect('dashboard_finance')

    try:
        amount_received = Decimal(amount_received_str)
    except InvalidOperation:
        messages.error(request, 'Invalid amount value.')
        return redirect('dashboard_finance')

    variance_reason = request.POST.get('variance_reason', '').strip()
    if milestone.amount and amount_received > milestone.amount and not variance_reason:
        variance_reason = 'Overpayment'

    milestone.status          = 'Received'
    milestone.received_date   = date.today()
    milestone.amount_received = amount_received
    milestone.variance_reason = variance_reason
    milestone.save(update_fields=['status', 'received_date', 'amount_received', 'variance_reason'])
    # Log milestone received — project accessed via FK to avoid extra query
    log_activity(milestone.project, request.user.profile, f"Marked {milestone.milestone_name} as Received", entity_type='Milestone', entity_id=milestone.pk)

    # Bidirectional sync: PaymentMilestone Received → Finance confirmation task Done.
    _MILESTONE_TO_FINANCE_TASK = {
        'M1': 'Advance Payment Confirmation',
        'M2': 'Finance Confirmation',
        'M3': '100% Payment Confirmation',
    }
    _sync_task_name = _MILESTONE_TO_FINANCE_TASK.get(milestone.milestone_name)
    if _sync_task_name:
        try:
            Task.objects.filter(
                phase__project=milestone.project,
                task_name=_sync_task_name,
                status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED],
            ).update(status=Task.DONE, completed_at=timezone.now())
        except Exception:
            pass  # Non-critical — never block the milestone update

    messages.success(request, f'{milestone.milestone_name} marked as Received.')
    return redirect('dashboard_finance')


@login_required
@role_required(['PM', 'Project Coordinator'])
def milestone_create(request, project_id):
    """
    Create the standard M1/M2/M3 milestones for a project if none exist yet.
    Milestones are normally created automatically during project_activate —
    this view exists as a fallback for projects activated before milestones were added.
    Access: assigned PM only. POST only.
    """
    if request.method != 'POST':
        return redirect('project_overview', project_id=project_id)

    project = get_object_or_404(Project, project_id=project_id)
    if not _pm_owns_project(request, project):
        raise Http404

    if not project.milestones.exists():
        milestone_defaults = [
            ('M1', 'On Survey Completion'),
            ('M2', 'On Material Supply'),
            ('M3', 'On Commissioning'),
        ]
        for name, desc in milestone_defaults:
            PaymentMilestone.objects.create(
                project=project,
                milestone_name=name,
                milestone_description=desc,
                created_by=request.user.profile,
            )
        messages.success(request, 'Payment milestones created.')
    else:
        messages.info(request, 'Milestones already exist for this project.')
    return redirect('project_overview', project_id=project_id)


# ---------------------------------------------------------------------------
# Payment Requests
# ---------------------------------------------------------------------------

@login_required
def raise_payment_request(request, project_id):
    """SCM raises a payment request against a vendor invoice. SCM role only. POST."""
    if request.method != 'POST':
        return redirect('dashboard_scm')

    # Raising a payment request is SCM-only — other roles get 403
    profile = request.user.profile
    if profile.role != 'SCM':
        return HttpResponse(status=403)

    project = get_object_or_404(Project, project_id=project_id)

    vendor_id      = request.POST.get('vendor_id', '').strip()
    boq_item_id    = request.POST.get('boq_item_id', '').strip()
    invoice_number = request.POST.get('invoice_number', '').strip()
    amount_str     = request.POST.get('amount', '').strip()
    note           = request.POST.get('note', '').strip()
    invoice_file   = request.FILES.get('invoice_document')

    # Server-side validation — invoice_document is mandatory (hard business rule, not a UI nicety)
    errors = []
    if not vendor_id:
        errors.append('Vendor is required.')
    if not boq_item_id:
        errors.append('BOQ item is required.')
    if not invoice_number:
        errors.append('Invoice number is required.')
    if not amount_str:
        errors.append('Amount is required.')
    if not invoice_file:
        errors.append('Invoice document is required.')

    if errors:
        messages.error(request, ' '.join(errors))
        return redirect('dashboard_scm')

    try:
        amount = Decimal(amount_str)
    except InvalidOperation:
        messages.error(request, 'Invalid amount.')
        return redirect('dashboard_scm')

    vendor = get_object_or_404(Vendor, pk=vendor_id, is_active=True)

    # BOQ item dropdown scoped to THIS project only — never show another
    # project's BOQ items in the raise-request form.
    boq_item = get_object_or_404(BOQItem, pk=boq_item_id, boq__project=project)

    # Upload invoice document to Supabase — reuse same pattern as ProjectDocument/TaskAttachment
    try:
        from .supabase_storage import get_supabase_client
        client = get_supabase_client()
    except (ValueError, ImportError) as exc:
        messages.error(request, f"Upload service unavailable. ({exc})")
        return redirect('dashboard_scm')

    supabase_path = (
        f"payment-requests/{project.project_id}/"
        f"{_uuid.uuid4()}_{invoice_file.name}"
    )
    try:
        _validate_and_upload(invoice_file, client, settings.SUPABASE_BUCKET, supabase_path)
        invoice_document_url = (
            f"{settings.SUPABASE_URL}/storage/v1/object/public/"
            f"{settings.SUPABASE_BUCKET}/{supabase_path}"
        )
    except ValueError as exc:
        messages.error(request, f"Invoice upload failed: {exc}")
        return redirect('dashboard_scm')

    pr = PaymentRequest.objects.create(
        project=project,
        vendor=vendor,
        boq_item=boq_item,
        invoice_number=invoice_number,
        invoice_document_name=invoice_file.name,
        invoice_document_url=invoice_document_url,
        invoice_document_path=supabase_path,
        amount=amount,
        note=note,
        requested_by=request.user,
        status=PaymentRequest.PENDING,
    )

    # Confirm actions are state-mutating — log them so they surface in the
    # project's Recent Activity panel for audit trail.
    log_activity(
        project, profile,
        f"Raised payment request to {vendor.name}: ₹{amount} (Invoice {invoice_number})",
        entity_type='PaymentRequest', entity_id=pr.pk,
    )
    messages.success(request, f'Payment request raised for ₹{amount} to {vendor.name}.')
    return redirect('dashboard_scm')


@login_required
def confirm_payment_request(request, project_id, request_id):
    """Finance confirms a vendor payment has been made. Finance role only. POST."""
    if request.method != 'POST':
        return redirect('project_overview', project_id=project_id)

    # Confirming payment is Finance-only — SCM and PM see status read-only
    profile = request.user.profile
    if profile.role != 'Finance':
        return HttpResponse(status=403)

    project = get_object_or_404(Project, project_id=project_id)
    pr = get_object_or_404(
        PaymentRequest.objects.select_related('vendor', 'boq_item'),
        pk=request_id, project=project, status=PaymentRequest.PENDING,
    )

    payment_date_str  = request.POST.get('payment_date', '').strip()
    payment_reference = request.POST.get('payment_reference', '').strip()

    if not payment_date_str:
        messages.error(request, 'Payment date is required.')
        return redirect('project_overview', project_id=project_id)

    try:
        payment_date = date.fromisoformat(payment_date_str)
    except ValueError:
        messages.error(request, 'Invalid payment date.')
        return redirect('project_overview', project_id=project_id)

    pr.status            = PaymentRequest.CONFIRMED
    pr.payment_date      = payment_date
    pr.payment_reference = payment_reference
    pr.confirmed_by      = request.user
    pr.save(update_fields=['status', 'payment_date', 'payment_reference', 'confirmed_by'])

    # Confirm actions are state-mutating — log them so they surface in the
    # project's Recent Activity panel for audit trail.
    log_activity(
        project, profile,
        f"Confirmed payment to {pr.vendor}: ₹{pr.amount} (Ref: {payment_reference or '—'})",
        entity_type='PaymentRequest', entity_id=pr.pk,
    )

    boq_desc = pr.boq_item.description if pr.boq_item else '(item)'
    seen_pks = set()
    invoice_recipients = list(UserProfile.objects.filter(role='SCM', is_active=True))
    invoice_recipients += project_managers(project)
    invoice_recipients += list(UserProfile.objects.filter(role='CEO', is_active=True))
    _ip_link = f'/projects/{project.project_id}/overview/'
    _ip_message = (
        f'Payment has been confirmed for project {project.customer_name}.\n\n'
        f'Vendor: {pr.vendor.name}. Invoice: {pr.invoice_number}. '
        f'Amount: Rs. {pr.amount}. Item: {boq_desc}.\n\n'
        f'Records have been updated.'
    )
    _ip_email_message = (
        f'{_ip_message}\n\nView in Horizon Solar PMS:\n'
        f'https://horizon-solar-pms-production.up.railway.app{_ip_link}'
    )
    for recipient in invoice_recipients:
        if recipient.pk in seen_pks:
            continue
        seen_pks.add(recipient.pk)
        send_notification(
            recipient=recipient,
            message=_ip_email_message,
            channels=['in_app', 'whatsapp', 'email'],
            link=_ip_link,
            subject=f'Invoice Payment Confirmed — {project.customer_name}',
            template='invoice_paid',
            template_params=[project.customer_name, boq_desc, pr.invoice_number, str(pr.amount), pr.vendor.name],
            related_project=project,
            actor=profile,
        )

    messages.success(request, f'Payment of ₹{pr.amount} to {pr.vendor} confirmed.')
    return redirect('project_overview', project_id=project_id)


# ---------------------------------------------------------------------------
# BD dashboard
# ---------------------------------------------------------------------------

@login_required
@role_required(['BD'])  # UserProfile.role = 'BD'; confirmed by Trigger 2
def dashboard_bd(request):
    """Sales & BD dashboard. One card per project, combined urgency circle. BD role only."""
    today = date.today()

    # Trigger 4: BD is department-level — no assigned_bd field on Project model confirmed.
    # BD sees all active/in-progress projects, same as Finance and SCM dashboards.
    #
    # Trigger 3: milestones on PaymentMilestone (related_name='milestones') —
    # prefetch reused exactly from Finance dashboard session; same pattern, same model.
    #
    # Trigger 6: annotate open issue count using Open+InProgress — same definition as
    # PM dashboard's open_issue_count. "Blocked issue" is BD's term for the same thing;
    # there is no separate "blocked" status on Issue.
    projects_qs = (
        Project.objects.filter(is_deleted=False, status__in=['Active', 'In Progress'])
        .prefetch_related(
            'milestones',
            Prefetch(
                'phases',
                queryset=ProjectPhase.objects.prefetch_related('tasks').order_by('phase_order'),
            ),
        )
        .select_related('assigned_pm__user')
        .annotate(
            open_issue_count=Count(
                'issues',
                filter=Q(issues__status__in=[Issue.OPEN, Issue.IN_PROGRESS]),
                distinct=True,
            )
        )
        .order_by('project_id')
    )

    # Trigger 5: No orc_status field on Project model — derived from BD task completion.
    # Task with assigned_role='BD / Sales' in "Sales & Documentation" phase represents ORC.
    # One batch query across all active projects avoids N+1 in the loop below.
    orc_done_project_ids = set(
        Task.objects.filter(
            phase__project__status__in=['Active', 'In Progress'],
            assigned_role=Task.BD,  # = 'BD / Sales'
            status=Task.DONE,
        ).values_list('phase__project_id', flat=True).distinct()
    )

    project_rows = []
    for project in projects_qs:
        # is_delayed: same view-computed logic as PM/SE/Finance/Design dashboards
        is_delayed = bool(
            project.target_commissioning_date
            and project.target_commissioning_date < today
        )
        delay_days = (
            (today - project.target_commissioning_date).days
            if is_delayed else None
        )

        # Current phase: first phase with an incomplete task; uses prefetched data — no extra query
        current_phase_name = None
        for phase in project.phases.all():
            for task in phase.tasks.all():
                if task.status != Task.DONE:
                    current_phase_name = phase.phase_name
                    break
            if current_phase_name:
                break

        # Trigger 3: milestone summary — same prefetch as Finance dashboard session.
        # milestones_awaiting = 'Invoiced' status: Finance has invoiced, BD nudges client for receipt.
        # (Finance's milestones_awaiting = 'Pending'; BD's perspective is post-invoice.)
        milestones_total    = 0
        milestones_received = 0
        milestones_awaiting = 0
        milestones_list     = []
        for m in project.milestones.all():
            milestones_total += 1
            if m.status == PaymentMilestone.RECEIVED:
                milestones_received += 1
            elif m.status == PaymentMilestone.INVOICED:
                # Invoice sent to client; BD follows up for payment receipt
                milestones_awaiting += 1
            # Variance for display: positive = short (expected > received), negative = excess
            if m.amount is not None and m.amount_received is not None:
                _var = m.amount - m.amount_received
                _var_abs = abs(_var)
            else:
                _var = None
                _var_abs = None
            milestones_list.append({
                'pk':                   m.pk,
                'milestone_name':       m.milestone_name,
                'milestone_description': m.milestone_description,
                'amount':               m.amount,
                'amount_received':      m.amount_received,
                'status':               m.status,
                'variance':             _var,
                'variance_abs':         _var_abs,
            })

        # Trigger 6: annotated open_issue_count = Open or In Progress — same as PM/SE dashboards
        blocked_issue_count = project.open_issue_count

        # Trigger 5: ORC derived from BD task completion — no orc_status field on Project.
        # orc_overdue hardcoded False — no SLA threshold defined; same pattern as Design dashboard.
        orc_status  = 'uploaded' if project.pk in orc_done_project_ids else 'pending'
        orc_overdue = False  # TODO: wire to real SLA once thresholds are defined in Claude Chat

        # Urgency formula from spec: blocked issues + invoiced-not-received milestones + overdue ORC
        urgency_count = (
            blocked_issue_count
            + milestones_awaiting
            + (1 if orc_status == 'pending' and orc_overdue else 0)
        )

        # Precedence: Blocked > Delayed > On-time, computed server-side.
        # Blocked overrides Delayed even when both are true — worse state wins.
        # Confirmed in Claude Chat, 19-June session.
        if blocked_issue_count > 0:
            status_badge = 'blocked'
        elif is_delayed:
            status_badge = 'delayed'
        else:
            status_badge = 'on_time'

        pm_name = None
        if project.assigned_pm:
            # Trigger 7: assigned_pm FK → UserProfile; full name with username fallback
            pm_name = (
                project.assigned_pm.user.get_full_name()
                or project.assigned_pm.user.username
            )

        project_rows.append({
            'project':             project,
            'project_id':          project.project_id,
            'customer_name':       project.customer_name,
            'phase':               current_phase_name,
            'pm_name':             pm_name,
            'is_delayed':          is_delayed,
            'delay_days':          delay_days,
            'orc_status':          orc_status,
            'orc_overdue':         orc_overdue,
            'milestones':          milestones_list,
            'milestones_total':    milestones_total,
            'milestones_received': milestones_received,
            'milestones_awaiting': milestones_awaiting,
            'blocked_issue_count': blocked_issue_count,
            'urgency_count':       urgency_count,
            'status_badge':        status_badge,
        })

    # Most-urgent-first; all-clear projects sink to bottom
    project_rows.sort(key=lambda r: r['urgency_count'], reverse=True)

    # Top-level summary — derived from per-project data, no extra queries
    total_blocked_projects    = sum(1 for r in project_rows if r['blocked_issue_count'] > 0)
    total_orc_overdue         = sum(1 for r in project_rows if r['orc_overdue'])
    total_milestones_awaiting = sum(r['milestones_awaiting'] for r in project_rows)

    return render(request, 'dashboard/bd.html', {
        'bd_first_name':             request.user.first_name,
        'project_rows':              project_rows,
        'total_blocked_projects':    total_blocked_projects,
        'total_orc_overdue':         total_orc_overdue,
        'total_milestones_awaiting': total_milestones_awaiting,
    })


@login_required
@role_required(['BD', 'PM'])
def set_milestone_amounts(request, project_id):
    """
    Set M1/M2/M3 agreed amounts for a project. Called by BD on Phase 1 Task 1 Done gate
    and by the BD dashboard inline edit pencil.
    Null values for a milestone key = skip that milestone (used for single-milestone edits).
    POST body: JSON {m1_amount, m2_amount, m3_amount} where each is a number or null.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST only'}, status=405)

    project = get_object_or_404(Project, project_id=project_id)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid request body.'})

    amounts = {}
    for name, key in [('M1', 'm1_amount'), ('M2', 'm2_amount'), ('M3', 'm3_amount')]:
        raw = data.get(key)
        if raw is None:
            continue  # null in JSON = skip this milestone
        raw_str = str(raw).strip()
        if raw_str == '':
            return JsonResponse({'success': False, 'error': 'All three milestone amounts are required.'})
        try:
            val = Decimal(raw_str)
            if val < 0:
                return JsonResponse({'success': False, 'error': 'Amounts must be 0 or greater.'})
            amounts[name] = val
        except InvalidOperation:
            return JsonResponse({'success': False, 'error': 'All three milestone amounts are required.'})

    if not amounts:
        return JsonResponse({'success': False, 'error': 'No amounts provided.'})

    # Contract value check — only when all three milestones will be non-null after this update
    existing = {m.milestone_name: m.amount for m in project.milestones.all() if m.amount is not None}
    merged = dict(existing)
    merged.update(amounts)
    if all(merged.get(n) is not None for n in ('M1', 'M2', 'M3')):
        total = merged['M1'] + merged['M2'] + merged['M3']
        contract = project.contract_value
        if contract is not None and total != contract:
            return JsonResponse({
                'success': False,
                'error': (
                    f'Milestone amounts must sum to ₹{contract:,.0f} (contract value). '
                    f'Current total: ₹{total:,.0f}. '
                    f'Difference: ₹{abs(contract - total):,.0f}.'
                ),
            })

    try:
        with transaction.atomic():
            for name, amount in amounts.items():
                PaymentMilestone.objects.update_or_create(
                    project=project, milestone_name=name,
                    defaults={'amount': amount},
                )
        log_activity(
            project, request.user.profile,
            'Set milestone amounts: ' + ', '.join(f'{n}=₹{a}' for n, a in amounts.items()),
            entity_type='Milestone',
        )
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': True})


# ---------------------------------------------------------------------------
# Individual project overview
# ---------------------------------------------------------------------------

@login_required
def project_overview(request, project_id):
    """
    Combined project page: info card, status blocks, documents, and phase/task list.
    Merges the former project_detail (task management) and project_overview (status summary)
    into a single URL. Access: all roles; PM and SE isolation applies.
    """
    project = get_object_or_404(
        Project.objects.select_related(
            'assigned_pm__user', 'assigned_design__user', 'created_by',
        ),
        project_id=project_id,
    )
    profile = request.user.profile
    role    = profile.role

    # Role isolation
    if role == 'PM' and not user_can_manage_project(request.user, project):
        raise Http404

    is_assigned_pm    = user_can_manage_project(request.user, project)
    can_assign_design = is_assigned_pm and project.status in ('Active', 'In Progress')

    # Handle POST actions — PM for all actions; Finance limited to update_milestone
    if request.method == 'POST' and (is_assigned_pm or (role == 'Finance' and request.POST.get('action') == 'update_milestone')):
        action = request.POST.get('action', '')

        if action == 'update_milestone':
            milestone_pk = request.POST.get('milestone_pk', '').strip()
            if milestone_pk:
                try:
                    milestone = project.milestones.get(pk=milestone_pk)
                    _actor_role = getattr(getattr(request.user, 'profile', None), 'role', None)

                    if _actor_role == 'Finance':
                        milestone.milestone_description = request.POST.get('milestone_description', '').strip()
                        _save_fields = ['milestone_description']
                        _amount_received_str = request.POST.get('amount_received', '').strip()
                        if _amount_received_str and milestone.status != 'Received':
                            try:
                                milestone.amount_received = Decimal(_amount_received_str)
                                milestone.status        = 'Received'
                                milestone.received_date = date.today()
                                _save_fields += ['amount_received', 'status', 'received_date']
                            except InvalidOperation:
                                pass
                        milestone.save(update_fields=_save_fields)
                        log_activity(project, request.user.profile, f"Updated {milestone.milestone_name}", entity_type='Milestone', entity_id=milestone.pk)
                        if 'status' in _save_fields:
                            _MILESTONE_TO_FINANCE_TASK = {
                                'M1': 'Advance Payment Confirmation',
                                'M2': 'Finance Confirmation',
                                'M3': '100% Payment Confirmation',
                            }
                            _sync_task_name = _MILESTONE_TO_FINANCE_TASK.get(milestone.milestone_name)
                            if _sync_task_name:
                                try:
                                    Task.objects.filter(
                                        phase__project=project,
                                        task_name=_sync_task_name,
                                        status__in=[Task.NOT_STARTED, Task.IN_PROGRESS, Task.BLOCKED],
                                    ).update(status=Task.DONE, completed_at=timezone.now())
                                except Exception:
                                    pass
                    else:
                        milestone.milestone_description = request.POST.get('milestone_description', '').strip()
                        amount_str   = request.POST.get('amount', '').strip()
                        due_date_str = request.POST.get('due_date', '').strip()
                        try:
                            milestone.amount = Decimal(amount_str) if amount_str else None
                        except InvalidOperation:
                            milestone.amount = None
                        try:
                            milestone.due_date = date.fromisoformat(due_date_str) if due_date_str else None
                        except ValueError:
                            milestone.due_date = None
                        milestone.save(update_fields=['milestone_description', 'amount', 'due_date'])
                    messages.success(request, f'{milestone.milestone_name} updated.')
                except PaymentMilestone.DoesNotExist:
                    messages.error(request, 'Milestone not found.')
            return redirect('project_overview', project_id=project.project_id)

        if action == 'assign_design' and can_assign_design:
            design_id = request.POST.get('assigned_design', '').strip()
            if design_id:
                try:
                    design_user = UserProfile.objects.get(pk=design_id, role='Design', is_active=True)
                except UserProfile.DoesNotExist:
                    messages.error(request, 'Invalid design user selected.')
                    return redirect('project_overview', project_id=project.project_id)
                project.assigned_design = design_user
                project.save(update_fields=['assigned_design'])
                _n = Task.objects.filter(
                    phase__project=project,
                    assigned_role=Task.DESIGN,
                    status__in=['Not Started', 'In Progress'],
                ).update(assigned_to=design_user)
                # One summary line for the whole bulk assignment — never per-task.
                _design_name = design_user.user.get_full_name() or design_user.user.username
                log_activity(
                    project, request.user.profile,
                    f"Assigned Design lead {_design_name} to {_n} tasks",
                    entity_type='Project', entity_id=project.pk,
                    action_code='design_bulk_assigned',
                )
            else:
                project.assigned_design = None
                project.save(update_fields=['assigned_design'])
                _n = Task.objects.filter(
                    phase__project=project,
                    assigned_role=Task.DESIGN,
                    status__in=['Not Started', 'In Progress'],
                ).update(assigned_to=None)
                # Clearing the Design lead is also a single summary line.
                log_activity(
                    project, request.user.profile,
                    f"Cleared Design lead from {_n} tasks",
                    entity_type='Project', entity_id=project.pk,
                    action_code='design_bulk_assigned',
                )
            messages.success(request, 'Design member updated.')
            return redirect('project_overview', project_id=project.project_id)

    # BOQ status + last event
    try:
        project.boq_status     = project.boq.status
        project.boq_last_event = project.boq.revisions.order_by('-revised_at').first()
        project.boq_url        = f'/projects/{project.project_id}/boq/'
    except Exception:
        project.boq_status     = None
        project.boq_last_event = None
        project.boq_url        = None

    # BOQ revision history for expanded block
    boq_revision_history = []
    try:
        boq_revision_history = list(
            project.boq.revisions.select_related('revised_by__user').order_by('-revised_at')
        )
    except Exception:
        pass

    # Milestones with variance (positive = short / expected > received; negative = excess)
    milestones = list(project.milestones.all())
    for m in milestones:
        if m.amount is not None and m.amount_received is not None:
            m.variance = m.amount - m.amount_received
            m.variance_abs = abs(m.variance)
        else:
            m.variance = None
            m.variance_abs = None

    # Dict of existing amounts keyed by milestone_name — used by BD modal pre-fill
    milestone_amounts = {m.milestone_name: str(m.amount) if m.amount is not None else '' for m in milestones}

    # Phases + task completion percentages (empty for Draft projects)
    phases = []
    phase_data_json = []
    if project.status != 'Draft':
        phases = list(
            ProjectPhase.objects.filter(project=project)
            .prefetch_related('tasks')
            .order_by('phase_order')
        )
        for phase in phases:
            tasks = list(phase.tasks.all())
            if not tasks:
                continue
            internal       = [t for t in tasks if t.task_type == 'Internal']
            internal_done  = sum(1 for t in internal if t.status == 'Done')
            internal_total = len(internal)
            pct            = int(internal_done / internal_total * 100) if internal_total else 0
            ext_pending    = sum(1 for t in tasks if t.task_type == 'External' and t.status != 'Done')
            phase_data_json.append({
                'pk':             phase.pk,
                'pct':            pct,
                'internal_done':  internal_done,
                'internal_total': internal_total,
                'ext_pending':    ext_pending,
            })

    # Recent activity from ActivityLog (authoritative audit trail).
    # Exclude task_unassigned — the lowest-signal of the new assignment events — so it
    # doesn't crowd status/issue activity out of this 8-item feed. The full trail
    # (including unassigns) is still visible in the per-project timeline and audit log.
    recent_activity = list(
        ActivityLog.objects.filter(project=project)
        .exclude(action_code='task_unassigned')
        .select_related('actor__user')
        .order_by('-timestamp')[:8]
    )

    documents = project.documents.filter(is_deleted=False) if project.status != 'Draft' else []

    project_issues = (
        Issue.objects.filter(project=project)
        .select_related('raised_by__user', 'assigned_to__user', 'task')
    )
    all_profiles = UserProfile.objects.select_related('user').filter(is_active=True).order_by('user__first_name')

    # Delivery challans with per-DC severity-aware summary text
    delivery_challans = list(
        project.delivery_challans
        .select_related('vendor', 'created_by__user')
        .prefetch_related('line_items')
        .all()
    )
    for dc in delivery_challans:
        li_list   = list(dc.line_items.all())
        total     = len(li_list)
        confirmed = [li for li in li_list if li.received_quantity is not None]
        n_conf    = len(confirmed)

        if not total:
            dc.line_items_summary = "No items logged"
        elif n_conf == 0:
            dc.line_items_summary = f"0 of {total} item{'s' if total > 1 else ''} confirmed"
        else:
            total_ordered  = sum(li.ordered_quantity for li in li_list)
            total_received = sum(li.received_quantity for li in confirmed)
            total_damaged  = sum(li.damaged_quantity for li in confirmed)
            # Use the DC's already-computed severity status for the summary label
            if dc.status == DeliveryChallan.RECEIVED:
                dc.line_items_summary = f"All {n_conf} item{'s' if n_conf > 1 else ''} received"
            elif dc.status == DeliveryChallan.REJECTED:
                # Severe: shortfall + damage, or nothing received
                dc.line_items_summary = (
                    f"Severe — {total_received:.0f}/{total_ordered:.0f} units"
                    + (f", {total_damaged} damaged" if total_damaged else "")
                )
            else:
                # Partially Received (AMBER): shortfall or damage, not both stacked
                if total_damaged:
                    dc.line_items_summary = (
                        f"Partial — {total_received:.0f}/{total_ordered:.0f} units, {total_damaged} damaged"
                    )
                else:
                    dc.line_items_summary = (
                        f"Partial — {total_received:.0f} of {total_ordered:.0f} units received"
                    )

    # Per-category material status (summary badge + detail quantities)
    material_status = get_material_status(project)

    cat_rows = (
        DCLineItem.objects.filter(challan__project=project)
        .values('boq_category')
        .annotate(
            total_ordered=Sum('ordered_quantity'),
            total_received=Sum('received_quantity'),
            total_damaged=Sum('damaged_quantity'),
            # Use damaged_quantity (precise numeric) — not the coarse condition string
            damage_count=Count('pk', filter=Q(damaged_quantity__gt=0)),
        )
    )
    material_status_by_category = []
    for row in cat_rows:
        received      = row['total_received'] or 0
        ordered       = row['total_ordered'] or 0
        total_damaged = row['total_damaged'] or 0
        has_damage    = row['damage_count'] > 0
        if received == 0:
            cat_status = 'Pending'
        elif received >= ordered and not has_damage:
            cat_status = 'Received'
        else:
            cat_status = 'Partial'
        material_status_by_category.append({
            'name':              row['boq_category'],
            'status':            cat_status,
            'received_quantity': received,
            'ordered_quantity':  ordered,
            'has_damage':        has_damage,
            'total_damaged':     total_damaged,
        })

    # Design assignment candidates (PM only, for assign-design dropdown)
    # Task.ROLE_CHOICES uses 'BD / Sales' but UserProfile.role stores 'BD'
    _TASK_TO_PROFILE_ROLE = {'BD / Sales': 'BD'}
    candidates_by_role = {}
    design_candidates  = UserProfile.objects.none()
    if is_assigned_pm:
        for role_key, _ in Task.ROLE_CHOICES:
            profile_role = _TASK_TO_PROFILE_ROLE.get(role_key, role_key)
            qs = UserProfile.objects.filter(role=profile_role, is_active=True).select_related('user')
            candidates_by_role[role_key] = [
                {'pk': p.pk, 'name': p.user.get_full_name() or p.user.username}
                for p in qs
            ]
        design_candidates = UserProfile.objects.filter(role='Design', is_active=True).select_related('user')

    dc_vendors = Vendor.objects.filter(is_active=True).order_by('name') if role == 'SCM' else []

    # Normalise UserProfile role → Task.ROLE_CHOICES value for template comparisons
    _PROFILE_TO_TASK_ROLE = {'BD': 'BD / Sales'}
    user_task_role = _PROFILE_TO_TASK_ROLE.get(role, role)

    # Cascade scheduling context — PM-only feature gate check
    _sys = SystemSettings.get()
    show_cascade_option = (_sys.cascade_scheduling_enabled and role == 'PM' and is_assigned_pm)

    # Payment requests for this project — Finance sees confirm actions, PM sees read-only
    # The queryset is shared; template role-gates control which actions are rendered.
    payment_requests = []
    if role in ('Finance', 'PM', 'SCM', 'Admin'):
        payment_requests = list(
            PaymentRequest.objects.filter(project=project)
            .select_related('vendor', 'boq_item', 'requested_by', 'confirmed_by')
            .order_by('-requested_date')
        )

    # ── Gantt (Residential only; computed live from activated_at + duration_days) ──
    # Internal view: every role that can see this project. Client (buffered/friendly)
    # view: PM / Project Coordinator / CEO only — computed and passed ONLY for them so
    # the buffered schedule is never in the DOM for other roles.
    gantt_available     = (project.project_type == 'Residential')
    gantt_not_activated = False
    gantt_internal      = None
    gantt_client        = None
    gantt_can_view_client = role in ('PM', 'Project Coordinator', 'CEO')
    if gantt_available:
        ext_min = _sys.gantt_external_min_display_days
        internal_rows = compute_gantt_schedule(project, 0, ext_min)
        gantt_not_activated = (project.activated_at is None)
        gantt_internal = build_gantt_view(internal_rows)
        if gantt_can_view_client:
            client_rows = compute_gantt_schedule(project, _sys.gantt_client_buffer_days, ext_min)
            gantt_client = build_gantt_view(
                client_rows, GANTT_PHASE_DISPLAY_NAME_MAP, GANTT_TASK_DISPLAY_NAME_MAP,
            )

    return render(request, 'projects/project_overview.html', {
        'project':                     project,
        'milestones':                  milestones,
        'milestone_amounts':           milestone_amounts,
        'phases':                      phases,
        'phase_data_json':             json.dumps(phase_data_json),
        'recent_activity':             recent_activity,
        'role':                        role,
        'user_role':                   role,
        'user_task_role':              user_task_role,
        'user_profile':                profile,
        'documents':                   documents,
        'project_issues':              project_issues,
        'all_profiles':                all_profiles,
        'delivery_challans':           delivery_challans,
        'material_status':             material_status,
        'material_status_by_category': material_status_by_category,
        'dc_vendors':                  dc_vendors,
        'dc_category_choices':         DCLineItem.CATEGORY_CHOICES,
        'dc_condition_choices':        DCLineItem.CONDITION_CHOICES,
        'is_assigned_pm':              is_assigned_pm,
        'can_assign_design':           can_assign_design,
        'design_candidates':           design_candidates,
        'task_status_choices':         Task.STATUS_CHOICES,
        'candidates_by_role':          candidates_by_role,
        'boq_revision_history':        boq_revision_history,
        'today':                       date.today(),
        'payment_requests':            payment_requests,
        'user_dashboard_url':          get_user_dashboard(request.user),
        'show_cascade_option':         show_cascade_option,
        'gantt_available':             gantt_available,
        'gantt_not_activated':         gantt_not_activated,
        'gantt_can_view_client':       gantt_can_view_client,
        'gantt_internal':              gantt_internal,
        'gantt_client':                gantt_client,
    })


# ---------------------------------------------------------------------------
# Zoho CRM Webhook
# ---------------------------------------------------------------------------

def _safe_decimal(value):
    """Strip non-numeric characters and convert to Decimal. Returns None on failure or empty input."""
    if not value:
        return None
    cleaned = re.sub(r'[^\d.]', '', str(value))
    try:
        return Decimal(cleaned)
    except Exception:
        return None


@csrf_exempt
def zoho_deal_closed_webhook(request):
    """
    Receive a Zoho CRM webhook when a deal reaches 'Closed Won'.
    Creates a Draft project from the deal fields.
    Security: token validated from X-Webhook-Token header or ?token= / ?secret= param.
    Duplicate guard: skips if a project with the same zoho_deal_id already exists.
    Always returns HTTP 200 even on application errors — non-200 causes Zoho to retry
    and would create duplicate projects.
    Access: public (unauthenticated), token-protected.
    """
    if request.method != 'POST':
        return HttpResponse(status=405)

    # Token validation — accept header, ?token=, or ?secret= (Zoho default param name)
    token = (
        request.headers.get('X-Webhook-Token', '')
        or request.GET.get('token', '')
        or request.GET.get('secret', '')
    )
    expected = settings.ZOHO_WEBHOOK_SECRET
    if not expected or token != expected:
        ip = request.META.get('REMOTE_ADDR', 'unknown')
        logger.warning('Webhook rejected: invalid token from IP %s', ip)
        return HttpResponse(status=403)

    try:
        payload = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return HttpResponse(status=400)

    # Zoho payload structure varies between webhook versions:
    # flat root (fields at top level) or wrapped in data[]/data[0].Deal
    data_wrapper = payload.get('data')
    if data_wrapper is not None:
        if isinstance(data_wrapper, list):
            data_wrapper = data_wrapper[0] if data_wrapper else {}
        deal = data_wrapper.get('Deal', data_wrapper)
    else:
        deal = payload  # flat root — fields are at the top level

    stage = deal.get('Stage', '')
    # Ignore all stages except Closed Won — return 200 so Zoho doesn't retry
    if stage != 'Closed Won':
        return HttpResponse(status=200)

    record_id = str(deal.get('id', '') or deal.get('Record_Id', '') or deal.get('zoho_deal_id', '')).strip()

    # Duplicate guard
    if record_id and Project.objects.filter(zoho_deal_id=record_id).exists():
        logger.info('Webhook: duplicate deal %s — skipped', record_id)
        return HttpResponse(status=200)

    # Field mapping
    account_name = (deal.get('Account_Name', '') or '').strip()
    contact_name = (deal.get('Contact_Name', '') or '').strip()
    if account_name:
        customer_name = account_name
        customer_contact_person = contact_name
    else:
        customer_name = contact_name
        customer_contact_person = ''

    raw_phone = str(deal.get('Mobile', '') or '').strip()
    digits_only = re.sub(r'\D', '', raw_phone)
    customer_phone = digits_only[-10:] if len(digits_only) >= 10 else digits_only

    raw_date = (deal.get('Closing_Date', '') or '').strip()
    target_commissioning_date = None
    if raw_date:
        try:
            target_commissioning_date = datetime.strptime(raw_date, '%Y-%m-%d').date()
        except ValueError:
            pass

    pm_email = (deal.get('Assign_PM', '') or '').strip()
    assigned_pm = None
    if pm_email:
        profile_match = UserProfile.objects.filter(user__email__iexact=pm_email).first()
        if profile_match:
            assigned_pm = profile_match

    # Create project
    try:
        project = Project.objects.create(
            customer_name=customer_name or 'Unknown',
            customer_contact_person=customer_contact_person,
            customer_phone=customer_phone,
            customer_email=(deal.get('Customer_Email', '') or '').strip() or None,
            site_address='',
            city=(deal.get('City', '') or '').strip(),
            state=(deal.get('State', '') or '').strip() or 'Uttar Pradesh',
            project_type='Residential',
            capacity_kw=_safe_decimal(deal.get('Capacity_kW') or deal.get('Capacity')),
            contract_value=_safe_decimal(deal.get('Amount')),
            assigned_pm=assigned_pm,
            target_commissioning_date=target_commissioning_date,
            status='Draft',
            zoho_deal_id=record_id,
            created_by=None,
        )
    except Exception as exc:
        logger.error('Webhook: project creation failed for deal %s — %s', record_id, exc)
        # Return 200 even on failure — non-200 would cause Zoho to retry and create duplicates
        return HttpResponse(status=200)

    logger.info('Webhook: project %s created for deal %s (pm=%s)',
                project.project_id, record_id,
                project.assigned_pm.user.email if project.assigned_pm else 'unassigned')

    try:
        ActivityLog.objects.create(
            project=project,
            actor=None,
            action=f"Project '{project.project_id}' created via Zoho CRM webhook",
            entity_type='Project',
            entity_id=project.pk,
        )
    except Exception:
        pass

    if project.assigned_pm is None:
        # Notify Admin in-app: PM could not be assigned
        try:
            admin_profile = UserProfile.objects.filter(role='Admin').first()
            if admin_profile:
                send_notification(
                    recipient=admin_profile,
                    message=(
                        f'New Draft project {project.project_id} created from Zoho CRM '
                        f'(deal {record_id}) — PM could not be assigned. Please assign a PM.'
                    ),
                    channels=['in_app'],
                    link=f'/projects/{project.project_id}/overview/',
                    related_project=project,
                )
        except Exception as exc:
            logger.error('Webhook: in-app notification failed for project %s — %s', project.project_id, exc)
        # Also email a fixed admin address so the alert lands even if no Admin is logged in
        try:
            send_raw_email(
                to_email='smzk07@gmail.com',
                subject=f'[SolarPMS] Unassigned Project: {project.project_id}',
                body=(
                    f'A new Draft project was created via Zoho CRM but no PM could be assigned.\n\n'
                    f'Project ID : {project.project_id}\n'
                    f'Customer   : {project.customer_name}\n'
                    f'City       : {project.city or "—"}\n'
                    f'Zoho Deal  : {record_id}\n'
                    f'PM email from Zoho: {pm_email or "(blank)"}\n\n'
                    f'Please log in to the Admin Panel and assign a PM:\n'
                    f'https://horizon-solar-pms-production.up.railway.app/projects/{project.project_id}/overview/'
                ),
            )
        except Exception as exc:
            logger.error('Webhook: admin alert email failed for project %s — %s', project.project_id, exc)
    else:
        # Notify the assigned PM about their new project
        try:
            pm_display_name = project.assigned_pm.user.get_full_name() or project.assigned_pm.user.username
            project_link = reverse('project_overview', args=[project.pk])
            project_url_abs = request.build_absolute_uri(project_link)
            _ap_link = f'/projects/{project.project_id}/'
            _ap_message = (
                f'Hi {pm_display_name},\n\n'
                f'You have been assigned as Project Manager for {project.customer_name} '
                f'({project.city}).\n\n'
                f'Please review the project details and begin onboarding.'
            )
            _ap_email_message = (
                f'{_ap_message}\n\nView in Horizon Solar PMS:\n'
                f'https://horizon-solar-pms-production.up.railway.app{_ap_link}'
            )
            send_notification(
                recipient=project.assigned_pm,
                message=_ap_email_message,
                channels=['in_app', 'whatsapp', 'email'],
                link=_ap_link,
                subject=f'New Project Assigned: {project.customer_name}',
                template='assign_project',
                template_params=[
                    project.customer_name,    # [0] header
                    pm_display_name,          # [1] body[0] — user_name
                    project_url_abs,          # [2] body[1] — project_url
                ],
                related_project=project,
            )
        except Exception as exc:
            logger.error('Webhook: PM notification failed for project %s — %s', project.project_id, exc)

    return HttpResponse(status=200)


# ---------------------------------------------------------------------------
# File upload helpers
# ---------------------------------------------------------------------------


def _validate_and_upload(file, supabase_client, bucket, supabase_path, allowed_extensions=None):
    """Validate one file and upload to Supabase. Raises ValueError on validation failure.
    Pass allowed_extensions (e.g. ALLOWED_PHOTO_EXTENSIONS) to restrict the accepted types
    for this call; defaults to ALLOWED_EXTENSIONS, preserving every existing caller."""
    allowed = allowed_extensions if allowed_extensions is not None else ALLOWED_EXTENSIONS
    ext = file.name.rsplit('.', 1)[-1].lower() if '.' in file.name else ''
    if ext not in allowed:
        raise ValueError(f"unsupported type (.{ext})")
    if file.size > MAX_FILE_SIZE_BYTES:
        raise ValueError("exceeds 20 MB limit")

    expected_mime = MIME_TYPE_MAP.get(ext, '')
    actual_mime   = (file.content_type or '').split(';')[0].strip()
    if expected_mime and actual_mime and actual_mime not in (expected_mime, 'application/octet-stream'):
        raise ValueError("MIME type does not match extension")

    file.seek(0)
    supabase_client.storage.from_(bucket).upload(
        path=supabase_path,
        file=file.read(),
        file_options={"content-type": MIME_TYPE_MAP.get(ext, 'application/octet-stream')},
    )
    return ext


# ---------------------------------------------------------------------------
# Task detail
# ---------------------------------------------------------------------------

@login_required
def task_detail(request, project_id, task_id):
    """
    Task detail page: attachments, issues, and threaded comments for this task.
    Access: all roles; PM isolation applies.
    """
    project = get_object_or_404(Project, project_id=project_id)
    profile = request.user.profile

    # PM isolation: PM sees only their own projects
    if profile.role == 'PM' and not user_can_manage_project(request.user, project):
        raise Http404

    task        = get_object_or_404(Task, pk=task_id, phase__project=project)
    attachments = task.attachments.filter(is_deleted=False)
    task_issues = (
        Issue.objects.filter(task=task)
        .select_related('raised_by__user', 'assigned_to__user', 'task')
    )
    all_profiles = UserProfile.objects.select_related('user').filter(is_active=True).order_by('user__first_name')

    # Prefetch replies to avoid N+1 — template iterates comment.replies.all()
    task_comments = (
        Comment.objects.filter(task=task, parent=None)
        .select_related('author__user')
        .prefetch_related(
            Prefetch('replies', queryset=Comment.objects.select_related('author__user'))
        )
    )

    is_assignee = task.assigned_to is not None and task.assigned_to == profile

    context = {
        'project':            project,
        'task':               task,
        'attachments':        attachments,
        'user_profile':       profile,
        'task_issues':        task_issues,
        'all_profiles':       all_profiles,
        'task_comments':      task_comments,
        'is_assignee':        is_assignee,
        'task_status_choices': Task.STATUS_CHOICES,
    }
    # Checklist — items come from the Checklist linked to this (task_name, project_type);
    # completion is per-(item, task). Shared with the HTMX swap via _checklist_context().
    context.update(_checklist_context(request, project, task))
    return render(request, 'projects/task_detail.html', context)


# ---------------------------------------------------------------------------
# Project document upload / delete
# ---------------------------------------------------------------------------

@login_required
def upload_project_document(request, project_id):
    """
    Upload one or more files to a project. Files are stored in Supabase under
    project-documents/{project_id}/{uuid}_{filename}. DB record is created after
    a successful Supabase upload. Partial success (some files uploaded) shows a warning.
    Access: all roles with project access; PM isolation applies.
    """
    if request.method != 'POST':
        return redirect('project_overview', project_id=project_id)

    project = get_object_or_404(Project, project_id=project_id)
    profile = request.user.profile

    if profile.role == 'PM' and not user_can_manage_project(request.user, project):
        raise Http404

    files = request.FILES.getlist('files')
    if not files:
        messages.error(request, 'No files selected.')
        return redirect('project_overview', project_id=project_id)

    try:
        from .supabase_storage import get_supabase_client
        client = get_supabase_client()
    except (ValueError, ImportError) as exc:
        messages.error(request, f"Upload service unavailable. Contact Admin. ({exc})")
        return redirect('project_overview', project_id=project_id)

    bucket     = settings.SUPABASE_BUCKET
    successes  = []
    failures   = []

    for file in files:
        supabase_path = (
            f"project-documents/{project.project_id}/"
            f"{_uuid.uuid4()}_{file.name}"
        )
        try:
            ext       = _validate_and_upload(file, client, bucket, supabase_path)
            file_type = 'Photo' if ext in ALLOWED_PHOTO_EXTENSIONS else 'Document'
            file_url  = (
                f"{settings.SUPABASE_URL}/storage/v1/object/public/"
                f"{settings.SUPABASE_BUCKET}/{supabase_path}"
            )
            ProjectDocument.objects.create(
                project=project,
                uploaded_by=profile,
                file_name=file.name,
                file_url=file_url,
                supabase_path=supabase_path,
                file_type=file_type,
                file_size_kb=max(1, file.size // 1024),
            )
            successes.append(file.name)
        except ValueError as exc:
            failures.append(f"{file.name} ({exc})")
        except Exception as exc:
            logger.error('Supabase upload failed for %s: %s', file.name, exc)
            failures.append(f"{file.name} (upload error)")

    # Build outcome summary
    if successes and failures:
        msg   = (f"{len(successes)} of {len(successes)+len(failures)} files uploaded. "
                 f"Failed: {', '.join(failures)}")
        is_ok = True
        log_activity(project, profile, f"Uploaded {len(successes)} file(s) to project",
                     entity_type='File', entity_id=None)
    elif successes:
        msg   = f"{len(successes)} file(s) uploaded successfully."
        is_ok = True
        log_activity(project, profile, f"Uploaded {len(successes)} file(s) to project",
                     entity_type='File', entity_id=None)
    else:
        msg   = f"No files uploaded. Failed: {', '.join(failures)}"
        is_ok = False

    # AJAX request — return JSON so the browser can show inline feedback
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': is_ok, 'message': msg})

    # Regular POST fallback — Django messages + redirect
    if successes and failures:
        messages.warning(request, msg)
    elif successes:
        messages.success(request, msg)
    else:
        messages.error(request, msg)

    next_url = request.POST.get('next', '')
    if next_url and not _urlparse(next_url).netloc:
        return redirect(next_url)
    return redirect('project_overview', project_id=project_id)


@login_required
def delete_project_document(request, project_id, doc_pk):
    """
    Soft-delete a project document (sets is_deleted=True). The Supabase file is not
    removed here — purge_deleted_files handles hard deletion after FILE_RETENTION_DAYS.
    Access: uploader or Admin only. POST only.
    """
    if request.method != 'POST':
        return redirect('project_overview', project_id=project_id)

    project = get_object_or_404(Project, project_id=project_id)
    profile = request.user.profile

    if profile.role == 'PM' and not user_can_manage_project(request.user, project):
        raise Http404

    doc = get_object_or_404(ProjectDocument, pk=doc_pk, project=project, is_deleted=False)

    if doc.uploaded_by != profile and profile.role != 'Admin':
        return HttpResponseForbidden()

    doc.is_deleted  = True
    doc.deleted_at  = timezone.now()
    doc.deleted_by  = profile
    doc.save(update_fields=['is_deleted', 'deleted_at', 'deleted_by'])

    # Log the project document deletion
    log_activity(project, profile, f"Deleted file: {doc.file_name}", entity_type='File', entity_id=doc.pk)
    messages.success(request, f'"{doc.file_name}" deleted.')
    next_url = request.POST.get('next', '')
    if next_url and not _urlparse(next_url).netloc:
        return redirect(next_url)
    return redirect('project_overview', project_id=project_id)


# ---------------------------------------------------------------------------
# Task attachment upload / delete
# ---------------------------------------------------------------------------

@login_required
def upload_task_attachment(request, project_id, task_id):
    """
    Upload one or more files to a task. Stored in Supabase under
    task-attachments/{project_id}/{task_pk}/{uuid}_{filename}.
    Cross-project guard prevents uploading to a task on a different project.
    Access: all roles with project access; PM isolation applies.
    """
    if request.method != 'POST':
        return redirect('task_detail', project_id=project_id, task_id=task_id)

    project = get_object_or_404(Project, project_id=project_id)
    profile = request.user.profile

    if profile.role == 'PM' and not user_can_manage_project(request.user, project):
        raise Http404

    task = get_object_or_404(Task, pk=task_id, phase__project=project)

    # Cross-project guard
    if task.phase.project.project_id != project_id:
        return HttpResponseForbidden()

    files = request.FILES.getlist('files')
    if not files:
        messages.error(request, 'No files selected.')
        if _is_hx(request):
            return _render_attachments_hx(request, project, task)
        return redirect('task_detail', project_id=project_id, task_id=task_id)

    try:
        from .supabase_storage import get_supabase_client
        client = get_supabase_client()
    except (ValueError, ImportError) as exc:
        messages.error(request, f"Upload service unavailable. Contact Admin. ({exc})")
        if _is_hx(request):
            return _render_attachments_hx(request, project, task)
        return redirect('task_detail', project_id=project_id, task_id=task_id)

    bucket    = settings.SUPABASE_BUCKET
    successes = []
    failures  = []

    for file in files:
        supabase_path = (
            f"task-attachments/{project.project_id}/{task.pk}/"
            f"{_uuid.uuid4()}_{file.name}"
        )
        try:
            ext       = _validate_and_upload(file, client, bucket, supabase_path)
            file_type = 'Photo' if ext in ALLOWED_PHOTO_EXTENSIONS else 'Document'
            file_url  = (
                f"{settings.SUPABASE_URL}/storage/v1/object/public/"
                f"{settings.SUPABASE_BUCKET}/{supabase_path}"
            )
            TaskAttachment.objects.create(
                task=task,
                uploaded_by=profile,
                file_name=file.name,
                file_url=file_url,
                supabase_path=supabase_path,
                file_type=file_type,
                file_size_kb=max(1, file.size // 1024),
            )
            successes.append(file.name)
        except ValueError as exc:
            failures.append(f"{file.name} ({exc})")
        except Exception as exc:
            logger.error('Supabase upload failed for %s: %s', file.name, exc)
            failures.append(f"{file.name} (upload error)")

    if successes and failures:
        messages.warning(
            request,
            f"{len(successes)} of {len(successes) + len(failures)} files uploaded. "
            f"Failed: {', '.join(failures)}"
        )
        # Log partial upload — at least some files were stored
        log_activity(project, profile, f"Uploaded {len(successes)} file(s) to task: {task.task_name}", entity_type='File', entity_id=task.pk)
    elif successes:
        messages.success(request, f"{len(successes)} file(s) uploaded successfully.")
        # Log successful task attachment upload
        log_activity(project, profile, f"Uploaded {len(successes)} file(s) to task: {task.task_name}", entity_type='File', entity_id=task.pk)
    else:
        messages.error(request, f"No files uploaded. Failed: {', '.join(failures)}")

    if _is_hx(request):
        return _render_attachments_hx(request, project, task)

    next_url = request.POST.get('next', '')
    if next_url and not _urlparse(next_url).netloc:
        return redirect(next_url)
    return redirect('task_detail', project_id=project_id, task_id=task_id)


@login_required
def delete_task_attachment(request, project_id, task_id, attach_pk):
    """
    Soft-delete a task attachment. Supabase file is purged later by purge_deleted_files.
    Access: uploader or Admin only. POST only.
    """
    if request.method != 'POST':
        return redirect('task_detail', project_id=project_id, task_id=task_id)

    project = get_object_or_404(Project, project_id=project_id)
    profile = request.user.profile

    if profile.role == 'PM' and not user_can_manage_project(request.user, project):
        raise Http404

    task   = get_object_or_404(Task, pk=task_id, phase__project=project)
    attach = get_object_or_404(TaskAttachment, pk=attach_pk, task=task, is_deleted=False)

    if attach.uploaded_by != profile and profile.role != 'Admin':
        return HttpResponseForbidden()

    attach.is_deleted = True
    attach.deleted_at = timezone.now()
    attach.deleted_by = profile
    attach.save(update_fields=['is_deleted', 'deleted_at', 'deleted_by'])

    # Log the task attachment deletion
    log_activity(project, profile, f"Deleted file: {attach.file_name}", entity_type='File', entity_id=attach.pk)
    messages.success(request, f'"{attach.file_name}" deleted.')

    if _is_hx(request):
        return _render_attachments_hx(request, project, task)

    next_url = request.POST.get('next', '')
    if next_url and not _urlparse(next_url).netloc:
        return redirect(next_url)
    return redirect('task_detail', project_id=project_id, task_id=task_id)


# ---------------------------------------------------------------------------
# Checklist completion — on the task-detail page
#
# The checklist itself (name + items) is authored in portal-admin and assigned to a
# task via ChecklistTaskLink; item CRUD lives there, NOT here. On task detail the only
# action is COMPLETION: tick an item + upload its required photo, writing a
# ChecklistItemCompletion row keyed by (item, task). Completion is gated by
# _user_can_complete_checklist_item (role-match OR PM/coordinator) and swaps
# #checklistSection via _render_checklist_hx. The task-closing gate is intentionally
# NOT here — that is a later change.
# ---------------------------------------------------------------------------

def _checklist_error(request, project, task, msg):
    """Surface a checklist error: HTMX → re-render the section with the flash inline;
    non-HTMX → message + redirect to the task detail page."""
    messages.error(request, msg)
    if _is_hx(request):
        return _render_checklist_hx(request, project, task)
    return redirect('task_detail', project_id=project.project_id, task_id=task.pk)


@login_required
def checklist_item_complete(request, project_id, task_id, item_id):
    """Tick one checklist item on this task AND upload its required photo as one atomic
    action, writing a ChecklistItemCompletion row keyed by (item, task). is_checked is set
    True only together with the three photo_* fields in the same save — a checked item can
    never lack a photo. The item must belong to the Checklist assigned to this task.
    Access: role-match OR PM/coordinator (_user_can_complete_checklist_item). POST only."""
    if request.method != 'POST':
        return redirect('task_detail', project_id=project_id, task_id=task_id)

    project = get_object_or_404(Project, project_id=project_id)
    task    = get_object_or_404(Task, pk=task_id, phase__project=project)

    # The item must belong to the checklist currently linked to this task — never trust a
    # raw item_id. _checklist_for_task enforces the (task_name, project_type) link + active.
    checklist = _checklist_for_task(task, project)
    if checklist is None:
        return _checklist_error(request, project, task, 'No checklist is assigned to this task.')
    item = get_object_or_404(ChecklistItem, pk=item_id, checklist=checklist)

    if not _user_can_complete_checklist_item(request.user, task, project):
        if _is_hx(request):
            return _checklist_error(request, project, task,
                                    'You do not have permission to complete items on this task.')
        return HttpResponseForbidden()

    profile = request.user.profile

    # Already completed on this task? Idempotent no-op (the template hides the form once done).
    existing = ChecklistItemCompletion.objects.filter(item=item, task=task).first()
    if existing is not None and existing.is_checked:
        if _is_hx(request):
            return _render_checklist_hx(request, project, task)
        return redirect('task_detail', project_id=project_id, task_id=task_id)

    photo = request.FILES.get('photo')
    if not photo:
        return _checklist_error(request, project, task,
                                'A photo is required to check this item.')

    try:
        from .supabase_storage import get_supabase_client
        client = get_supabase_client()
    except (ValueError, ImportError) as exc:
        return _checklist_error(request, project, task,
                                f"Upload service unavailable. Contact Admin. ({exc})")

    bucket = settings.SUPABASE_BUCKET
    supabase_path = (
        f"checklist-photos/{project.project_id}/{task.pk}/{item.pk}/"
        f"{_uuid.uuid4()}_{photo.name}"
    )
    try:
        # Photo-only allow-list — reuses the shared helper (audit Point 2)
        _validate_and_upload(photo, client, bucket, supabase_path,
                             allowed_extensions=ALLOWED_PHOTO_EXTENSIONS)
    except ValueError as exc:
        return _checklist_error(request, project, task, f"Photo rejected: {exc}")
    except Exception as exc:
        logger.error('Supabase checklist photo upload failed for %s: %s', photo.name, exc)
        return _checklist_error(request, project, task, 'Photo upload failed. Please try again.')

    file_url = (
        f"{settings.SUPABASE_URL}/storage/v1/object/public/"
        f"{settings.SUPABASE_BUCKET}/{supabase_path}"
    )

    # Atomic completion: tick + all three photo fields written together on the (item, task) row.
    completion, _created = ChecklistItemCompletion.objects.get_or_create(item=item, task=task)
    completion.photo_file_name     = photo.name
    completion.photo_url           = file_url
    completion.photo_supabase_path = supabase_path
    completion.is_checked          = True
    completion.checked_by          = request.user
    completion.checked_at          = timezone.now()
    completion.save(update_fields=[
        'photo_file_name', 'photo_url', 'photo_supabase_path',
        'is_checked', 'checked_by', 'checked_at',
    ])

    log_activity(project, profile,
                 f"Completed checklist item '{item.label}' on task: {task.task_name}",
                 entity_type='ChecklistItemCompletion', entity_id=completion.pk)
    messages.success(request, 'Checklist item checked.')

    if _is_hx(request):
        return _render_checklist_hx(request, project, task)
    return redirect('task_detail', project_id=project_id, task_id=task_id)


# ---------------------------------------------------------------------------
# Issues
# ---------------------------------------------------------------------------

def _issue_base_qs():
    """Base Issue queryset with actor relations pre-joined — used in issue detail views."""
    return Issue.objects.select_related('raised_by__user', 'assigned_to__user', 'task')


def _is_project_pm(profile, project):
    """Return True if profile manages this project (PM or Project Coordinator).

    Used for issue close/reopen guards. Full drizzle-down: coordinators get the same
    issue-lifecycle authority as the PM on their projects. The ownership comparison
    routes through the canonical user_can_manage_project(); the role gate is kept
    (and now includes Project Coordinator) so the pathological webhook case — where
    assigned_pm could point to a non-manager profile — is still excluded.
    """
    return profile.role in ('PM', 'Project Coordinator') and user_can_manage_project(profile.user, project)


@login_required
def create_project_issue(request, project_id):
    """
    Raise a project-level issue (not tied to any specific task).
    Access: all roles with project access; PM isolation applies. POST only.
    """
    if request.method != 'POST':
        return redirect('project_overview', project_id=project_id)

    project = get_object_or_404(Project, project_id=project_id)
    profile = request.user.profile

    if profile.role == 'PM' and not user_can_manage_project(request.user, project):
        raise Http404

    title       = request.POST.get('title', '').strip()
    description = request.POST.get('description', '').strip()
    severity    = request.POST.get('severity', Issue.MEDIUM)
    due_date_s  = request.POST.get('due_date', '').strip()
    assignee_id = request.POST.get('assigned_to', '').strip()

    if not title:
        messages.error(request, 'Issue title is required.')
        return redirect('project_overview', project_id=project_id)

    if severity not in dict(Issue.SEVERITY_CHOICES):
        severity = Issue.MEDIUM

    due_date = None
    if due_date_s:
        try:
            due_date = date.fromisoformat(due_date_s)
        except ValueError:
            pass

    assigned_to = None
    if assignee_id:
        try:
            assigned_to = UserProfile.objects.get(pk=assignee_id)
        except UserProfile.DoesNotExist:
            pass

    issue = Issue.objects.create(
        project=project,
        task=None,
        title=title,
        description=description,
        severity=severity,
        status=Issue.OPEN,
        raised_by=profile,
        assigned_to=assigned_to,
        due_date=due_date,
    )
    log_activity(project, profile, f"Raised issue: {title} ({severity})", entity_type='Issue', entity_id=issue.pk, action_code='issue_created')
    if assigned_to and assigned_to != profile:
        raiser_name = profile.user.get_full_name() or profile.user.username
        recipient_name = assigned_to.user.get_full_name() or assigned_to.user.username
        _ic_link = f'/issues/{issue.pk}/'
        _ic_message = (
            f'A new issue has been raised on project {project.customer_name}: '
            f'"{title}".\n\n'
            f'Please login to review and coordinate resolution.'
        )
        _ic_email_message = (
            f'{_ic_message}\n\nView in Horizon Solar PMS:\n'
            f'https://horizon-solar-pms-production.up.railway.app{_ic_link}'
        )
        send_notification(
            recipient=assigned_to,
            message=_ic_email_message,
            channels=['in_app', 'whatsapp', 'email'],
            link=_ic_link,
            subject=f'New Issue Raised — {project.customer_name}',
            template='issue_created',
            template_params=[project.customer_name, recipient_name, project.customer_name, raiser_name],
            related_project=project,
            actor=profile,
        )
    # Notify every project manager (PM + coordinators), skipping the raiser and the
    # assignee (who are notified through other paths). project_managers() dedupes.
    for _mgr in project_managers(project):
        if _mgr != profile and _mgr != assigned_to:
            send_notification(
                recipient=_mgr,
                message=f'Issue "{title}" raised on {project.project_id} — {project.customer_name}.',
                channels=['in_app'],
                link=f'/issues/{issue.pk}/',
                related_project=project,
                actor=profile,
            )
    messages.success(request, f'Issue "{title}" raised successfully.')
    return redirect('project_overview', project_id=project_id)


@login_required
def create_task_issue(request, project_id, task_id):
    """
    Raise an issue linked to a specific task (task.issues relation).
    Access: all roles with project access; PM isolation applies. POST only.
    """
    if request.method != 'POST':
        return redirect('task_detail', project_id=project_id, task_id=task_id)

    project = get_object_or_404(Project, project_id=project_id)
    profile = request.user.profile

    if profile.role == 'PM' and not user_can_manage_project(request.user, project):
        raise Http404

    task = get_object_or_404(Task, pk=task_id, phase__project=project)

    title       = request.POST.get('title', '').strip()
    description = request.POST.get('description', '').strip()
    severity    = request.POST.get('severity', Issue.MEDIUM)
    due_date_s  = request.POST.get('due_date', '').strip()
    assignee_id = request.POST.get('assigned_to', '').strip()

    if not title:
        messages.error(request, 'Issue title is required.')
        return redirect('task_detail', project_id=project_id, task_id=task_id)

    if severity not in dict(Issue.SEVERITY_CHOICES):
        severity = Issue.MEDIUM

    due_date = None
    if due_date_s:
        try:
            due_date = date.fromisoformat(due_date_s)
        except ValueError:
            pass

    assigned_to = None
    if assignee_id:
        try:
            assigned_to = UserProfile.objects.get(pk=assignee_id)
        except UserProfile.DoesNotExist:
            pass

    issue = Issue.objects.create(
        project=project,
        task=task,
        title=title,
        description=description,
        severity=severity,
        status=Issue.OPEN,
        raised_by=profile,
        assigned_to=assigned_to,
        due_date=due_date,
    )
    log_activity(project, profile, f"Raised issue: {title} ({severity})", entity_type='Issue', entity_id=issue.pk, action_code='issue_created')
    if assigned_to and assigned_to != profile:
        raiser_name = profile.user.get_full_name() or profile.user.username
        recipient_name = assigned_to.user.get_full_name() or assigned_to.user.username
        _ic_link = f'/issues/{issue.pk}/'
        _ic_message = (
            f'A new issue has been raised on project {project.customer_name}: '
            f'"{title}".\n\n'
            f'Please login to review and coordinate resolution.'
        )
        _ic_email_message = (
            f'{_ic_message}\n\nView in Horizon Solar PMS:\n'
            f'https://horizon-solar-pms-production.up.railway.app{_ic_link}'
        )
        send_notification(
            recipient=assigned_to,
            message=_ic_email_message,
            channels=['in_app', 'whatsapp', 'email'],
            link=_ic_link,
            subject=f'New Issue Raised — {project.customer_name}',
            template='issue_created',
            template_params=[project.customer_name, recipient_name, project.customer_name, raiser_name],
            related_project=project,
            actor=profile,
        )
    # Notify every project manager (PM + coordinators), skipping the raiser and the
    # assignee (who are notified through other paths). project_managers() dedupes.
    for _mgr in project_managers(project):
        if _mgr != profile and _mgr != assigned_to:
            send_notification(
                recipient=_mgr,
                message=f'Issue "{title}" raised on {project.project_id} — {project.customer_name}.',
                channels=['in_app'],
                link=f'/issues/{issue.pk}/',
                related_project=project,
                actor=profile,
            )
    messages.success(request, f'Issue "{title}" raised.')
    return redirect('task_detail', project_id=project_id, task_id=task_id)


@login_required
def create_delivery_issue(request, project_id, dc_id):
    """
    Raise an issue linked to a specific DeliveryChallan (delivery_challan.issues relation).
    Follows the same pattern as create_task_issue but scoped to DC instead of Task.
    Access: all roles with project access; PM isolation applies. POST only.
    """
    if request.method != 'POST':
        return redirect('delivery_challan_detail', project_id=project_id, dc_id=dc_id)

    project = get_object_or_404(Project, project_id=project_id)
    profile = request.user.profile

    # PM isolation: PMs can only interact with their own projects
    if profile.role == 'PM' and not user_can_manage_project(request.user, project):
        raise Http404

    # Cross-project guard: DC must belong to the project in the URL
    challan = get_object_or_404(DeliveryChallan, pk=dc_id)
    if challan.project.project_id != project_id:
        raise Http404

    title       = request.POST.get('title', '').strip()
    description = request.POST.get('description', '').strip()
    severity    = request.POST.get('severity', Issue.MEDIUM)
    due_date_s  = request.POST.get('due_date', '').strip()
    assignee_id = request.POST.get('assigned_to', '').strip()

    if not title:
        messages.error(request, 'Issue title is required.')
        return redirect('delivery_challan_detail', project_id=project_id, dc_id=dc_id)

    if severity not in dict(Issue.SEVERITY_CHOICES):
        severity = Issue.MEDIUM

    due_date = None
    if due_date_s:
        try:
            due_date = date.fromisoformat(due_date_s)
        except ValueError:
            pass

    assigned_to = None
    if assignee_id:
        try:
            assigned_to = UserProfile.objects.get(pk=assignee_id)
        except UserProfile.DoesNotExist:
            pass

    issue = Issue.objects.create(
        project=project,
        delivery_challan=challan,
        task=None,
        title=title,
        description=description,
        severity=severity,
        status=Issue.OPEN,
        raised_by=profile,
        assigned_to=assigned_to,
        due_date=due_date,
    )
    log_activity(
        project, profile,
        f"Raised delivery issue: {title} ({severity}) on DC {challan.dc_number}",
        entity_type='Issue', entity_id=issue.pk, action_code='issue_created',
    )
    if assigned_to and assigned_to != profile:
        raiser_name = profile.user.get_full_name() or profile.user.username
        recipient_name = assigned_to.user.get_full_name() or assigned_to.user.username
        _ic_link = f'/issues/{issue.pk}/'
        _ic_message = (
            f'A new issue has been raised on project {project.customer_name}: '
            f'"{title}".\n\n'
            f'Please login to review and coordinate resolution.'
        )
        _ic_email_message = (
            f'{_ic_message}\n\nView in Horizon Solar PMS:\n'
            f'https://horizon-solar-pms-production.up.railway.app{_ic_link}'
        )
        send_notification(
            recipient=assigned_to,
            message=_ic_email_message,
            channels=['in_app', 'whatsapp', 'email'],
            link=_ic_link,
            subject=f'New Issue Raised — {project.customer_name}',
            template='issue_created',
            template_params=[project.customer_name, recipient_name, project.customer_name, raiser_name],
            related_project=project,
            actor=profile,
        )
    # Notify every project manager (PM + coordinators), skipping the raiser and the
    # assignee (who are notified through other paths). project_managers() dedupes.
    for _mgr in project_managers(project):
        if _mgr != profile and _mgr != assigned_to:
            send_notification(
                recipient=_mgr,
                message=f'Issue "{title}" raised on {project.project_id} — {project.customer_name}.',
                channels=['in_app'],
                link=f'/issues/{issue.pk}/',
                related_project=project,
                actor=profile,
            )
    messages.success(request, f'Issue "{title}" raised for DC {challan.dc_number}.')
    return redirect('delivery_challan_detail', project_id=project_id, dc_id=dc_id)


@login_required
def issue_detail(request, issue_id):
    """
    Issue detail page with comments and status/assignee controls.
    Access: all roles; PM isolation applies.
    """
    issue   = get_object_or_404(_issue_base_qs(), pk=issue_id)
    project = issue.project
    profile = request.user.profile

    if profile.role == 'PM' and not user_can_manage_project(request.user, project):
        raise Http404

    all_profiles = UserProfile.objects.select_related('user').filter(is_active=True).order_by('user__first_name')
    is_pm        = _is_project_pm(profile, project)

    # Prefetch replies to avoid N+1 — template iterates comment.replies.all()
    issue_comments = (
        Comment.objects.filter(issue=issue, parent=None)
        .select_related('author__user')
        .prefetch_related(
            Prefetch('replies', queryset=Comment.objects.select_related('author__user'))
        )
    )

    return render(request, 'projects/issue_detail.html', {
        'issue':          issue,
        'project':        project,
        'user_profile':   profile,
        'all_profiles':   all_profiles,
        'is_pm':          is_pm,
        'issue_comments': issue_comments,
    })


@login_required
def update_issue_status(request, issue_id):
    """
    Advance an Open issue to In Progress. Requires the issue to have an assignee.
    filter().update() used to prevent race condition on concurrent status changes.
    Access: all roles with project access. POST only.
    """
    if request.method != 'POST':
        return redirect('issue_detail', issue_id=issue_id)

    issue   = get_object_or_404(Issue, pk=issue_id)
    project = issue.project
    profile = request.user.profile

    if profile.role == 'PM' and not user_can_manage_project(request.user, project):
        raise Http404

    if issue.status == Issue.CLOSED:
        return HttpResponseForbidden('Issue is closed.')

    if issue.status != Issue.OPEN:
        messages.warning(request, 'Issue is not in Open status.')
        return redirect('issue_detail', issue_id=issue_id)

    if issue.assigned_to is None:
        messages.warning(request, 'Please assign this issue before starting work.')
        return redirect('issue_detail', issue_id=issue_id)

    # filter().update() with status condition prevents race condition — updated==0 means
    # another request already changed the status between our read and this write
    updated = Issue.objects.filter(pk=issue.pk, status=Issue.OPEN).update(status=Issue.IN_PROGRESS)
    if updated == 0:
        messages.warning(request, 'Issue status was already updated.')
    else:
        log_activity(project, profile, f"Moved issue to In Progress: {issue.title}", entity_type='Issue', entity_id=issue.pk)
        messages.success(request, 'Issue moved to In Progress.')
    return redirect('issue_detail', issue_id=issue_id)


@login_required
def resolve_issue(request, issue_id):
    """
    Mark an In Progress issue as Resolved. Resolution note is required.
    Notifies the PM (if they are not the resolver) so they can review and close.
    filter().update() used to prevent race condition on concurrent status changes.
    Access: any role with project access. POST only.
    """
    if request.method != 'POST':
        return redirect('issue_detail', issue_id=issue_id)

    issue   = get_object_or_404(Issue, pk=issue_id)
    project = issue.project
    profile = request.user.profile

    if profile.role == 'PM' and not user_can_manage_project(request.user, project):
        raise Http404

    if issue.status == Issue.CLOSED:
        return HttpResponseForbidden('Issue is closed.')

    if issue.status != Issue.IN_PROGRESS:
        messages.warning(request, 'Issue must be In Progress to resolve.')
        return redirect('issue_detail', issue_id=issue_id)

    resolution_note = request.POST.get('resolution_note', '').strip()
    if not resolution_note:
        messages.error(request, 'Resolution note is required.')
        return redirect('issue_detail', issue_id=issue_id)

    updated = Issue.objects.filter(pk=issue.pk, status=Issue.IN_PROGRESS).update(
        status=Issue.RESOLVED,
        resolved_at=timezone.now(),
        resolution_note=resolution_note,
    )
    if updated == 0:
        messages.warning(request, 'Issue status was already updated.')
    else:
        log_activity(project, profile, f"Resolved issue: {issue.title}", entity_type='Issue', entity_id=issue.pk, action_code='issue_resolved')
        resolver_name = profile.user.get_full_name() or profile.user.username
        issue_link = f'/issues/{issue.pk}/'
        issue_link_abs = request.build_absolute_uri(issue_link)
        resolved_params = [project.customer_name, issue.title, resolver_name, issue_link_abs]
        _ir_message = (
            f'The issue "{issue.title}" on project {project.customer_name} '
            f'has been marked as resolved by {resolver_name}.'
        )
        _ir_email_message = (
            f'{_ir_message}\n\nView in Horizon Solar PMS:\n'
            f'https://horizon-solar-pms-production.up.railway.app{issue_link}'
        )
        notified_pks = {profile.pk}
        for notify_recipient in project_managers(project) + [issue.assigned_to, issue.raised_by]:
            if notify_recipient and notify_recipient.pk not in notified_pks:
                notified_pks.add(notify_recipient.pk)
                send_notification(
                    recipient=notify_recipient,
                    message=_ir_email_message,
                    channels=['in_app', 'whatsapp', 'email'],
                    link=issue_link,
                    subject=f'Issue Resolved — {project.customer_name}',
                    template='issue_resolved',
                    template_params=resolved_params,
                    related_project=project,
                    actor=profile,
                )
        messages.success(request, 'Issue marked as Resolved. PM has been notified.')
    return redirect('issue_detail', issue_id=issue_id)


@login_required
def close_issue(request, issue_id):
    """
    Close a Resolved issue. Only the project PM can close.
    filter().update() used to prevent race condition on concurrent status changes.
    Access: project PM only. POST only.
    """
    if request.method != 'POST':
        return redirect('issue_detail', issue_id=issue_id)

    issue   = get_object_or_404(Issue, pk=issue_id)
    project = issue.project
    profile = request.user.profile

    if not _is_project_pm(profile, project):
        return HttpResponseForbidden('Only the project PM can close issues.')

    if issue.status != Issue.RESOLVED:
        messages.warning(request, 'Issue must be Resolved before it can be closed.')
        return redirect('issue_detail', issue_id=issue_id)

    updated = Issue.objects.filter(pk=issue.pk, status=Issue.RESOLVED).update(
        status=Issue.CLOSED,
        closed_at=timezone.now(),
    )
    if updated == 0:
        messages.warning(request, 'Issue status was already updated.')
    else:
        log_activity(project, profile, f"PM closed issue: {issue.title}", entity_type='Issue', entity_id=issue.pk, action_code='issue_closed')
        messages.success(request, 'Issue closed.')
    return redirect('issue_detail', issue_id=issue_id)


@login_required
def reopen_issue(request, issue_id):
    """
    Reopen a Resolved issue back to Open (clears resolved_at and resolution_note).
    Only the project PM can reopen. filter().update() prevents race conditions.
    Access: project PM only. POST only.
    """
    if request.method != 'POST':
        return redirect('issue_detail', issue_id=issue_id)

    issue   = get_object_or_404(Issue, pk=issue_id)
    project = issue.project
    profile = request.user.profile

    if not _is_project_pm(profile, project):
        return HttpResponseForbidden('Only the project PM can reopen issues.')

    if issue.status != Issue.RESOLVED:
        messages.warning(request, 'Only Resolved issues can be reopened.')
        return redirect('issue_detail', issue_id=issue_id)

    updated = Issue.objects.filter(pk=issue.pk, status=Issue.RESOLVED).update(
        status=Issue.OPEN,
        resolved_at=None,
        resolution_note='',
    )
    if updated == 0:
        messages.warning(request, 'Issue status was already updated.')
    else:
        log_activity(project, profile, f"PM reopened issue: {issue.title}", entity_type='Issue', entity_id=issue.pk, action_code='issue_reopened')
        messages.success(request, 'Issue reopened.')
    return redirect('issue_detail', issue_id=issue_id)


@login_required
def assign_issue(request, issue_id):
    """
    Assign or unassign an issue. Closed issues cannot be reassigned.
    filter().update() used for atomic update without loading the full object.
    Access: all roles with project access. POST only.
    """
    if request.method != 'POST':
        return redirect('issue_detail', issue_id=issue_id)

    issue   = get_object_or_404(Issue, pk=issue_id)
    project = issue.project
    profile = request.user.profile

    if profile.role == 'PM' and not user_can_manage_project(request.user, project):
        raise Http404

    if issue.status == Issue.CLOSED:
        return HttpResponseForbidden('Cannot reassign a closed issue.')

    assignee_id = request.POST.get('assigned_to', '').strip()
    if assignee_id:
        try:
            new_assignee = UserProfile.objects.get(pk=assignee_id)
        except UserProfile.DoesNotExist:
            messages.error(request, 'Invalid user selected.')
            return redirect('issue_detail', issue_id=issue_id)
        Issue.objects.filter(pk=issue.pk).update(assigned_to=new_assignee)
        log_activity(
            project, profile,
            f"Assigned issue to {new_assignee.user.get_full_name() or new_assignee.user.username}",
            entity_type='Issue', entity_id=issue.pk,
        )
        messages.success(request, 'Issue assigned.')
    else:
        Issue.objects.filter(pk=issue.pk).update(assigned_to=None)
        log_activity(project, profile, f"Unassigned issue: {issue.title}", entity_type='Issue', entity_id=issue.pk)
        messages.success(request, 'Issue unassigned.')
    return redirect('issue_detail', issue_id=issue_id)


# ---------------------------------------------------------------------------
# Comments
# ---------------------------------------------------------------------------

@login_required
def create_task_comment(request, project_id, task_id):
    """Create a top-level comment or reply on a task. All roles. POST only."""
    if request.method != 'POST':
        return redirect('task_detail', project_id=project_id, task_id=task_id)

    project = get_object_or_404(Project, project_id=project_id)
    task    = get_object_or_404(Task, pk=task_id, phase__project=project)
    profile = request.user.profile

    # PM isolation: PM can only access their own projects
    if profile.role == 'PM' and not user_can_manage_project(request.user, project):
        raise Http404

    body = request.POST.get('body', '').strip()
    if not body:
        messages.error(request, 'Comment body is required.')
        if _is_hx(request):
            return _render_comments_hx(request, project, task)
        return redirect('task_detail', project_id=project_id, task_id=task_id)

    parent_pk      = request.POST.get('parent', '').strip()
    parent_comment = None

    if parent_pk:
        parent_comment = get_object_or_404(Comment, pk=parent_pk)
        # Reject replies to replies — only one level of nesting supported
        if parent_comment.parent is not None:
            if _is_hx(request):
                messages.error(request, 'Replies to replies are not supported.')
                return _render_comments_hx(request, project, task)
            return HttpResponse("Replies to replies are not supported.", status=400)
        # Parent must belong to this same task, not another
        if parent_comment.task != task:
            if _is_hx(request):
                messages.error(request, 'Parent comment mismatch.')
                return _render_comments_hx(request, project, task)
            return HttpResponse("Parent comment mismatch.", status=400)

    Comment.objects.create(
        project=project,
        task=task,
        issue=None,
        parent=parent_comment,
        author=profile,
        body=body,
    )

    # log_activity() is wrapped in try/except internally —
    # a failed log must never block the primary action
    if parent_comment:
        log_activity(project, profile, f"Replied to comment on task: {task.task_name}", entity_type='Comment', entity_id=task.pk)
    else:
        log_activity(project, profile, f"Commented on task: {task.task_name}", entity_type='Comment', entity_id=task.pk)

    if _is_hx(request):
        return _render_comments_hx(request, project, task)

    return redirect('task_detail', project_id=project_id, task_id=task_id)


@login_required
def create_issue_comment(request, issue_id):
    """Create a top-level comment or reply on an issue. All roles. POST only."""
    if request.method != 'POST':
        return redirect('issue_detail', issue_id=issue_id)

    issue   = get_object_or_404(Issue, pk=issue_id)
    project = issue.project
    profile = request.user.profile

    # PM isolation: PM can only access their own projects
    if profile.role == 'PM' and not user_can_manage_project(request.user, project):
        raise Http404

    body = request.POST.get('body', '').strip()
    if not body:
        messages.error(request, 'Comment body is required.')
        return redirect('issue_detail', issue_id=issue_id)

    parent_pk      = request.POST.get('parent', '').strip()
    parent_comment = None

    if parent_pk:
        parent_comment = get_object_or_404(Comment, pk=parent_pk)
        # Reject replies to replies — only one level of nesting supported
        if parent_comment.parent is not None:
            return HttpResponse("Replies to replies are not supported.", status=400)
        # Parent must belong to this same issue, not another
        if parent_comment.issue != issue:
            return HttpResponse("Parent comment mismatch.", status=400)

    Comment.objects.create(
        project=project,
        task=None,
        issue=issue,
        parent=parent_comment,
        author=profile,
        body=body,
    )

    # log_activity() is wrapped in try/except internally —
    # a failed log must never block the primary action
    if parent_comment:
        log_activity(project, profile, f"Replied to comment on issue: {issue.title}", entity_type='Comment', entity_id=issue.pk)
    else:
        log_activity(project, profile, f"Commented on issue: {issue.title}", entity_type='Comment', entity_id=issue.pk)

    return redirect('issue_detail', issue_id=issue_id)


@login_required
def delete_comment(request, comment_id):
    """Soft-delete a comment. Author or Admin only. POST only."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    comment = get_object_or_404(Comment, pk=comment_id)
    profile = request.user.profile

    # Only author or Admin can delete — other roles get 403
    if comment.author != profile and profile.role != 'Admin':
        return HttpResponse(status=403)

    comment.is_deleted = True
    comment.deleted_at = timezone.now()
    comment.save(update_fields=['is_deleted', 'deleted_at'])

    # Log then redirect to the originating detail page
    if comment.task:
        log_activity(
            comment.project, profile,
            f"Deleted comment on task: {comment.task.task_name}",
            entity_type='Comment', entity_id=comment.task.pk,
        )
        return redirect('task_detail', project_id=comment.project.project_id, task_id=comment.task.pk)
    else:
        log_activity(
            comment.project, profile,
            f"Deleted comment on issue: {comment.issue.title}",
            entity_type='Comment', entity_id=comment.issue.pk,
        )
        return redirect('issue_detail', issue_id=comment.issue.pk)


# ---------------------------------------------------------------------------
# Activity log views
# ---------------------------------------------------------------------------

@login_required
def project_timeline(request, project_id):
    """Project activity timeline. All roles — PM isolation applies."""
    from django.core.paginator import Paginator

    project = get_object_or_404(Project, project_id=project_id)
    profile = request.user.profile

    # PM isolation: PM can only view their own projects
    if profile.role == 'PM' and not user_can_manage_project(request.user, project):
        raise Http404

    # All activity logs for this project, newest first, with actor data
    logs = (
        project.activity_logs
        .select_related('actor__user')
        .order_by('-timestamp')
    )

    paginator   = Paginator(logs, 20)
    page_number = request.GET.get('page')
    page_obj    = paginator.get_page(page_number)

    return render(request, 'projects/project_timeline.html', {
        'project':  project,
        'page_obj': page_obj,
    })


@login_required
@role_required(['Admin'])
def portal_activity_log(request):
    """Cross-project activity log with filters. Admin only."""
    from django.core.paginator import Paginator

    # All logs across all projects, newest first
    logs = ActivityLog.objects.select_related(
        'project', 'actor__user'
    ).order_by('-timestamp')

    # Conditional filter values — all optional, empty = show all
    project_id  = request.GET.get('project_id',  '').strip()
    actor_id    = request.GET.get('actor_id',    '').strip()
    entity_type = request.GET.get('entity_type', '').strip()
    date_from   = request.GET.get('date_from',   '').strip()
    date_to     = request.GET.get('date_to',     '').strip()

    if project_id:
        logs = logs.filter(project__project_id=project_id)
    if actor_id:
        logs = logs.filter(actor__pk=actor_id)
    if entity_type:
        logs = logs.filter(entity_type=entity_type)
    if date_from:
        try:
            logs = logs.filter(timestamp__date__gte=date.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            logs = logs.filter(timestamp__date__lte=date.fromisoformat(date_to))
        except ValueError:
            pass

    paginator   = Paginator(logs, 50)
    page_number = request.GET.get('page')
    page_obj    = paginator.get_page(page_number)

    # Actor dropdown: only users who have actually logged actions
    actors = (
        UserProfile.objects
        .filter(activity_logs__isnull=False)
        .distinct()
        .select_related('user')
        .order_by('user__first_name')
    )
    all_projects = Project.objects.order_by('project_id')
    # Hardcoded entity type list matches what views log
    entity_types = ['Issue', 'Comment', 'Task', 'BOQ', 'Milestone', 'File']

    return render(request, 'projects/portal_activity_log.html', {
        'page_obj':     page_obj,
        'all_projects': all_projects,
        'actors':       actors,
        'entity_types': entity_types,
        'filters': {
            'project_id':  project_id,
            'actor_id':    actor_id,
            'entity_type': entity_type,
            'date_from':   date_from,
            'date_to':     date_to,
        },
    })


@login_required
@role_required(['Admin'])
def admin_whatsapp_log(request):
    """WhatsApp diagnostic log — API send status + Interakt delivery status. Admin only."""
    from django.core.paginator import Paginator

    status_param     = request.GET.get('status',     '').strip()
    project_id_param = request.GET.get('project_id', '').strip()
    date_from_param  = request.GET.get('date_from',  '').strip()
    date_to_param    = request.GET.get('date_to',    '').strip()

    default_date_from = (timezone.now().date() - timedelta(days=7)).isoformat()
    date_from = date_from_param or default_date_from

    qs = (
        NotificationLog.objects
        .filter(channel='whatsapp')
        .select_related('recipient__user', 'related_project', 'actor__user')
        .order_by('-created_at')
    )

    if status_param:
        qs = qs.filter(status=status_param)
    if project_id_param:
        qs = qs.filter(related_project_id=project_id_param)
    if date_from:
        try:
            qs = qs.filter(created_at__date__gte=date.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to_param:
        try:
            qs = qs.filter(created_at__date__lte=date.fromisoformat(date_to_param))
        except ValueError:
            pass

    sent_count            = qs.filter(status='sent').count()
    failed_count          = qs.filter(status='failed').count()
    skipped_count         = qs.filter(status='skipped').count()
    delivered_count       = qs.filter(delivery_status='message_api_delivered').count()
    read_count            = qs.filter(delivery_status='message_api_read').count()
    delivery_failed_count = qs.filter(delivery_status='message_api_failed').count()
    pending_count         = qs.filter(status='sent', delivery_status='').count()

    project_list = Project.objects.filter(is_deleted=False).order_by('customer_name')

    paginator   = Paginator(qs, 50)
    page_number = request.GET.get('page')
    page_obj    = paginator.get_page(page_number)

    return render(request, 'projects/admin_whatsapp_log.html', {
        'page_obj':             page_obj,
        'sent_count':           sent_count,
        'failed_count':         failed_count,
        'skipped_count':        skipped_count,
        'delivered_count':      delivered_count,
        'read_count':           read_count,
        'delivery_failed_count': delivery_failed_count,
        'pending_count':        pending_count,
        'project_list':         project_list,
        'current_filters': {
            'status':     status_param,
            'project_id': project_id_param,
            'date_from':  date_from,
            'date_to':    date_to_param,
        },
    })


# ---------------------------------------------------------------------------
# Delivery Challans (SCM Delivery Tracker — Day 9)
# ---------------------------------------------------------------------------

@login_required
@role_required(['SCM'])
def create_delivery_challan(request, project_id):
    """
    Create a new Delivery Challan for a project. Access: SCM only.
    Dynamic line items are submitted with indexed field names (line_item_*_N).
    Minimum 1 line item required — zero items rejected with a validation error.
    GET renders the form; POST creates the DC + line items in one transaction.
    """
    project = get_object_or_404(Project, project_id=project_id)
    profile = request.user.profile
    vendors = Vendor.objects.filter(is_active=True).order_by('name')

    if request.method != 'POST':
        return render(request, 'projects/delivery_challan_create.html', {
            'project':          project,
            'vendors':          vendors,
            'category_choices': DCLineItem.CATEGORY_CHOICES,
        })

    # Parse DC header fields
    vendor_id              = request.POST.get('vendor_id', '').strip()
    po_number              = request.POST.get('po_number', '').strip()
    dc_number              = request.POST.get('dc_number', '').strip()
    dc_date_s              = request.POST.get('dc_date', '').strip()
    expected_delivery_s    = request.POST.get('expected_delivery_date', '').strip()
    notes                  = request.POST.get('notes', '').strip()

    if not dc_number or not dc_date_s:
        messages.error(request, 'DC Number and DC Date are required.')
        return render(request, 'projects/delivery_challan_create.html', {
            'project': project, 'vendors': vendors,
            'category_choices': DCLineItem.CATEGORY_CHOICES,
        })

    try:
        dc_date = date.fromisoformat(dc_date_s)
    except ValueError:
        messages.error(request, 'Invalid DC Date format.')
        return render(request, 'projects/delivery_challan_create.html', {
            'project': project, 'vendors': vendors,
            'category_choices': DCLineItem.CATEGORY_CHOICES,
        })

    expected_delivery_date = None
    if expected_delivery_s:
        try:
            expected_delivery_date = date.fromisoformat(expected_delivery_s)
        except ValueError:
            pass

    vendor = None
    if vendor_id:
        try:
            vendor = Vendor.objects.get(pk=vendor_id, is_active=True)
        except Vendor.DoesNotExist:
            pass

    # Parse dynamically indexed line item fields from POST
    # Field names follow the pattern: line_item_category_N, line_item_description_N, etc.
    line_items_data = []
    i = 0
    while f'line_item_category_{i}' in request.POST:
        category    = request.POST.get(f'line_item_category_{i}', '').strip()
        description = request.POST.get(f'line_item_description_{i}', '').strip()
        qty_str     = request.POST.get(f'line_item_qty_{i}', '').strip()
        unit        = request.POST.get(f'line_item_unit_{i}', 'Nos').strip() or 'Nos'
        qty         = _safe_decimal(qty_str)
        if category and description and qty:
            line_items_data.append({
                'boq_category':     category,
                'item_description': description,
                'ordered_quantity': qty,
                'unit':             unit,
            })
        i += 1

    # Minimum 1 line item — DC with zero items rejected
    if not line_items_data:
        messages.error(request, 'At least one line item is required.')
        return render(request, 'projects/delivery_challan_create.html', {
            'project': project, 'vendors': vendors,
            'category_choices': DCLineItem.CATEGORY_CHOICES,
        })

    with transaction.atomic():
        challan = DeliveryChallan.objects.create(
            project=project,
            vendor=vendor,
            po_number=po_number,
            dc_number=dc_number,
            dc_date=dc_date,
            expected_delivery_date=expected_delivery_date,
            status=DeliveryChallan.EXPECTED,
            notes=notes,
            created_by=profile,
        )
        # Create line items after challan exists; recalculate_dc_status is NOT called
        # here because all new items have no received_quantity → status stays Expected
        for item_data in line_items_data:
            DCLineItem.objects.create(challan=challan, **item_data)

    vendor_name = vendor.name if vendor else 'Unknown Vendor'
    log_activity(
        project, profile,
        f"SCM created Delivery Challan {dc_number} for {vendor_name}",
        entity_type='DeliveryChallan', entity_id=challan.pk,
    )
    messages.success(request, f'Delivery Challan {dc_number} created successfully.')
    return redirect('delivery_challan_detail', project_id=project_id, dc_id=challan.pk)


@login_required
def delivery_challan_detail(request, project_id, dc_id):
    """
    DC detail page: line items table, GRN confirmation form (SE), Edit GRN button (SCM).
    Cross-project guard: DC must belong to the project in the URL — returns 404 on mismatch
    to prevent cross-project data leakage via URL manipulation.
    Access: SCM, PM, SE, Admin.
    """
    project = get_object_or_404(Project, project_id=project_id)
    profile = request.user.profile

    # Only SCM, PM, SE, Admin can view DC pages
    if profile.role not in ('SCM', 'PM', 'Project Coordinator', 'Site Engineer', 'Admin'):
        return HttpResponseForbidden()

    # PM isolation: PM sees only their own projects
    if profile.role == 'PM' and not user_can_manage_project(request.user, project):
        raise Http404

    # Cross-project guard: DC must belong to the project in the URL
    challan    = get_object_or_404(DeliveryChallan, pk=dc_id)
    if challan.project.project_id != project_id:
        raise Http404

    line_items = challan.line_items.select_related('grn_confirmed_by__user').all()

    # Issues raised against this specific DC — shown below line items for visibility
    dc_issues = (
        Issue.objects.filter(delivery_challan=challan)
        .select_related('raised_by__user', 'assigned_to__user')
        .order_by('-raised_at')
    )

    return render(request, 'projects/delivery_challan_detail.html', {
        'project':           project,
        'challan':           challan,
        'line_items':        line_items,
        'dc_issues':         dc_issues,
        'role':              profile.role,
        'user_profile':      profile,
        'condition_choices': DCLineItem.CONDITION_CHOICES,
        'all_profiles':      UserProfile.objects.select_related('user').filter(is_active=True).order_by('user__first_name'),
    })


@login_required
@role_required(['Site Engineer'])
def confirm_grn(request, project_id, dc_id):
    """
    SE confirms receipt of materials at site (GRN confirmation).
    Captures received_quantity and damaged_quantity separately — two numeric inputs
    give precise information rather than a coarse condition dropdown.
    condition field is derived for backward compatibility with code that reads it.
    recalculate_dc_status() called ONCE after all line items saved — never inside the loop.
    Access: Site Engineer only. POST only.
    """
    if request.method != 'POST':
        return redirect('delivery_challan_detail', project_id=project_id, dc_id=dc_id)

    project = get_object_or_404(Project, project_id=project_id)
    profile = request.user.profile

    # Cross-project guard
    challan = get_object_or_404(DeliveryChallan, pk=dc_id)
    if challan.project.project_id != project_id:
        raise Http404

    # SE cannot confirm an already-Received DC — button is hidden in UI but guard here too
    if challan.status == DeliveryChallan.RECEIVED:
        return HttpResponseForbidden()

    today      = date.today()
    line_items = challan.line_items.all()

    for item in line_items:
        qty_str        = request.POST.get(f'received_qty_{item.pk}', '').strip()
        damaged_qty_str = request.POST.get(f'damaged_qty_{item.pk}', '0').strip()
        grn_notes      = request.POST.get(f'grn_notes_{item.pk}', '').strip()

        if not qty_str:
            continue  # Skip items where SE didn't enter a received quantity

        received_qty = _safe_decimal(qty_str)
        if received_qty is None:
            continue

        # Parse damaged_quantity; clamp to [0, received_qty]
        try:
            damaged_qty = max(0, int(float(damaged_qty_str)))
        except (ValueError, TypeError):
            damaged_qty = 0
        damaged_qty = min(damaged_qty, int(received_qty))

        # Derive condition for backward compatibility with code that reads this field
        # condition string is now secondary — damaged_quantity is the precise source of truth
        if damaged_qty == 0:
            derived_condition = DCLineItem.GOOD
        elif damaged_qty >= received_qty:
            derived_condition = DCLineItem.DAMAGED
        else:
            derived_condition = DCLineItem.PARTIAL

        item.received_quantity = received_qty
        item.damaged_quantity  = damaged_qty
        item.condition         = derived_condition
        item.grn_date          = today
        item.grn_confirmed_by  = profile
        item.grn_notes         = grn_notes
        item.save()

    # Recalculate DC status ONCE after all line items saved — never inside the loop
    # (calling it inside the loop causes status to oscillate incorrectly)
    recalculate_dc_status(challan)
    challan.refresh_from_db()

    log_activity(
        project, profile,
        f"SE confirmed GRN for DC {challan.dc_number} — {challan.status}",
        entity_type='DeliveryChallan', entity_id=challan.pk,
    )
    messages.success(request, f'GRN confirmed. DC status: {challan.status}.')
    return redirect('delivery_challan_detail', project_id=project_id, dc_id=dc_id)


@login_required
@role_required(['SCM'])
def override_grn(request, project_id, dc_id):
    """
    SCM overrides an SE-submitted GRN. grn_confirmed_by is NOT overwritten —
    original SE submitter is preserved. ActivityLog records who made the override.
    No status restriction: SCM can override even on Received/Rejected DCs (to correct mistakes).
    Captures received_quantity and damaged_quantity separately — mirrors confirm_grn logic.
    recalculate_dc_status() called ONCE after all items saved.
    Access: SCM only. POST only.
    """
    if request.method != 'POST':
        return redirect('delivery_challan_detail', project_id=project_id, dc_id=dc_id)

    project = get_object_or_404(Project, project_id=project_id)
    profile = request.user.profile

    # Cross-project guard
    challan = get_object_or_404(DeliveryChallan, pk=dc_id)
    if challan.project.project_id != project_id:
        raise Http404

    today      = date.today()
    line_items = challan.line_items.all()

    for item in line_items:
        qty_str        = request.POST.get(f'received_qty_{item.pk}', '').strip()
        damaged_qty_str = request.POST.get(f'damaged_qty_{item.pk}', '0').strip()
        grn_notes      = request.POST.get(f'grn_notes_{item.pk}', '').strip()

        if not qty_str:
            continue  # Skip items where SCM didn't enter a received quantity

        received_qty = _safe_decimal(qty_str)
        if received_qty is None:
            continue

        # Parse damaged_quantity; clamp to [0, received_qty]
        try:
            damaged_qty = max(0, int(float(damaged_qty_str)))
        except (ValueError, TypeError):
            damaged_qty = 0
        damaged_qty = min(damaged_qty, int(received_qty))

        # Derive condition for backward compatibility
        if damaged_qty == 0:
            derived_condition = DCLineItem.GOOD
        elif damaged_qty >= received_qty:
            derived_condition = DCLineItem.DAMAGED
        else:
            derived_condition = DCLineItem.PARTIAL

        item.received_quantity = received_qty
        item.damaged_quantity  = damaged_qty
        item.condition         = derived_condition
        item.grn_date          = today
        item.grn_notes         = grn_notes
        # grn_confirmed_by NOT overwritten — original SE submitter is preserved
        item.save()

    # Recalculate DC status ONCE after all line items saved
    recalculate_dc_status(challan)
    challan.refresh_from_db()

    log_activity(
        project, profile,
        f"SCM overrode GRN for DC {challan.dc_number} — override by {profile.user.get_full_name() or profile.user.username}",
        entity_type='DeliveryChallan', entity_id=challan.pk,
    )
    messages.success(request, f'GRN overridden. DC status: {challan.status}.')
    return redirect('delivery_challan_detail', project_id=project_id, dc_id=dc_id)


# ---------------------------------------------------------------------------
# Interakt Delivery Webhook — publicly reachable, no auth required
# ---------------------------------------------------------------------------

INTERAKT_DELIVERY_PRIORITY = {
    'message_api_read':      4,
    'message_api_delivered': 3,
    'message_api_sent':      2,
    'message_api_failed':    1,
    '':                      0,
}

INTERAKT_VALID_DELIVERY_TYPES = frozenset(INTERAKT_DELIVERY_PRIORITY) - {''}


@csrf_exempt
def interakt_webhook(request):
    import hmac as _hmac
    import hashlib
    import os

    if request.method != 'POST':
        return HttpResponse(status=405)

    secret = os.environ.get('INTERAKT_WEBHOOK_SECRET', '')
    if secret:
        signature = request.headers.get('Interakt-Signature', '')
        expected = 'sha256=' + _hmac.new(
            secret.encode('utf-8'),
            request.body,
            hashlib.sha256,
        ).hexdigest()
        if not _hmac.compare_digest(signature, expected):
            logger.warning(
                'Interakt webhook signature mismatch. Got: %s', signature
            )
            return HttpResponse(status=200)

    try:
        payload = json.loads(request.body)
    except Exception:
        return HttpResponse(status=200)

    event_type = payload.get('type', '')
    message_id = payload.get('data', {}).get('message', {}).get('id', '')

    if event_type not in INTERAKT_VALID_DELIVERY_TYPES or not message_id:
        return HttpResponse(status=200)

    try:
        log = NotificationLog.objects.filter(interakt_message_id=message_id).first()
        if log:
            current_priority = INTERAKT_DELIVERY_PRIORITY.get(log.delivery_status, 0)
            new_priority      = INTERAKT_DELIVERY_PRIORITY.get(event_type, 0)
            if new_priority > current_priority:
                log.delivery_status = event_type
                log.save(update_fields=['delivery_status'])
    except Exception as e:
        logger.error('Interakt webhook DB update failed: %s', e)

    return HttpResponse(status=200)


# ---------------------------------------------------------------------------
# My Documents — personal record-keeping page
# ---------------------------------------------------------------------------

@login_required
def my_documents(request):
    """Personal document archive for the logged-in user. Role-aware sections."""
    profile = request.user.profile
    role    = profile.role

    # Section A — Uploaded Files (all roles)
    task_attachments = TaskAttachment.objects.filter(
        uploaded_by=profile,
        is_deleted=False,
    ).select_related('task', 'task__phase', 'task__phase__project').order_by('-uploaded_at')[:50]

    project_docs = ProjectDocument.objects.filter(
        uploaded_by=profile,
        is_deleted=False,
    ).select_related('project').order_by('-uploaded_at')[:50]

    # Section B — BOQ Submissions (Design only)
    boq_list = []
    if role == 'Design':
        boq_list = BOQ.objects.filter(
            submitted_by=profile,
        ).select_related('project').order_by('-submitted_at')[:50]

    # Section C — Design Submissions (Design only)
    design_list = []
    if role == 'Design':
        design_list = DesignSubmission.objects.filter(
            submitted_by=profile,
        ).select_related('project').order_by('-submitted_at')[:50]

    # Section D — Delivery Challans (SCM only)
    dc_list = []
    if role == 'SCM':
        dc_list = DeliveryChallan.objects.filter(
            created_by=profile,
        ).select_related('project').order_by('-created_at')[:50]

    # Section E — Payment Requests (SCM only)
    # requested_by is FK to auth.User (not UserProfile)
    pr_list = []
    if role == 'SCM':
        pr_list = PaymentRequest.objects.filter(
            requested_by=request.user,
        ).select_related('project', 'vendor').order_by('-requested_date')[:50]

    context = {
        'task_attachments': task_attachments,
        'project_docs':     project_docs,
        'boq_list':         boq_list,
        'design_list':      design_list,
        'dc_list':          dc_list,
        'pr_list':          pr_list,
        'role':             role,
    }
    return render(request, 'projects/my_documents.html', context)


@login_required
def design_submission_detail(request, pk):
    """Read-only detail view for a DesignSubmission. Submitter, PM, or Admin only."""
    submission = get_object_or_404(DesignSubmission, pk=pk)
    profile    = request.user.profile
    if profile != submission.submitted_by and profile.role not in ('PM', 'Admin'):
        messages.error(request, "You don't have access to this submission.")
        return redirect('my_documents')
    return render(request, 'projects/design_submission_detail.html', {'submission': submission})


@login_required
def payment_request_detail(request, project_id, request_id):
    """Read-only detail view for a PaymentRequest. SCM, Finance, PM, or Admin only."""
    project = get_object_or_404(Project, project_id=project_id)
    pr      = get_object_or_404(PaymentRequest, pk=request_id, project=project)
    profile = request.user.profile
    if profile.role not in ('SCM', 'Finance', 'PM', 'Admin'):
        messages.error(request, "You don't have access to this payment request.")
        return redirect('my_documents')
    return render(request, 'projects/payment_request_detail.html', {
        'pr':      pr,
        'project': project,
    })


# ---------------------------------------------------------------------------
# Admin Panel (portal-admin/) — Admin role only
# ---------------------------------------------------------------------------

@login_required
@role_required(['Admin'])
def admin_master_switches(request):
    """Screen 1: Master switches for WhatsApp, email, in-app notifications, maintenance mode, cascade scheduling."""
    settings = SystemSettings.get()

    FIELD_LABELS = {
        'whatsapp_enabled':             'WhatsApp notifications (Interakt)',
        'email_enabled':                'Email notifications (ZeptoMail)',
        'in_app_notifications_enabled': 'In-app notifications',
        'maintenance_mode':             'Maintenance mode',
    }

    if request.method == 'POST':
        actor = request.user.profile
        for field, label in FIELD_LABELS.items():
            old_value = getattr(settings, field)
            new_value = field in request.POST
            if old_value != new_value:
                setattr(settings, field, new_value)
                log_activity(
                    project=None,
                    actor=actor,
                    action=f"Master switch '{label}' set to {'ON' if new_value else 'OFF'}",
                    entity_type='System',
                    entity_id=None,
                )
        # Cascade scheduling feature gate — handled separately for specific log message
        old_cascade = settings.cascade_scheduling_enabled
        new_cascade = 'cascade_scheduling_enabled' in request.POST
        if old_cascade != new_cascade:
            settings.cascade_scheduling_enabled = new_cascade
            log_activity(
                project=None,
                actor=actor,
                action=f"Cascading scheduling feature gate set to {'ON' if new_cascade else 'OFF'}",
                entity_type='System',
                entity_id=None,
            )
        # Gantt integer settings — parse defensively; keep current value on bad/negative input.
        GANTT_INT_FIELDS = {
            'gantt_client_buffer_days':       'Client Gantt buffer (days)',
            'gantt_external_min_display_days': 'Gantt external min bar width (days)',
        }
        for field, label in GANTT_INT_FIELDS.items():
            raw = request.POST.get(field, '').strip()
            try:
                new_int = int(raw)
            except (TypeError, ValueError):
                continue
            if new_int < 0:
                continue
            old_int = getattr(settings, field)
            if old_int != new_int:
                setattr(settings, field, new_int)
                log_activity(
                    project=None,
                    actor=actor,
                    action=f"{label} set to {new_int}",
                    entity_type='System',
                    entity_id=None,
                )
        settings.save()
        messages.success(request, 'Settings saved.')
        return redirect('admin_master_switches')

    return render(request, 'projects/admin/master_switches.html', {'settings': settings})


@login_required
@role_required(['Admin'])
def admin_user_management(request):
    """Screen 2: Activate/deactivate users and change roles."""
    from django.contrib.auth.models import User as AuthUser

    role_choices = UserProfile.ROLE_CHOICES

    if request.method == 'POST':
        action      = request.POST.get('action', '')
        target_id   = request.POST.get('user_id', '')
        actor       = request.user.profile

        try:
            target_user = AuthUser.objects.select_related('profile').get(pk=target_id)
        except AuthUser.DoesNotExist:
            messages.error(request, 'User not found.')
            return redirect('admin_user_management')

        # Never allow Admin to deactivate themselves
        if target_user == request.user and action == 'deactivate':
            messages.error(request, 'You cannot deactivate your own account.')
            return redirect('admin_user_management')

        if action == 'deactivate':
            target_user.is_active = False
            target_user.save()
            log_activity(
                project=None,
                actor=actor,
                action=f"User '{target_user.get_full_name() or target_user.username}' ({target_user.profile.role}) deactivated",
                entity_type='User',
                entity_id=target_user.id,
            )
            messages.success(request, f'{target_user.get_full_name() or target_user.username} deactivated.')

        elif action == 'reactivate':
            target_user.is_active = True
            target_user.save()
            log_activity(
                project=None,
                actor=actor,
                action=f"User '{target_user.get_full_name() or target_user.username}' ({target_user.profile.role}) reactivated",
                entity_type='User',
                entity_id=target_user.id,
            )
            messages.success(request, f'{target_user.get_full_name() or target_user.username} reactivated.')

        elif action == 'change_role':
            new_role = request.POST.get('new_role', '').strip()
            valid_roles = [r[0] for r in role_choices]
            if new_role not in valid_roles:
                messages.error(request, 'Invalid role selected.')
                return redirect('admin_user_management')
            old_role = target_user.profile.role
            if old_role != new_role:
                target_user.profile.role = new_role
                target_user.profile.save()
                log_activity(
                    project=None,
                    actor=actor,
                    action=f"User '{target_user.get_full_name() or target_user.username}' role changed from '{old_role}' to '{new_role}'",
                    entity_type='User',
                    entity_id=target_user.id,
                )
                messages.success(request, f'Role updated for {target_user.get_full_name() or target_user.username}.')
            else:
                messages.success(request, 'No change — role is already set to that value.')

        return redirect('admin_user_management')

    users = (
        AuthUser.objects
        .select_related('profile')
        .order_by('profile__role', 'first_name', 'last_name')
    )
    return render(request, 'projects/admin/user_management.html', {
        'users':        users,
        'role_choices': role_choices,
        'current_user': request.user,
    })


@login_required
@role_required(['Admin'])
def admin_notification_prefs(request):
    """Screen 3: Per-user WhatsApp and email notification toggles."""
    if request.method == 'POST':
        target_profile_id = request.POST.get('profile_id', '')
        actor = request.user.profile
        try:
            target_profile = UserProfile.objects.select_related('user').get(pk=target_profile_id)
        except UserProfile.DoesNotExist:
            messages.error(request, 'User not found.')
            return redirect('admin_notification_prefs')

        new_wa    = 'whatsapp_notifications' in request.POST
        new_email = 'email_notifications' in request.POST

        for field, label, new_val in [
            ('whatsapp_notifications', 'WhatsApp', new_wa),
            ('email_notifications',    'Email',    new_email),
        ]:
            old_val = getattr(target_profile, field)
            if old_val != new_val:
                setattr(target_profile, field, new_val)
                log_activity(
                    project=None,
                    actor=actor,
                    action=(
                        f"Notification pref updated for "
                        f"'{target_profile.user.get_full_name() or target_profile.user.username}': "
                        f"{label} set to {'ON' if new_val else 'OFF'}"
                    ),
                    entity_type='Notification',
                    entity_id=target_profile.id,
                )
        target_profile.save()
        messages.success(request, f'Preferences updated for {target_profile.user.get_full_name() or target_profile.user.username}.')
        return redirect('admin_notification_prefs')

    profiles = UserProfile.objects.select_related('user').order_by('role', 'user__first_name')
    return render(request, 'projects/admin/notification_prefs.html', {'profiles': profiles})


@login_required
@role_required(['Admin'])
def admin_departments(request):
    """Departments view — users grouped by role, with inline deactivate/reactivate/role-change."""
    from django.contrib.auth.models import User as AuthUser
    from itertools import groupby

    DEPT_NAMES = {
        'PM':            'Project Management',
        'Site Engineer': 'Site Execution',
        'SCM':           'Supply Chain',
        'Finance':       'Finance',
        'Design':        'Design',
        'CEO':           'Leadership',
        'BD':            'Sales & Business Development',
        'Admin':         'Administration',
    }

    if request.method == 'POST':
        action    = request.POST.get('action', '')
        target_id = request.POST.get('user_id', '')
        actor     = request.user.profile

        try:
            target_user = AuthUser.objects.select_related('profile').get(pk=target_id)
        except AuthUser.DoesNotExist:
            messages.error(request, 'User not found.')
            return redirect('admin_departments')

        if target_user == request.user and action == 'deactivate':
            messages.error(request, 'You cannot deactivate your own account.')
            return redirect('admin_departments')

        if action == 'deactivate':
            target_user.is_active = False
            target_user.save()
            log_activity(
                project=None,
                actor=actor,
                action=f"User '{target_user.get_full_name() or target_user.username}' ({target_user.profile.role}) deactivated",
                entity_type='User',
                entity_id=target_user.id,
            )
            messages.success(request, f'{target_user.get_full_name() or target_user.username} deactivated.')

        elif action == 'reactivate':
            target_user.is_active = True
            target_user.save()
            log_activity(
                project=None,
                actor=actor,
                action=f"User '{target_user.get_full_name() or target_user.username}' ({target_user.profile.role}) reactivated",
                entity_type='User',
                entity_id=target_user.id,
            )
            messages.success(request, f'{target_user.get_full_name() or target_user.username} reactivated.')

        elif action == 'change_role':
            new_role    = request.POST.get('new_role', '').strip()
            valid_roles = [r[0] for r in UserProfile.ROLE_CHOICES]
            if new_role not in valid_roles:
                messages.error(request, 'Invalid role selected.')
                return redirect('admin_departments')
            old_role = target_user.profile.role
            if old_role != new_role:
                target_user.profile.role = new_role
                target_user.profile.save()
                log_activity(
                    project=None,
                    actor=actor,
                    action=f"User '{target_user.get_full_name() or target_user.username}' role changed from '{old_role}' to '{new_role}'",
                    entity_type='User',
                    entity_id=target_user.id,
                )
                messages.success(request, f'Role updated for {target_user.get_full_name() or target_user.username}.')

        return redirect('admin_departments')

    all_profiles = (
        UserProfile.objects
        .select_related('user')
        .order_by('role', 'user__first_name', 'user__last_name')
    )

    grouped = []
    for role_key, members in groupby(all_profiles, key=lambda p: p.role):
        member_list = list(members)
        active_count = sum(1 for m in member_list if m.user.is_active)
        grouped.append({
            'role':         role_key,
            'dept_name':    DEPT_NAMES.get(role_key, role_key),
            'members':      member_list,
            'active_count': active_count,
        })

    return render(request, 'projects/admin/departments.html', {
        'grouped':      grouped,
        'role_choices': UserProfile.ROLE_CHOICES,
        'current_user': request.user,
    })


@login_required
@role_required(['Admin'])
def admin_user_edit(request, user_id):
    """Edit a user's full profile from the Admin Panel: name, username, email, phone, role, password."""
    from django.contrib.auth.models import User as AuthUser
    target_user = get_object_or_404(AuthUser, pk=user_id)
    try:
        profile = target_user.profile
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=target_user)

    if request.method == 'POST':
        form = AdminUserEditForm(request.POST, instance_user=target_user)
        if form.is_valid():
            cd = form.cleaned_data
            actor = request.user.profile

            target_user.first_name = cd['first_name']
            target_user.last_name  = cd['last_name']
            target_user.username   = cd['username']
            target_user.email      = cd['email']
            target_user.is_staff   = (cd['role'] == 'Admin')
            if cd['new_password']:
                target_user.set_password(cd['new_password'])
            target_user.save()

            profile.role            = cd['role']
            profile.phone_number    = cd['phone_number']
            profile.is_design_head  = cd['is_design_head']
            profile.save()

            log_activity(
                project=None,
                actor=actor,
                action=f"User '{target_user.get_full_name() or target_user.username}' profile updated by admin",
                entity_type='User',
                entity_id=target_user.id,
            )
            if cd['new_password']:
                log_activity(
                    project=None,
                    actor=actor,
                    action=f"Password reset for user '{target_user.username}' by admin",
                    entity_type='User',
                    entity_id=target_user.id,
                )

            messages.success(request, f'{target_user.get_full_name() or target_user.username} updated successfully.')
            return redirect('admin_departments')
    else:
        form = AdminUserEditForm(
            initial={
                'first_name':   target_user.first_name,
                'last_name':    target_user.last_name,
                'username':     target_user.username,
                'email':        target_user.email,
                'phone_number': profile.phone_number,
                'role':         profile.role,
                'is_design_head': profile.is_design_head,
            },
            instance_user=target_user,
        )

    return render(request, 'projects/admin/user_edit.html', {
        'form':        form,
        'target_user': target_user,
    })


@login_required
@role_required(['Admin'])
def admin_send_records(request):
    """Screen 5: Notification send log — all channels with filters and CSV export."""
    import csv
    from django.http import HttpResponse
    from django.core.paginator import Paginator

    channel_param   = request.GET.get('channel',   '').strip()
    status_param    = request.GET.get('status',    '').strip()
    date_from_param = request.GET.get('date_from', '').strip()
    date_to_param   = request.GET.get('date_to',   '').strip()
    export          = request.GET.get('export',    '').strip()

    default_date_from = (timezone.now().date() - timedelta(days=7)).isoformat()
    date_from = date_from_param or default_date_from

    qs = (
        NotificationLog.objects
        .select_related('recipient__user', 'related_project', 'actor__user')
        .order_by('-created_at')
    )

    if channel_param and channel_param != 'all':
        qs = qs.filter(channel=channel_param)
    if status_param and status_param != 'all':
        qs = qs.filter(status=status_param)
    if date_from:
        try:
            qs = qs.filter(created_at__date__gte=date.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to_param:
        try:
            qs = qs.filter(created_at__date__lte=date.fromisoformat(date_to_param))
        except ValueError:
            pass

    # Stat cards — always last 7 days, unfiltered by channel/status
    seven_days_ago = timezone.now() - timedelta(days=7)
    stats_qs = NotificationLog.objects.filter(created_at__gte=seven_days_ago)
    stat_total     = stats_qs.count()
    stat_whatsapp  = stats_qs.filter(channel='whatsapp').count()
    stat_email     = stats_qs.filter(channel='email').count()
    stat_failed    = stats_qs.filter(status='failed').count()

    if export == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="send_records.csv"'
        writer = csv.writer(response)
        writer.writerow(['timestamp', 'recipient', 'role', 'channel', 'template', 'api_status', 'delivery_status'])
        for log in qs.iterator():
            writer.writerow([
                log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                log.recipient.user.get_full_name() or log.recipient.user.username,
                log.recipient.role,
                log.channel,
                log.template_name,
                log.status,
                log.delivery_status,
            ])
        return response

    paginator   = Paginator(qs, 50)
    page_number = request.GET.get('page')
    page_obj    = paginator.get_page(page_number)

    return render(request, 'projects/admin/send_records.html', {
        'page_obj':    page_obj,
        'stat_total':  stat_total,
        'stat_wa':     stat_whatsapp,
        'stat_email':  stat_email,
        'stat_failed': stat_failed,
        'filters': {
            'channel':   channel_param,
            'status':    status_param,
            'date_from': date_from,
            'date_to':   date_to_param,
        },
    })


# ---------------------------------------------------------------------------
# Admin — Audit Log
# ---------------------------------------------------------------------------

def _export_audit_csv(qs):
    import csv
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="audit_log.csv"'
    writer = csv.writer(response)
    writer.writerow(['Timestamp', 'User', 'Role', 'Action', 'Entity Type', 'Entity ID', 'Project'])
    for entry in qs:
        writer.writerow([
            entry.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            entry.actor.user.get_full_name() if entry.actor else '—',
            entry.actor.role if entry.actor else '—',
            entry.action,
            entry.entity_type,
            entry.entity_id or '—',
            entry.project.project_id if entry.project else '—',
        ])
    return response


@login_required
@role_required(['Admin'])
def admin_audit_log(request):
    """Full audit log across all users and projects. Admin only."""
    from django.core.paginator import Paginator

    qs = ActivityLog.objects.select_related('actor__user', 'project').order_by('-timestamp')

    user_id   = request.GET.get('user', '').strip()
    entity    = request.GET.get('entity_type', '').strip()
    project   = request.GET.get('project', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to   = request.GET.get('date_to', '').strip()
    keyword   = request.GET.get('keyword', '').strip()

    if user_id:
        qs = qs.filter(actor__user__id=user_id)
    if entity:
        qs = qs.filter(entity_type=entity)
    if project:
        qs = qs.filter(project__project_id=project)
    if date_from:
        try:
            qs = qs.filter(timestamp__date__gte=date.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            qs = qs.filter(timestamp__date__lte=date.fromisoformat(date_to))
        except ValueError:
            pass
    if keyword:
        qs = qs.filter(action__icontains=keyword)

    if request.GET.get('export') == 'csv':
        return _export_audit_csv(qs)

    seven_days_ago = timezone.now() - timedelta(days=7)
    stats = {
        'today': ActivityLog.objects.filter(timestamp__date=timezone.now().date()).count(),
        'this_week': ActivityLog.objects.filter(timestamp__gte=seven_days_ago).count(),
        'admin_actions': ActivityLog.objects.filter(
            timestamp__gte=seven_days_ago,
            entity_type__in=['System', 'User', 'Notification'],
        ).count(),
        'most_active': (
            ActivityLog.objects
            .filter(timestamp__gte=seven_days_ago)
            .values('actor__user__first_name', 'actor__user__last_name', 'actor__role')
            .annotate(cnt=Count('id'))
            .order_by('-cnt')
            .first()
        ),
    }

    paginator = Paginator(qs, 50)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    all_users    = UserProfile.objects.select_related('user').order_by('user__first_name')
    all_projects = Project.objects.filter(is_deleted=False).order_by('project_id')
    entity_types = ['Task', 'Issue', 'Project', 'Milestone', 'File', 'BOQ', 'Comment',
                    'User', 'System', 'Notification']

    return render(request, 'projects/admin/audit_log.html', {
        'page_obj':     page_obj,
        'stats':        stats,
        'all_users':    all_users,
        'all_projects': all_projects,
        'entity_types': entity_types,
        'filters': {
            'user':        user_id,
            'entity_type': entity,
            'project':     project,
            'date_from':   date_from,
            'date_to':     date_to,
            'keyword':     keyword,
        },
    })


@login_required
@role_required(['Admin'])
def admin_project_list(request):
    """All projects table inside the Admin Panel. Admin only."""
    projects = (
        Project.objects
        .filter(is_deleted=False)
        .select_related('assigned_pm__user')
        .prefetch_related(
            Prefetch('phases',
                     queryset=ProjectPhase.objects.prefetch_related('tasks').order_by('phase_order'))
        )
        .order_by('-created_at')
    )
    pm_users = (
        UserProfile.objects
        .filter(role='PM', is_active=True)
        .select_related('user')
        .order_by('user__first_name')
    )
    return render(request, 'projects/admin/projects_list.html', {
        'projects': projects,
        'pm_users': pm_users,
    })


@login_required
@role_required(['Admin'])
def admin_assign_pm(request, project_id):
    """Assign (or reassign) a PM to a Draft project. POST only. Admin only."""
    if request.method != 'POST':
        return redirect('admin_project_list')

    project = get_object_or_404(Project, project_id=project_id, is_deleted=False)

    pm_user_id = request.POST.get('pm_user_id', '').strip()
    if not pm_user_id:
        messages.error(request, 'Please select a PM.')
        return redirect('admin_project_list')

    try:
        pm_profile = UserProfile.objects.select_related('user').get(pk=pm_user_id, role='PM', is_active=True)
    except UserProfile.DoesNotExist:
        messages.error(request, 'Selected user is not a valid active PM.')
        return redirect('admin_project_list')

    old_pm = project.assigned_pm
    project.assigned_pm = pm_profile
    project.save(update_fields=['assigned_pm'])

    log_activity(
        project=project,
        actor=request.user.profile,  # was request.user (a User) — silently rejected by the FK
        action=(
            f"PM assigned to {pm_profile.user.get_full_name() or pm_profile.user.username}"
            + (f" (previously {old_pm.user.get_full_name() or old_pm.user.username})" if old_pm else "")
        ),
        entity_type='Project',
        entity_id=project.pk,
        action_code='pm_assigned',
    )

    # Notify the newly assigned PM
    try:
        pm_display_name = pm_profile.user.get_full_name() or pm_profile.user.username
        _link = f'/projects/{project.project_id}/overview/'
        _abs_link = request.build_absolute_uri(_link)
        _body = (
            f'Hi {pm_display_name},\n\n'
            f'You have been assigned as Project Manager for {project.customer_name}'
            + (f' ({project.city})' if project.city else '') + '.\n\n'
            f'Please review the project details and activate when ready.'
            f'\n\nView in Horizon Solar PMS:\nhttps://horizon-solar-pms-production.up.railway.app{_link}'
        )
        send_notification(
            recipient=pm_profile,
            message=_body,
            channels=['in_app', 'whatsapp', 'email'],
            link=_link,
            subject=f'New Project Assigned: {project.customer_name}',
            template='assign_project',
            template_params=[
                project.customer_name,
                pm_display_name,
                _abs_link,
            ],
            related_project=project,
        )
    except Exception as exc:
        logger.error('admin_assign_pm: notification failed for project %s — %s', project.project_id, exc)

    messages.success(
        request,
        f'PM assigned to {project.project_id}: {pm_profile.user.get_full_name() or pm_profile.user.username}.'
    )
    return redirect('admin_project_list')


@login_required
@role_required(['Admin'])
def admin_task_durations(request):
    """
    Admin view to edit default duration_days for each task in the residential project template.
    Changes apply to new projects only — existing project tasks are never modified.
    GET: render grouped table. POST: validate and save changed values.
    """
    if request.method == 'POST':
        actor = request.user.profile
        changed = 0
        errors = []

        for key, raw_val in request.POST.items():
            if not key.startswith('duration_'):
                continue
            try:
                pk = int(key.split('_', 1)[1])
            except (ValueError, IndexError):
                continue

            raw_val = raw_val.strip()
            if not raw_val.isdigit():
                errors.append(f"Invalid value '{raw_val}' — must be a non-negative whole number.")
                continue
            new_days = int(raw_val)

            try:
                record = TaskDurationTemplate.objects.get(pk=pk)
            except TaskDurationTemplate.DoesNotExist:
                continue

            if new_days == record.duration_days:
                continue

            old_days = record.duration_days
            record.duration_days = new_days
            record.updated_by = request.user
            record.save(update_fields=['duration_days', 'updated_by', 'updated_at'])
            changed += 1

            log_activity(
                project=None,
                actor=actor,
                action=(
                    f"Updated task duration: '{record.task_name}' ({record.phase_name}) "
                    f"changed from {old_days}d to {new_days}d [residential template]"
                ),
                entity_type='TaskDurationTemplate',
                entity_id=record.pk,
            )

        if errors:
            for msg in errors:
                messages.error(request, msg)
        else:
            messages.success(request, f'Saved. {changed} duration(s) updated.')

        return redirect('admin_task_durations')

    # GET — group records by phase_name preserving the natural phase order
    PHASE_ORDER = [
        'Sales & Documentation',
        'Detail Engineering Visit',
        'Design',
        'Pre-Installation Approvals',
        'Procurement',
        'Delivery',
        'Installation',
        'Commissioning',
        'Finance Closure',
    ]
    records = list(
        TaskDurationTemplate.objects
        .filter(project_type='residential')
        .select_related('updated_by')
        .order_by('phase_name', 'task_name')
    )

    phase_groups = {phase: [] for phase in PHASE_ORDER}
    for rec in records:
        if rec.phase_name in phase_groups:
            phase_groups[rec.phase_name].append(rec)

    # Build ordered list of (phase_name, [records]) skipping empty phases
    grouped = [(phase, phase_groups[phase]) for phase in PHASE_ORDER if phase_groups[phase]]

    return render(request, 'projects/admin/task_durations.html', {
        'grouped': grouped,
        'total': len(records),
    })


# ---------------------------------------------------------------------------
# Portal-admin: reusable Checklists
#
# A Checklist is authored once here (name + ordered items) and surfaced on a task
# by linking it to one or more (task_name, project_type) pairs. UNIQUE on that pair
# — a task can have at most one checklist; a second assignment is rejected here with a
# clear error (not just at the DB level). Task names for the picker are sourced from the
# hardcoded Residential template via utils.get_residential_template_task_names(). All
# mutations are Admin-only and log_activity(entity_type='Checklist'). Item CRUD lives
# here (NOT on task detail); task detail is completion-only.
# ---------------------------------------------------------------------------

def _checklist_task_name_choices():
    """Ordered (phase_name, task_name) pairs the admin may assign — from the Residential
    template. Reported source: projects.utils.get_residential_template_task_names(), which
    reads the single hardcoded PHASES structure shared with attach_residential_template()."""
    return get_residential_template_task_names()


@login_required
@role_required(['Admin'])
def admin_checklists(request):
    """List all Checklists (name, item count, assigned task_name/project_type pairs, active
    toggle). Access: Admin only."""
    checklists = (
        Checklist.objects
        .prefetch_related('items', 'task_links')
        .all()
    )
    return render(request, 'projects/admin/checklists.html', {
        'checklists': checklists,
    })


@login_required
@role_required(['Admin'])
def admin_checklist_create(request):
    """Create a new empty Checklist, then redirect to its editor. Access: Admin only. POST only."""
    if request.method != 'POST':
        return redirect('admin_checklists')

    name = request.POST.get('name', '').strip()
    if not name:
        messages.error(request, 'Please enter a checklist name.')
        return redirect('admin_checklists')

    checklist = Checklist.objects.create(name=name, created_by=request.user)
    log_activity(None, request.user.profile,
                 f"Created checklist '{name}'",
                 entity_type='Checklist', entity_id=checklist.pk)
    messages.success(request, f'Checklist "{name}" created. Add items and assign tasks below.')
    return redirect('admin_checklist_edit', checklist_id=checklist.pk)


@login_required
@role_required(['Admin'])
def admin_checklist_edit(request, checklist_id):
    """Editor for one Checklist: rename/active toggle, item add/edit/delete/reorder, and
    task-link assign/unassign. GET only — mutations POST to the dedicated actions below."""
    checklist = get_object_or_404(Checklist, pk=checklist_id)
    items = checklist.items.all()
    links = checklist.task_links.all()

    return render(request, 'projects/admin/checklist_edit.html', {
        'checklist':       checklist,
        'items':           items,
        'links':           links,
        'task_name_pairs': _checklist_task_name_choices(),
        'project_types':   Project.PROJECT_TYPE_CHOICES,
    })


@login_required
@role_required(['Admin'])
def admin_checklist_update(request, checklist_id):
    """Rename a Checklist and/or toggle its active flag. Access: Admin only. POST only."""
    if request.method != 'POST':
        return redirect('admin_checklist_edit', checklist_id=checklist_id)

    checklist = get_object_or_404(Checklist, pk=checklist_id)
    name = request.POST.get('name', '').strip()
    if not name:
        messages.error(request, 'Checklist name cannot be empty.')
        return redirect('admin_checklist_edit', checklist_id=checklist_id)

    is_active = request.POST.get('is_active') == 'on'
    checklist.name = name
    checklist.is_active = is_active
    checklist.save(update_fields=['name', 'is_active'])

    log_activity(None, request.user.profile,
                 f"Updated checklist '{name}' (active={is_active})",
                 entity_type='Checklist', entity_id=checklist.pk)
    messages.success(request, 'Checklist saved.')
    return redirect('admin_checklist_edit', checklist_id=checklist_id)


@login_required
@role_required(['Admin'])
def admin_checklist_delete(request, checklist_id):
    """Delete a Checklist. Cascades to its items, task links, and item completions.
    Access: Admin only. POST only."""
    if request.method != 'POST':
        return redirect('admin_checklists')

    checklist = get_object_or_404(Checklist, pk=checklist_id)
    name = checklist.name
    checklist.delete()  # CASCADE: items → completions, and task_links

    log_activity(None, request.user.profile,
                 f"Deleted checklist '{name}' (and its items, links, and completions)",
                 entity_type='Checklist', entity_id=None)
    messages.success(request, f'Checklist "{name}" deleted.')
    return redirect('admin_checklists')


@login_required
@role_required(['Admin'])
def admin_checklist_item_add(request, checklist_id):
    """Append one item to a Checklist. Access: Admin only. POST only."""
    if request.method != 'POST':
        return redirect('admin_checklist_edit', checklist_id=checklist_id)

    checklist = get_object_or_404(Checklist, pk=checklist_id)
    label = request.POST.get('label', '').strip()
    if not label:
        messages.error(request, 'Please enter a label for the item.')
        return redirect('admin_checklist_edit', checklist_id=checklist_id)

    next_order = (checklist.items.aggregate(Max('order'))['order__max'] or 0) + 1
    item = ChecklistItem.objects.create(checklist=checklist, label=label, order=next_order)

    log_activity(None, request.user.profile,
                 f"Added item '{label}' to checklist '{checklist.name}'",
                 entity_type='Checklist', entity_id=checklist.pk)
    messages.success(request, 'Item added.')
    return redirect('admin_checklist_edit', checklist_id=checklist_id)


@login_required
@role_required(['Admin'])
def admin_checklist_item_edit(request, checklist_id, item_id):
    """Edit one Checklist item's label. Access: Admin only. POST only."""
    if request.method != 'POST':
        return redirect('admin_checklist_edit', checklist_id=checklist_id)

    checklist = get_object_or_404(Checklist, pk=checklist_id)
    item = get_object_or_404(ChecklistItem, pk=item_id, checklist=checklist)
    label = request.POST.get('label', '').strip()
    if not label:
        messages.error(request, 'Label cannot be empty.')
        return redirect('admin_checklist_edit', checklist_id=checklist_id)

    item.label = label
    item.save(update_fields=['label'])

    log_activity(None, request.user.profile,
                 f"Edited item to '{label}' on checklist '{checklist.name}'",
                 entity_type='Checklist', entity_id=checklist.pk)
    messages.success(request, 'Item updated.')
    return redirect('admin_checklist_edit', checklist_id=checklist_id)


@login_required
@role_required(['Admin'])
def admin_checklist_item_delete(request, checklist_id, item_id):
    """Delete one Checklist item (cascades to its completions). Access: Admin only. POST only."""
    if request.method != 'POST':
        return redirect('admin_checklist_edit', checklist_id=checklist_id)

    checklist = get_object_or_404(Checklist, pk=checklist_id)
    item = get_object_or_404(ChecklistItem, pk=item_id, checklist=checklist)
    label = item.label
    item.delete()

    log_activity(None, request.user.profile,
                 f"Deleted item '{label}' from checklist '{checklist.name}'",
                 entity_type='Checklist', entity_id=checklist.pk)
    messages.success(request, 'Item deleted.')
    return redirect('admin_checklist_edit', checklist_id=checklist_id)


@login_required
@role_required(['Admin'])
def admin_checklist_item_move(request, checklist_id, item_id):
    """Reorder a Checklist item up or down by swapping `order` with its neighbour (no drag
    library). Access: Admin only. POST only."""
    if request.method != 'POST':
        return redirect('admin_checklist_edit', checklist_id=checklist_id)

    checklist = get_object_or_404(Checklist, pk=checklist_id)
    item = get_object_or_404(ChecklistItem, pk=item_id, checklist=checklist)
    direction = request.POST.get('direction', '')

    if direction == 'up':
        neighbor = (checklist.items
                    .filter(order__lt=item.order).order_by('-order', '-pk').first())
    elif direction == 'down':
        neighbor = (checklist.items
                    .filter(order__gt=item.order).order_by('order', 'pk').first())
    else:
        messages.error(request, 'Invalid move direction.')
        return redirect('admin_checklist_edit', checklist_id=checklist_id)

    # No neighbour → already at the top/bottom; silent no-op
    if neighbor is not None:
        with transaction.atomic():
            item.order, neighbor.order = neighbor.order, item.order
            item.save(update_fields=['order'])
            neighbor.save(update_fields=['order'])
        log_activity(None, request.user.profile,
                     f"Reordered item '{item.label}' {direction} on checklist '{checklist.name}'",
                     entity_type='Checklist', entity_id=checklist.pk)

    return redirect('admin_checklist_edit', checklist_id=checklist_id)


@login_required
@role_required(['Admin'])
def admin_checklist_link_add(request, checklist_id):
    """Assign this Checklist to a (task_name, project_type) pair. Enforces the
    one-checklist-per-task-name+type rule at this layer with a clear error before the DB
    unique constraint is ever hit. Access: Admin only. POST only."""
    if request.method != 'POST':
        return redirect('admin_checklist_edit', checklist_id=checklist_id)

    checklist = get_object_or_404(Checklist, pk=checklist_id)
    task_name = request.POST.get('task_name', '').strip()
    project_type = request.POST.get('project_type', '').strip()

    valid_task_names = {name for _phase, name in _checklist_task_name_choices()}
    valid_types = {value for value, _label in Project.PROJECT_TYPE_CHOICES}
    if task_name not in valid_task_names or project_type not in valid_types:
        messages.error(request, 'Please choose a valid task name and project type.')
        return redirect('admin_checklist_edit', checklist_id=checklist_id)

    # Uniqueness enforced here with a clear message — a second checklist on an already-linked
    # (task_name, project_type) is rejected, never silently overwritten.
    existing = (ChecklistTaskLink.objects
                .select_related('checklist')
                .filter(task_name=task_name, project_type=project_type)
                .first())
    if existing is not None:
        if existing.checklist_id == checklist.pk:
            messages.error(request, f'"{task_name}" ({project_type}) is already assigned to this checklist.')
        else:
            messages.error(
                request,
                f'"{task_name}" ({project_type}) is already assigned to checklist '
                f'"{existing.checklist.name}". Unassign it there first.'
            )
        return redirect('admin_checklist_edit', checklist_id=checklist_id)

    link = ChecklistTaskLink.objects.create(
        checklist=checklist, task_name=task_name, project_type=project_type,
    )
    log_activity(None, request.user.profile,
                 f"Assigned checklist '{checklist.name}' to task '{task_name}' ({project_type})",
                 entity_type='Checklist', entity_id=checklist.pk)
    messages.success(request, f'Assigned to "{task_name}" ({project_type}).')
    return redirect('admin_checklist_edit', checklist_id=checklist_id)


@login_required
@role_required(['Admin'])
def admin_checklist_link_delete(request, checklist_id, link_id):
    """Unassign a Checklist from a task (delete one ChecklistTaskLink). Access: Admin only.
    POST only."""
    if request.method != 'POST':
        return redirect('admin_checklist_edit', checklist_id=checklist_id)

    checklist = get_object_or_404(Checklist, pk=checklist_id)
    link = get_object_or_404(ChecklistTaskLink, pk=link_id, checklist=checklist)
    task_name, project_type = link.task_name, link.project_type
    link.delete()

    log_activity(None, request.user.profile,
                 f"Unassigned checklist '{checklist.name}' from task '{task_name}' ({project_type})",
                 entity_type='Checklist', entity_id=checklist.pk)
    messages.success(request, f'Unassigned from "{task_name}" ({project_type}).')
    return redirect('admin_checklist_edit', checklist_id=checklist_id)


# ---------------------------------------------------------------------------
# Admin Panel — BOQ Item Master (catalogue)
#
# The catalogue BOQ line items reference. Editing an entry here never touches an
# existing BOQItem: BOQItem.description is a point-in-time snapshot taken at BOQ
# creation, and nothing below writes to BOQItem.
#
# Deactivate, never delete — BOQItem.item_master is SET_NULL, so deleting a row
# would silently drop rows out of cross-site quantity aggregation. Hence no delete
# view exists.
# ---------------------------------------------------------------------------

@login_required
@role_required(['Admin'])
def admin_boq_items(request):
    """List all BOQItemMaster entries in template order, with an optional active filter
    and the count of BOQ rows linked to each. Access: Admin only."""
    items = BOQItemMaster.objects.annotate(linked_count=Count('boq_items'))

    active_filter = request.GET.get('active', '')
    if active_filter == '1':
        items = items.filter(is_active=True)
    elif active_filter == '0':
        items = items.filter(is_active=False)

    return render(request, 'projects/admin/boq_items.html', {
        'items':         items,
        'active_filter': active_filter,
        'active_count':  BOQItemMaster.objects.filter(is_active=True).count(),
        'total_count':   BOQItemMaster.objects.count(),
    })


@login_required
@role_required(['Admin'])
def admin_boq_item_create(request):
    """Add one catalogue entry. Access: Admin only."""
    if request.method == 'POST':
        form = BOQItemMasterForm(request.POST)
        if form.is_valid():
            item = form.save()
            log_activity(None, request.user.profile,
                         f"Created BOQ catalogue item '{item.code} — {item.description}'",
                         entity_type='BOQItemMaster', entity_id=item.pk)
            messages.success(request, f'Catalogue item "{item.code}" added.')
            return redirect('admin_boq_items')
    else:
        # Suggest the next slot at the end of the template rather than 0
        next_order = (BOQItemMaster.objects.aggregate(m=Max('sort_order'))['m'] or 0) + 1
        form = BOQItemMasterForm(initial={'sort_order': next_order, 'is_active': True})

    return render(request, 'projects/admin/boq_item_form.html', {
        'form':  form,
        'title': 'Add BOQ Catalogue Item',
        'item':  None,
    })


@login_required
@role_required(['Admin'])
def admin_boq_item_edit(request, item_id):
    """Edit one catalogue entry. `code` is create-only and is removed from the form here
    — reassigning it would break the identifier existing BOQ rows were grouped under.
    Existing BOQItem rows are never modified. Access: Admin only."""
    item = get_object_or_404(BOQItemMaster, pk=item_id)

    if request.method == 'POST':
        form = BOQItemMasterForm(request.POST, instance=item)
        del form.fields['code']
        if form.is_valid():
            form.save()
            log_activity(None, request.user.profile,
                         f"Updated BOQ catalogue item '{item.code}' "
                         f"(active={item.is_active})",
                         entity_type='BOQItemMaster', entity_id=item.pk)
            messages.success(request, f'Catalogue item "{item.code}" saved.')
            return redirect('admin_boq_items')
    else:
        form = BOQItemMasterForm(instance=item)
        del form.fields['code']

    return render(request, 'projects/admin/boq_item_form.html', {
        'form':         form,
        'title':        f'Edit BOQ Catalogue Item — {item.code}',
        'item':         item,
        'linked_count': item.boq_items.count(),
    })


@login_required
@role_required(['Admin'])
def admin_boq_item_toggle(request, item_id):
    """Deactivate / reactivate one catalogue entry. Deactivated entries drop out of
    get_standard_boq_items(), so new BOQs stop including them; BOQ rows already created
    from the entry keep their description, quantity and item_master link untouched.
    Access: Admin only. POST only."""
    if request.method != 'POST':
        return redirect('admin_boq_items')

    item = get_object_or_404(BOQItemMaster, pk=item_id)
    item.is_active = not item.is_active
    item.save(update_fields=['is_active', 'updated_at'])

    state = 'activated' if item.is_active else 'deactivated'
    log_activity(None, request.user.profile,
                 f"{state.capitalize()} BOQ catalogue item '{item.code}'",
                 entity_type='BOQItemMaster', entity_id=item.pk)
    messages.success(request, f'Catalogue item "{item.code}" {state}.')
    return redirect('admin_boq_items')


# ---------------------------------------------------------------------------
# Password management
# ---------------------------------------------------------------------------

@login_required
def change_password(request):
    """Allow any authenticated user to change their own password."""
    from django.contrib.auth import update_session_auth_hash
    from django.contrib.auth.forms import PasswordChangeForm

    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Password changed successfully.')
            return redirect('my_documents')
    else:
        form = PasswordChangeForm(request.user)

    return render(request, 'projects/change_password.html', {'form': form})


@login_required
@role_required(['Admin'])
def admin_reset_password(request, user_id):
    """Admin-only: set a new password for any user."""
    from django.contrib.auth.models import User as AuthUser

    target_user = get_object_or_404(AuthUser, pk=user_id)
    actor = request.user.profile

    if request.method == 'POST':
        new_password    = request.POST.get('new_password', '').strip()
        confirm_password = request.POST.get('confirm_password', '').strip()

        if not new_password:
            messages.error(request, 'Password cannot be empty.')
        elif len(new_password) < 8:
            messages.error(request, 'Password must be at least 8 characters.')
        elif new_password != confirm_password:
            messages.error(request, 'Passwords do not match.')
        else:
            target_user.set_password(new_password)
            target_user.save()
            log_activity(
                project=None,
                actor=actor,
                action=f"Admin reset password for '{target_user.get_full_name() or target_user.username}'",
                entity_type='User',
                entity_id=target_user.id,
            )
            messages.success(
                request,
                f'Password reset for {target_user.get_full_name() or target_user.username}.'
            )
            return redirect('admin_user_management')

    return render(request, 'projects/admin/reset_password.html', {
        'target_user': target_user,
    })


# ---------------------------------------------------------------------------
# System Admin Panel
# ---------------------------------------------------------------------------

from projects.decorators import system_admin_required  # noqa: E402
from django.core.exceptions import PermissionDenied  # noqa: E402


# Roles that System Admin must never see, query, or be able to assign
_SA_EXCLUDED_ROLES = ['Admin', 'System Admin']

# Operational roles System Admin may create/edit — never includes Admin or System Admin
_SA_EDITABLE_ROLE_CHOICES = [
    ('PM',                  'PM'),
    ('Project Coordinator', 'Project Coordinator'),
    ('Site Engineer',       'Site Engineer'),
    ('Design',              'Design'),
    ('Finance',             'Finance'),
    ('SCM',                 'SCM'),
    ('CEO',                 'CEO'),
    ('BD',                  'BD'),
]

_SA_DEPT_NAMES = {
    'PM':                  'Project Management',
    'Project Coordinator': 'Project Management',  # Coordinators sit under PM in the org
    'Site Engineer':       'Site Execution',
    'SCM':                 'Supply Chain',
    'Finance':             'Finance',
    'Design':              'Design',
    'CEO':                 'Leadership',
    'BD':                  'Sales & Business Development',
}

# Phase order for task duration template
_PHASE_ORDER = [
    'Sales & Documentation',
    'Detail Engineering Visit',
    'Design',
    'Pre-Installation Approvals',
    'Procurement',
    'Delivery',
    'Installation',
    'Commissioning',
    'Finance Closure',
]


@system_admin_required
def subadmin_projects(request):
    """System Admin: view all projects and assign unassigned ones to a PM (first-time only)."""
    if request.method == 'POST':
        project_pk  = request.POST.get('project_id', '').strip()
        pm_prof_pk  = request.POST.get('pm_user_id', '').strip()

        if not project_pk or not pm_prof_pk:
            messages.error(request, 'Please select a PM before assigning.')
            return redirect('subadmin_projects')

        project = get_object_or_404(Project, pk=project_pk, is_deleted=False)

        # Only first-time assignment — prevent reassignment of already-assigned projects
        if project.assigned_pm_id:
            messages.error(request, 'This project already has a PM assigned. Reassignment is not permitted here.')
            return redirect('subadmin_projects')

        pm_profile = get_object_or_404(UserProfile, pk=pm_prof_pk, role='PM', is_active=True)
        project.assigned_pm = pm_profile
        project.save(update_fields=['assigned_pm'])

        log_activity(
            project=project,
            actor=request.user.profile,
            action=f"PM '{pm_profile.user.get_full_name() or pm_profile.user.username}' assigned to project '{project.project_id}' by System Admin",
            entity_type='Project',
            entity_id=project.pk,
        )
        messages.success(request, f'PM assigned to {project.project_id}.')
        return redirect('subadmin_projects')

    unassigned  = (
        Project.objects
        .filter(assigned_pm__isnull=True, is_deleted=False)
        .order_by('-created_at')
    )
    assigned = (
        Project.objects
        .filter(assigned_pm__isnull=False, is_deleted=False)
        .select_related('assigned_pm__user')
        .order_by('-created_at')
    )
    pm_profiles = (
        UserProfile.objects
        .filter(role='PM', is_active=True)
        .select_related('user')
        .order_by('user__first_name')
    )

    return render(request, 'projects/subadmin/projects.html', {
        'unassigned':  unassigned,
        'assigned':    assigned,
        'pm_profiles': pm_profiles,
    })


@system_admin_required
def subadmin_departments(request):
    """System Admin departments view — excludes Admin and System Admin from every query."""
    from django.contrib.auth.models import User as AuthUser
    from itertools import groupby

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'create_user':
            return _subadmin_create_user(request)

        elif action == 'change_role':
            target_id = request.POST.get('user_id', '').strip()
            try:
                target_user = AuthUser.objects.select_related('profile').get(pk=target_id)
            except AuthUser.DoesNotExist:
                messages.error(request, 'User not found.')
                return redirect('subadmin_departments')

            # Defense-in-depth: never allow editing Admin or System Admin accounts
            if target_user.profile.role in _SA_EXCLUDED_ROLES:
                raise PermissionDenied

            new_role = request.POST.get('new_role', '').strip()
            valid_roles = [r[0] for r in _SA_EDITABLE_ROLE_CHOICES]

            # Defense-in-depth: cannot promote anyone to Admin or System Admin
            if new_role not in valid_roles:
                messages.error(request, 'Invalid role selection.')
                return redirect('subadmin_departments')

            old_role = target_user.profile.role
            if old_role != new_role:
                target_user.profile.role = new_role
                target_user.profile.save()
                log_activity(
                    project=None,
                    actor=request.user.profile,
                    action=f"User '{target_user.get_full_name() or target_user.username}' role changed from '{old_role}' to '{new_role}' by System Admin",
                    entity_type='User',
                    entity_id=target_user.id,
                )
                messages.success(request, f'Role updated for {target_user.get_full_name() or target_user.username}.')
            return redirect('subadmin_departments')

        elif action == 'edit_user':
            target_id = request.POST.get('user_id', '').strip()
            try:
                target_user = AuthUser.objects.select_related('profile').get(pk=target_id)
            except AuthUser.DoesNotExist:
                messages.error(request, 'User not found.')
                return redirect('subadmin_departments')

            # Defense-in-depth: never allow editing Admin or System Admin accounts
            if target_user.profile.role in _SA_EXCLUDED_ROLES:
                raise PermissionDenied

            new_role = request.POST.get('new_role', '').strip()
            valid_roles = [r[0] for r in _SA_EDITABLE_ROLE_CHOICES]

            # Defense-in-depth: cannot promote anyone to Admin or System Admin
            if new_role not in valid_roles:
                messages.error(request, 'Invalid role selection.')
                return redirect('subadmin_departments')

            old_role = target_user.profile.role
            target_user.first_name = request.POST.get('first_name', target_user.first_name).strip()
            target_user.last_name  = request.POST.get('last_name',  target_user.last_name).strip()
            target_user.save()

            target_user.profile.phone_number   = request.POST.get('phone_number', target_user.profile.phone_number).strip()
            target_user.profile.is_design_head = request.POST.get('is_design_head') == 'on'
            if old_role != new_role:
                target_user.profile.role = new_role
            target_user.profile.save()

            log_activity(
                project=None,
                actor=request.user.profile,
                action=(
                    f"User '{target_user.get_full_name() or target_user.username}' details edited by System Admin"
                    + (f" (role: {old_role} → {new_role})" if old_role != new_role else "")
                ),
                entity_type='User',
                entity_id=target_user.id,
            )
            messages.success(request, f'{target_user.get_full_name() or target_user.username} updated.')
            return redirect('subadmin_departments')

        elif action == 'deactivate':
            target_id = request.POST.get('user_id', '').strip()
            try:
                target_user = AuthUser.objects.select_related('profile').get(pk=target_id)
            except AuthUser.DoesNotExist:
                messages.error(request, 'User not found.')
                return redirect('subadmin_departments')

            if target_user.profile.role in _SA_EXCLUDED_ROLES:
                raise PermissionDenied

            target_user.is_active = False
            target_user.save()
            log_activity(
                project=None,
                actor=request.user.profile,
                action=f"User '{target_user.get_full_name() or target_user.username}' deactivated by System Admin",
                entity_type='User',
                entity_id=target_user.id,
            )
            messages.success(request, f'{target_user.get_full_name() or target_user.username} deactivated.')
            return redirect('subadmin_departments')

        elif action == 'reactivate':
            target_id = request.POST.get('user_id', '').strip()
            try:
                target_user = AuthUser.objects.select_related('profile').get(pk=target_id)
            except AuthUser.DoesNotExist:
                messages.error(request, 'User not found.')
                return redirect('subadmin_departments')

            if target_user.profile.role in _SA_EXCLUDED_ROLES:
                raise PermissionDenied

            target_user.is_active = True
            target_user.save()
            log_activity(
                project=None,
                actor=request.user.profile,
                action=f"User '{target_user.get_full_name() or target_user.username}' reactivated by System Admin",
                entity_type='User',
                entity_id=target_user.id,
            )
            messages.success(request, f'{target_user.get_full_name() or target_user.username} reactivated.')
            return redirect('subadmin_departments')

        messages.error(request, 'Unknown action.')
        return redirect('subadmin_departments')

    # GET — queryset always excludes Admin and System Admin at DB level
    all_profiles = (
        UserProfile.objects
        .select_related('user')
        .exclude(role__in=_SA_EXCLUDED_ROLES)
        .order_by('role', 'user__first_name', 'user__last_name')
    )

    grouped = []
    for role_key, members in groupby(all_profiles, key=lambda p: p.role):
        member_list  = list(members)
        active_count = sum(1 for m in member_list if m.user.is_active)
        grouped.append({
            'role':         role_key,
            'dept_name':    _SA_DEPT_NAMES.get(role_key, role_key),
            'members':      member_list,
            'active_count': active_count,
        })

    return render(request, 'projects/subadmin/departments.html', {
        'grouped':               grouped,
        'editable_role_choices': _SA_EDITABLE_ROLE_CHOICES,
        'current_user':          request.user,
    })


def _subadmin_create_user(request):
    """Helper called by subadmin_departments to create a new user. Never creates Admin or System Admin."""
    from django.contrib.auth.models import User as AuthUser

    first_name   = request.POST.get('first_name', '').strip()
    last_name    = request.POST.get('last_name', '').strip()
    username     = request.POST.get('username', '').strip()
    phone_number = request.POST.get('phone_number', '').strip()
    role         = request.POST.get('role', '').strip()
    password     = request.POST.get('password', '')

    # Defense-in-depth: cannot create Admin or System Admin
    if role in _SA_EXCLUDED_ROLES:
        messages.error(request, 'Invalid role selection.')
        return redirect('subadmin_departments')

    valid_roles = [r[0] for r in _SA_EDITABLE_ROLE_CHOICES]
    if role not in valid_roles:
        messages.error(request, 'Invalid role selection.')
        return redirect('subadmin_departments')

    if not username:
        messages.error(request, 'Username is required.')
        return redirect('subadmin_departments')

    if not password or len(password) < 8:
        messages.error(request, 'Password must be at least 8 characters.')
        return redirect('subadmin_departments')

    if AuthUser.objects.filter(username=username).exists():
        messages.error(request, f"Username '{username}' is already taken.")
        return redirect('subadmin_departments')

    new_user = AuthUser.objects.create_user(
        username=username,
        password=password,
        first_name=first_name,
        last_name=last_name,
        is_active=True,
    )

    # UserProfile is auto-created by post_save signal — just update the fields
    new_user.profile.role         = role
    new_user.profile.phone_number = phone_number
    new_user.profile.created_by   = request.user
    new_user.profile.save()

    log_activity(
        project=None,
        actor=request.user.profile,
        action=f"User '{new_user.get_full_name() or username}' ({role}) created by System Admin",
        entity_type='User',
        entity_id=new_user.id,
    )
    messages.success(request, f"User '{username}' created successfully as {role}.")
    return redirect('subadmin_departments')


@system_admin_required
def subadmin_task_durations(request):
    """System Admin view for task duration templates — same data as admin version, own template."""
    if request.method == 'POST':
        actor   = request.user.profile
        changed = 0
        errors  = []

        for key, raw_val in request.POST.items():
            if not key.startswith('duration_'):
                continue
            try:
                pk = int(key.split('_', 1)[1])
            except (ValueError, IndexError):
                continue

            raw_val = raw_val.strip()
            if not raw_val.isdigit():
                errors.append(f"Invalid value '{raw_val}' — must be a non-negative whole number.")
                continue
            new_days = int(raw_val)

            try:
                record = TaskDurationTemplate.objects.get(pk=pk)
            except TaskDurationTemplate.DoesNotExist:
                continue

            if new_days == record.duration_days:
                continue

            old_days = record.duration_days
            record.duration_days = new_days
            record.updated_by    = request.user
            record.save(update_fields=['duration_days', 'updated_by', 'updated_at'])
            changed += 1

            log_activity(
                project=None,
                actor=actor,
                action=(
                    f"Updated task duration: '{record.task_name}' ({record.phase_name}) "
                    f"changed from {old_days}d to {new_days}d [residential template]"
                ),
                entity_type='TaskDurationTemplate',
                entity_id=record.pk,
            )

        if errors:
            for msg in errors:
                messages.error(request, msg)
        else:
            messages.success(request, f'Saved. {changed} duration(s) updated.')

        return redirect('subadmin_task_durations')

    records = list(
        TaskDurationTemplate.objects
        .filter(project_type='residential')
        .select_related('updated_by')
        .order_by('phase_name', 'task_name')
    )

    phase_groups = {phase: [] for phase in _PHASE_ORDER}
    for rec in records:
        if rec.phase_name in phase_groups:
            phase_groups[rec.phase_name].append(rec)

    grouped = [(phase, phase_groups[phase]) for phase in _PHASE_ORDER if phase_groups[phase]]

    return render(request, 'projects/subadmin/task_durations.html', {
        'grouped': grouped,
        'total':   len(records),
    })
